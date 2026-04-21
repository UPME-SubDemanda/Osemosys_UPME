from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

import app.simulation.core.data_processing as data_processing_module
import app.services.scenario_service as scenario_service_module
from app.core.exceptions import ForbiddenError
from app.models import Emission, Fuel, OsemosysParamValue, Scenario, Technology
from app.services.csv_scenario_import_service import CsvScenarioImportService
from app.services.official_import_service import OfficialImportService
from app.services.scenario_service import ScenarioService

from factories import (
    create_osemosys_value,
    create_permission,
    create_region,
    create_scenario,
    create_user,
)


def _build_small_excel(*, region_name: str, param_name: str, year: int, value: float) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parameters"
    sheet.append(["Parameter", "Region", year])
    sheet.append([param_name, region_name, value])
    data = BytesIO()
    workbook.save(data)
    return data.getvalue()


def test_scenario_visibility_metadata_and_changed_params(db_session, monkeypatch) -> None:
    monkeypatch.setattr(scenario_service_module, "osemosys_table", lambda table_name: table_name)

    owner = create_user(db_session, username="owner")
    viewer = create_user(db_session, username="viewer")

    owner_only = create_scenario(
        db_session, name="Privado", owner=owner.username, edit_policy="OWNER_ONLY"
    )
    open_scenario = create_scenario(
        db_session, name="Abierto", owner=owner.username, edit_policy="OPEN"
    )
    restricted = create_scenario(
        db_session, name="Restringido", owner=owner.username, edit_policy="RESTRICTED"
    )
    create_permission(
        db_session,
        scenario_id=restricted.id,
        user=viewer,
        can_edit_direct=True,
        can_manage_values=True,
    )

    listed = ScenarioService.list(
        db_session,
        current_user=viewer,
        busqueda=None,
        owner=None,
        edit_policy=None,
        permission_scope=None,
        cantidad=50,
        offset=1,
    )

    listed_names = {row["name"] for row in listed["data"]}
    assert "Abierto" in listed_names
    assert "Restringido" in listed_names
    assert "Privado" not in listed_names

    restricted_public = ScenarioService.get_public(
        db_session, scenario_id=restricted.id, current_user=viewer
    )
    assert restricted_public["effective_access"]["can_view"] is True
    assert restricted_public["effective_access"]["can_edit_direct"] is True

    updated = ScenarioService.update_metadata(
        db_session,
        scenario_id=restricted.id,
        current_user=viewer,
        payload={
            "name": "Restringido editado",
            "description": "Nueva descripción",
            "edit_policy": "OPEN",
            "simulation_type": "REGIONAL",
        },
    )
    assert updated["name"] == "Restringido editado"
    assert updated["description"] == "Nueva descripción"
    assert updated["edit_policy"] == "OPEN"
    assert updated["simulation_type"] == "REGIONAL"

    parent = create_scenario(
        db_session, name="Padre", owner=owner.username, edit_policy="RESTRICTED"
    )
    create_permission(
        db_session,
        scenario_id=parent.id,
        user=viewer,
        can_edit_direct=True,
        can_manage_values=True,
    )
    region = create_region(db_session, name="Norte")
    base_row = create_osemosys_value(
        db_session,
        scenario_id=parent.id,
        param_name="CapitalCost",
        id_region=region.id,
        year=2025,
        value=10.0,
    )

    child = ScenarioService.clone(
        db_session,
        source_scenario_id=parent.id,
        current_user=viewer,
        name="Hijo",
        description="Copia",
        edit_policy="OWNER_ONLY",
    )
    assert child["base_scenario_id"] == parent.id
    assert child["base_scenario_name"] == parent.name
    assert child["changed_param_names"] == []

    cloned_row = (
        db_session.query(OsemosysParamValue)
        .filter(
            OsemosysParamValue.id_scenario == child["id"],
            OsemosysParamValue.param_name == "CapitalCost",
            OsemosysParamValue.id_region == region.id,
            OsemosysParamValue.year == 2025,
        )
        .one()
    )
    updated_child_row = ScenarioService.update_osemosys_value(
        db_session,
        scenario_id=child["id"],
        value_id=cloned_row.id,
        current_user=viewer,
        payload={
            "param_name": "CapitalCost",
            "region_name": region.name,
            "year": 2025,
            "value": 20.0,
        },
    )
    assert updated_child_row["value"] == 20.0

    created_child_row = ScenarioService.create_osemosys_value(
        db_session,
        scenario_id=parent.id,
        current_user=owner,
        payload={
            "param_name": "Demand",
            "region_name": region.name,
            "year": 2025,
            "value": 7.0,
        },
    )
    assert created_child_row["param_name"] == "Demand"

    created_lineage_row = ScenarioService.create_osemosys_value(
        db_session,
        scenario_id=child["id"],
        current_user=viewer,
        payload={
            "param_name": "ResidualCapacity",
            "region_name": region.name,
            "year": 2025,
            "value": 5.0,
        },
    )
    assert created_lineage_row["param_name"] == "ResidualCapacity"
    ScenarioService.deactivate_osemosys_value(
        db_session,
        scenario_id=child["id"],
        value_id=created_lineage_row["id"],
        current_user=viewer,
    )

    child_public = ScenarioService.get_public(
        db_session, scenario_id=child["id"], current_user=viewer
    )
    assert child_public["changed_param_names"] == ["CapitalCost", "ResidualCapacity"]
    assert base_row.id != cloned_row.id


def test_excel_preview_apply_and_update_change_values(db_session) -> None:
    owner = create_user(db_session, username="excel-owner")
    scenario = create_scenario(
        db_session, name="Excel pequeño", owner=owner.username, edit_policy="OWNER_ONLY"
    )
    region = create_region(db_session, name="Centro")
    seeded_row = create_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        param_name="CapitalCost",
        id_region=region.id,
        year=2025,
        value=11.0,
    )

    workbook_v1 = _build_small_excel(
        region_name=region.name,
        param_name="CapitalCost",
        year=2025,
        value=12.5,
    )
    preview = OfficialImportService.preview_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="small.xlsx",
        content=workbook_v1,
        selected_sheet_name="Parameters",
    )
    assert preview["total_rows_read"] == 1
    assert preview["not_found"] == 0
    assert len(preview["changes"]) == 1
    assert preview["changes"][0]["action"] == "update"
    assert preview["changes"][0]["row_id"] == seeded_row.id
    assert preview["changes"][0]["old_value"] == 11.0
    assert preview["changes"][0]["new_value"] == 12.5

    applied = OfficialImportService.apply_excel_changes(
        db_session,
        scenario_id=scenario.id,
        changes=preview["changes"],
    )
    assert applied["updated"] == 1
    assert applied["inserted"] == 0
    assert applied["skipped"] == 0
    db_session.refresh(seeded_row)
    assert seeded_row.value == 12.5

    workbook_v2 = _build_small_excel(
        region_name=region.name,
        param_name="CapitalCost",
        year=2025,
        value=14.0,
    )
    updated = OfficialImportService.update_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="small.xlsx",
        content=workbook_v2,
        selected_sheet_name="Parameters",
    )
    assert updated["updated"] == 1
    assert updated["not_found"] == 0
    db_session.refresh(seeded_row)
    assert seeded_row.value == 14.0


def test_root_scenario_tracks_changed_param_names(db_session) -> None:
    owner = create_user(db_session, username="root-track-owner")
    scenario = create_scenario(
        db_session,
        name="Raiz con tracking",
        owner=owner.username,
        edit_policy="OWNER_ONLY",
    )
    region = create_region(db_session, name="Region root")

    created_row = ScenarioService.create_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        current_user=owner,
        payload={
            "param_name": "ReserveMargin",
            "region_name": region.name,
            "year": 2025,
            "value": 1.0,
        },
    )
    assert created_row["param_name"] == "ReserveMargin"

    refreshed = ScenarioService.get_public(
        db_session,
        scenario_id=scenario.id,
        current_user=owner,
    )
    assert refreshed["changed_param_names"] == ["ReserveMargin"]


def _write_csv(root: Path, name: str, rows: list[str]) -> None:
    (root / name).write_text("\n".join(rows) + "\n", encoding="utf-8")


def test_import_csv_directory_creates_scenario_and_values(db_session, tmp_path: Path) -> None:
    owner = create_user(db_session, username="csv-owner")
    csv_root = tmp_path / "csv-model"
    csv_root.mkdir(parents=True)

    _write_csv(csv_root, "YEAR.csv", ["VALUE", "2025"])
    _write_csv(csv_root, "REGION.csv", ["VALUE", "R1"])
    _write_csv(csv_root, "TECHNOLOGY.csv", ["VALUE", "T1"])
    _write_csv(csv_root, "TIMESLICE.csv", ["VALUE", "TS1"])
    _write_csv(csv_root, "MODE_OF_OPERATION.csv", ["VALUE", "1"])
    _write_csv(csv_root, "FUEL.csv", ["VALUE", "F1"])
    _write_csv(csv_root, "SpecifiedAnnualDemand.csv", ["REGION,FUEL,YEAR,VALUE", "R1,F1,2025,100"])
    _write_csv(
        csv_root,
        "OutputActivityRatio.csv",
        ["REGION,TECHNOLOGY,FUEL,MODE_OF_OPERATION,YEAR,VALUE", "R1,T1,F1,1,2025,1"],
    )
    _write_csv(csv_root, "CapitalCost.csv", ["REGION,TECHNOLOGY,YEAR,VALUE", "R1,T1,2025,55"])

    created = CsvScenarioImportService.import_from_directory(
        db_session,
        current_user=owner,
        csv_root=csv_root,
        scenario_name="Escenario CSV regional",
        description="Importado desde ZIP",
        edit_policy="OWNER_ONLY",
        tag_id=None,
        simulation_type="REGIONAL",
    )

    assert created["simulation_type"] == "REGIONAL"
    assert created["name"] == "Escenario CSV regional"
    rows = (
        db_session.query(OsemosysParamValue)
        .filter(OsemosysParamValue.id_scenario == created["id"])
        .all()
    )
    created_scenario = db_session.get(Scenario, created["id"])
    assert created_scenario is not None
    assert created_scenario.processing_mode == "PREPROCESSED_CSV"
    assert len(rows) == 3
    assert {row.param_name for row in rows} == {
        "SpecifiedAnnualDemand",
        "OutputActivityRatio",
        "CapitalCost",
    }


def test_import_csv_directory_reuses_existing_catalogs_without_duplicate_violation(
    db_session,
    tmp_path: Path,
) -> None:
    owner = create_user(db_session, username="csv-existing-owner")
    existing_technology = Technology(name="EXPOIL_1LIV", is_active=True)
    db_session.add(existing_technology)
    db_session.commit()

    csv_root = tmp_path / "csv-existing-model"
    csv_root.mkdir(parents=True)

    _write_csv(csv_root, "YEAR.csv", ["VALUE", "2025"])
    _write_csv(csv_root, "REGION.csv", ["VALUE", "R1"])
    _write_csv(csv_root, "TECHNOLOGY.csv", ["VALUE", "EXPOIL_1LIV"])
    _write_csv(csv_root, "TIMESLICE.csv", ["VALUE", "TS1"])
    _write_csv(csv_root, "MODE_OF_OPERATION.csv", ["VALUE", "1"])
    _write_csv(csv_root, "FUEL.csv", ["VALUE", "F1"])
    _write_csv(csv_root, "SpecifiedAnnualDemand.csv", ["REGION,FUEL,YEAR,VALUE", "R1,F1,2025,100"])
    _write_csv(
        csv_root,
        "OutputActivityRatio.csv",
        ["REGION,TECHNOLOGY,FUEL,MODE_OF_OPERATION,YEAR,VALUE", "R1,EXPOIL_1LIV,F1,1,2025,1"],
    )

    created = CsvScenarioImportService.import_from_directory(
        db_session,
        current_user=owner,
        csv_root=csv_root,
        scenario_name="Escenario con catálogos existentes",
        description=None,
        edit_policy="OWNER_ONLY",
        tag_id=None,
        simulation_type="NATIONAL",
    )

    assert created["name"] == "Escenario con catálogos existentes"
    technologies = db_session.query(Technology).filter(Technology.name == "EXPOIL_1LIV").all()
    assert len(technologies) == 1


def test_run_data_processing_skips_postprocessing_for_preprocessed_csv_scenario(
    db_session,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    owner = create_user(db_session, username="preprocessed-owner")
    scenario = create_scenario(
        db_session,
        name="Escenario preprocesado",
        owner=owner.username,
        edit_policy="OWNER_ONLY",
        processing_mode="PREPROCESSED_CSV",
    )

    export_result = data_processing_module.ProcessingResult(
        has_storage=False,
        has_udc=False,
        sets={"YEAR": [2025], "REGION": ["R1"]},
        param_count=3,
    )
    called_steps: list[str] = []

    def fake_export(db, *, scenario_id: int, csv_dir: str):
        assert scenario_id == scenario.id
        return export_result

    def _mark(name: str):
        def inner(*args, **kwargs):
            called_steps.append(name)
        return inner

    monkeypatch.setattr(data_processing_module, "export_scenario_to_csv", fake_export)
    monkeypatch.setattr(data_processing_module, "normalize_mode_of_operation_in_csv_dir", _mark("normalize"))
    monkeypatch.setattr(data_processing_module, "eliminar_valores_fuera_de_indices", _mark("filter"))
    monkeypatch.setattr(data_processing_module, "completar_Matrix_Act_Ratio", _mark("act_ratio"))
    monkeypatch.setattr(data_processing_module, "completar_Matrix_Emission", _mark("emission"))
    monkeypatch.setattr(data_processing_module, "completar_Matrix_Storage", _mark("storage"))
    monkeypatch.setattr(data_processing_module, "completar_Matrix_Cost", _mark("cost"))
    monkeypatch.setattr(data_processing_module, "process_and_save_emission_ratios", _mark("emission_ratios"))
    monkeypatch.setattr(data_processing_module, "ensure_udc_csvs", _mark("udc"))
    monkeypatch.setattr(data_processing_module, "apply_udc_config", _mark("apply_udc"))
    monkeypatch.setattr(data_processing_module, "reorder_activity_ratio_csvs_for_dataportal", _mark("reorder"))

    result = data_processing_module.run_data_processing(
        db_session,
        scenario_id=scenario.id,
        csv_dir=str(tmp_path),
    )

    assert result is export_result
    assert called_steps == []


def test_excel_preview_detects_inserts_and_apply_creates_missing_catalogs(db_session) -> None:
    owner = create_user(db_session, username="excel-insert-owner")
    scenario = create_scenario(
        db_session, name="Excel insert", owner=owner.username, edit_policy="OWNER_ONLY"
    )
    region = create_region(db_session, name="Sur")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parameters"
    sheet.append(["Parameter", "Region", "Technology", "Fuel", "Emission", 2026])
    sheet.append(["VariableCost", region.name, "TECH_NEW_TEST", "FUEL_NEW_TEST", "EMI_NEW_TEST", 9.75])
    data = BytesIO()
    workbook.save(data)

    preview = OfficialImportService.preview_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="insert.xlsx",
        content=data.getvalue(),
        selected_sheet_name="Parameters",
    )
    assert preview["total_rows_read"] == 1
    assert len(preview["changes"]) == 1
    row = preview["changes"][0]
    assert row["action"] == "insert"
    assert row["row_id"] is None
    assert row["old_value"] is None
    assert row["new_value"] == 9.75

    applied = OfficialImportService.apply_excel_changes(
        db_session,
        scenario_id=scenario.id,
        changes=preview["changes"],
    )
    assert applied["updated"] == 0
    assert applied["inserted"] == 1
    assert applied["skipped"] == 0

    inserted_row = (
        db_session.query(OsemosysParamValue)
        .filter(
            OsemosysParamValue.id_scenario == scenario.id,
            OsemosysParamValue.param_name == "VariableCost",
            OsemosysParamValue.year == 2026,
        )
        .one_or_none()
    )
    assert inserted_row is not None
    assert float(inserted_row.value) == 9.75
    assert db_session.query(Technology).filter(Technology.name == "TECH_NEW_TEST").one_or_none() is not None
    assert db_session.query(Fuel).filter(Fuel.name == "FUEL_NEW_TEST").one_or_none() is not None
    assert db_session.query(Emission).filter(Emission.name == "EMI_NEW_TEST").one_or_none() is not None


def test_excel_new_rows_use_sand_processing_defaults_for_preview_and_apply(db_session) -> None:
    owner = create_user(db_session, username="excel-default-owner")
    scenario = create_scenario(
        db_session, name="Excel defaults", owner=owner.username, edit_policy="OWNER_ONLY"
    )
    region = create_region(db_session, name="Oriente")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parameters"
    sheet.append(["Parameter", "Region", "Technology", "Fuel", "Emission", 2026])
    # Celda vacía en VariableCost debe resolverse con default SAND (0.000001),
    # por lo tanto debe detectarse como insert real.
    sheet.append(["VariableCost", region.name, "TECH_DEF_TEST", "FUEL_DEF_TEST", "EMI_DEF_TEST", None])
    data = BytesIO()
    workbook.save(data)

    preview = OfficialImportService.preview_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="defaults.xlsx",
        content=data.getvalue(),
        selected_sheet_name="Parameters",
    )
    assert preview["total_rows_read"] == 1
    assert len(preview["changes"]) == 1
    row = preview["changes"][0]
    assert row["action"] == "insert"
    assert abs(float(row["new_value"]) - 0.000001) < 1e-12

    applied = OfficialImportService.apply_excel_changes(
        db_session,
        scenario_id=scenario.id,
        changes=preview["changes"],
    )
    assert applied["updated"] == 0
    assert applied["inserted"] == 1
    assert applied["skipped"] == 0

    inserted_row = (
        db_session.query(OsemosysParamValue)
        .filter(
            OsemosysParamValue.id_scenario == scenario.id,
            OsemosysParamValue.param_name == "VariableCost",
            OsemosysParamValue.year == 2026,
        )
        .one_or_none()
    )
    assert inserted_row is not None
    assert abs(float(inserted_row.value) - 0.000001) < 1e-12


def test_excel_new_rows_with_timeslice_all_zero_are_skipped_by_sand_aggregation(db_session) -> None:
    owner = create_user(db_session, username="excel-timeslice-owner")
    scenario = create_scenario(
        db_session, name="Excel timeslice", owner=owner.username, edit_policy="OWNER_ONLY"
    )
    region = create_region(db_session, name="CentroOriente")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parameters"
    sheet.append(["Parameter", "Region", "Technology", "Fuel", "Emission", "TIMESLICE", 2026])
    sheet.append(["FixedCost", region.name, "TECH_TS_TEST", "FUEL_TS_TEST", "EMI_TS_TEST", "TS1", 0])
    sheet.append(["FixedCost", region.name, "TECH_TS_TEST", "FUEL_TS_TEST", "EMI_TS_TEST", "TS2", 0])
    data = BytesIO()
    workbook.save(data)

    preview = OfficialImportService.preview_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="timeslice_zero.xlsx",
        content=data.getvalue(),
        selected_sheet_name="Parameters",
    )
    assert preview["total_rows_read"] == 2
    assert preview["changes"] == []

    applied = OfficialImportService.apply_excel_changes(
        db_session,
        scenario_id=scenario.id,
        changes=preview["changes"],
    )
    assert applied["updated"] == 0
    assert applied["inserted"] == 0
    assert applied["skipped"] == 0


def test_open_scenario_allows_authenticated_non_owner_to_manage_values(db_session) -> None:
    owner = create_user(db_session, username="open-owner")
    outsider = create_user(db_session, username="open-outsider")
    scenario = create_scenario(
        db_session, name="Escenario abierto", owner=owner.username, edit_policy="OPEN"
    )
    region = create_region(db_session, name="Occidente")
    seeded_row = create_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        param_name="CapitalCost",
        id_region=region.id,
        year=2025,
        value=10.0,
    )

    scenario_public = ScenarioService.get_public(
        db_session,
        scenario_id=scenario.id,
        current_user=outsider,
    )
    assert scenario_public["effective_access"]["can_manage_values"] is True
    assert scenario_public["effective_access"]["can_edit_direct"] is False

    created_row = ScenarioService.create_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        current_user=outsider,
        payload={
            "param_name": "Demand",
            "region_name": region.name,
            "year": 2025,
            "value": 4.0,
        },
    )
    assert created_row["param_name"] == "Demand"

    updated_row = ScenarioService.update_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        value_id=seeded_row.id,
        current_user=outsider,
        payload={
            "param_name": "CapitalCost",
            "region_name": region.name,
            "year": 2025,
            "value": 12.0,
        },
    )
    assert updated_row["value"] == 12.0

    ScenarioService.deactivate_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        value_id=created_row["id"],
        current_user=outsider,
    )
    assert db_session.get(OsemosysParamValue, created_row["id"]) is None

    with pytest.raises(ForbiddenError):
        ScenarioService.update_metadata(
            db_session,
            scenario_id=scenario.id,
            current_user=outsider,
            payload={"description": "No debería poder editar metadatos"},
        )


def test_open_scenario_excel_updates_do_not_require_explicit_permission(db_session) -> None:
    owner = create_user(db_session, username="open-excel-owner")
    outsider = create_user(db_session, username="open-excel-outsider")
    scenario = create_scenario(
        db_session, name="Escenario abierto excel", owner=owner.username, edit_policy="OPEN"
    )
    region = create_region(db_session, name="Caribe")
    seeded_row = create_osemosys_value(
        db_session,
        scenario_id=scenario.id,
        param_name="CapitalCost",
        id_region=region.id,
        year=2025,
        value=11.0,
    )

    workbook_v1 = _build_small_excel(
        region_name=region.name,
        param_name="CapitalCost",
        year=2025,
        value=13.5,
    )
    preview = OfficialImportService.preview_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="open-small.xlsx",
        content=workbook_v1,
        selected_sheet_name="Parameters",
    )
    assert preview["changes"][0]["action"] == "update"
    assert preview["changes"][0]["row_id"] == seeded_row.id

    ScenarioService._require_manage_values(
        db_session,
        scenario_id=scenario.id,
        current_user=outsider,
    )

    applied = OfficialImportService.apply_excel_changes(
        db_session,
        scenario_id=scenario.id,
        changes=preview["changes"],
    )
    assert applied["updated"] == 1
    assert applied["inserted"] == 0
    assert applied["skipped"] == 0
    db_session.refresh(seeded_row)
    assert seeded_row.value == 13.5

    workbook_v2 = _build_small_excel(
        region_name=region.name,
        param_name="CapitalCost",
        year=2025,
        value=15.0,
    )
    updated = OfficialImportService.update_scenario_from_excel(
        db_session,
        scenario_id=scenario.id,
        filename="open-small.xlsx",
        content=workbook_v2,
        selected_sheet_name="Parameters",
    )
    assert updated["updated"] == 1
    db_session.refresh(seeded_row)
    assert seeded_row.value == 15.0
