"""Chart service — lógica central de visualización.

Reemplaza ``graficas.py`` y ``graficas_comparacion.py`` del paquete
``osemosys_src`` para funcionar directamente contra la BD.  Cada función
pública devuelve un Pydantic schema listo para serialización en FastAPI.

Funciones públicas:
  - ``build_chart_data``       → gráfica single-escenario
  - ``build_comparison_data``  → gráfica multi-escenario / subplots por año
  - ``get_result_summary``     → KPIs de cabecera
  - ``get_chart_catalog``      → catálogo de tipos de gráfica disponibles
"""

from __future__ import annotations

import logging
import os
from typing import Any

import pandas as pd
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.models import OsemosysOutputParamValue, SimulationJob


# ──────────────────────────────────────────────────────────────────────────
#  Tipografía: registramos Nunito (bundled en `fonts/`) y la dejamos como
#  font.family default para todos los renderers matplotlib del módulo.
# ──────────────────────────────────────────────────────────────────────────
def _register_nunito_font() -> None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        from matplotlib import font_manager

        fonts_dir = os.path.join(os.path.dirname(__file__), "fonts")
        if not os.path.isdir(fonts_dir):
            return
        for fname in os.listdir(fonts_dir):
            if fname.lower().endswith((".ttf", ".otf")):
                font_manager.fontManager.addfont(os.path.join(fonts_dir, fname))
        # Si Nunito quedó disponible, lo activamos como default.
        names = {f.name for f in font_manager.fontManager.ttflist}
        if "Nunito" in names:
            matplotlib.rcParams["font.family"] = "Nunito"
            matplotlib.rcParams["font.sans-serif"] = [
                "Nunito",
                "DejaVu Sans",
                "sans-serif",
            ]
    except Exception:  # pragma: no cover — no romper el módulo si falla
        pass


_register_nunito_font()
from app.schemas.scenario import ScenarioTagPublic
from app.schemas.visualization import (
    ChartCatalogItem,
    ChartDataResponse,
    ChartSeries,
    CompareChartFacetResponse,
    CompareChartResponse,
    FacetData,
    ParetoChartResponse,
    ResultSummaryResponse,
    SubplotData,
)
from app.visualization.colors import (
    asignar_grupo,
    generar_colores_tecnologias,
    _color_electricidad,
    _color_por_grupo_fijo,
    _color_por_sector,
    _color_por_emision,
)
from app.visualization.labels import get_label
from app.visualization.configs import (
    CONFIGS,
    TITULOS_VARIABLES_CAPACIDAD,
    NOMBRES_COMBUSTIBLES,
    _map_h2_verde_azul_gris,
)
from app.visualization.configs_comparacion import CONFIGS_COMPARACION
from app.visualization.catalog_reader import (
    get_colores_emisiones,
    get_colores_grupos,
    get_colores_sector,
    get_mapa_sector,
)

logger = logging.getLogger(__name__)

# Variables principales (columnas tipadas en la BD)
_MAIN_TYPED_VARIABLES = {"Dispatch", "NewCapacity", "UnmetDemand", "AnnualEmissions"}


# ═══════════════════════════════════════════════════════════════════════════
# 1. DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════


def _load_variable_data(
    db: Session,
    job_id: int,
    variable_name: str,
) -> pd.DataFrame:
    """Carga datos de ``osemosys_output_param_value`` y devuelve un DataFrame.

    • Variables principales (Dispatch, NewCapacity, …): usa columnas tipadas
      (``technology_name``, ``fuel_name``, ``year``, ``value``).
    • Variables intermedias (ProductionByTechnology, TotalCapacityAnnual, …):
      extrae TECHNOLOGY, FUEL, YEAR del campo ``index_json``.

    Parámetros
    ----------
    db : Session
        Sesión SQLAlchemy activa.
    job_id : int
        ID del simulation_job.
    variable_name : str
        Nombre exacto de la variable a cargar.

    Retorna
    -------
    pd.DataFrame
        Columnas garantizadas: TECHNOLOGY, YEAR, VALUE.
        Columna opcional: FUEL (cuando está disponible).
    """

    # ── Consulta BD ──────────────────────────────────────────────────────
    rows = (
        db.query(OsemosysOutputParamValue)
        .filter(
            OsemosysOutputParamValue.id_simulation_job == job_id,
            OsemosysOutputParamValue.variable_name == variable_name,
        )
        .all()
    )

    if not rows:
        return pd.DataFrame(columns=["TECHNOLOGY", "FUEL", "YEAR", "VALUE"])

    # ── Construir DataFrame ──────────────────────────────────────────────
    if variable_name in _MAIN_TYPED_VARIABLES:
        records = []
        for r in rows:
            records.append(
                {
                    "TECHNOLOGY": r.technology_name or "",
                    "FUEL": r.fuel_name or "",
                    "YEAR": r.year,
                    "VALUE": float(r.value),
                }
            )
        df = pd.DataFrame(records)

    else:
        # Variable intermedia → extraer de index_json
        records = []
        for r in rows:
            idx_raw = r.index_json if r.index_json else []
            idx = idx_raw if isinstance(idx_raw, (list, tuple)) else []
            # Convenciones del pipeline:
            #   ProductionByTechnology / UseByTechnology / TotalCapacityAnnual /
            #   AccumulatedNewCapacity / AnnualTechnologyEmission:
            #     index_json = [REGION, TECHNOLOGY, FUEL?, YEAR?, ...]
            # La posición del YEAR puede variar: posición 2, 3 o 4.
            technology = str(idx[1]) if len(idx) > 1 else ""
            fuel = ""
            year = None

            if len(idx) >= 5:
                # [REGION, TECH, FUEL, ?, YEAR]  (5-element index)
                fuel = str(idx[2]) if idx[2] is not None else ""
                year = _safe_int(idx[4]) or _safe_int(idx[3])
            elif len(idx) >= 4:
                # [REGION, TECH, FUEL, YEAR]  (4-element index)
                fuel = str(idx[2]) if idx[2] is not None else ""
                year = _safe_int(idx[3])
            elif len(idx) >= 3:
                # [REGION, TECH, YEAR]  (3-element index)
                year = _safe_int(idx[2])

            records.append(
                {
                    "TECHNOLOGY": technology,
                    "FUEL": fuel,
                    "YEAR": year,
                    "VALUE": float(r.value),
                }
            )
        df = pd.DataFrame(records)

    # Limpiar: descartar filas sin YEAR útil
    df = df.dropna(subset=["YEAR"])
    df["YEAR"] = df["YEAR"].astype(int)

    return df


def _safe_int(val: Any) -> int | None:
    """Intenta convertir un valor a int, retorna None si falla."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def format_axis_3sig(v: Any) -> str:
    """Formatea un valor numérico con **mínimo 3 cifras significativas**.

    Uso típico: ``ax.yaxis.set_major_formatter(FuncFormatter(format_axis_3sig))``.

    Reglas:
      * ``|v| >= 100``  → entero con separador de miles ("1,234")
      * ``10 <= |v| < 100`` → 1 decimal ("12.3")
      * ``1 <= |v| < 10``  → 2 decimales ("1.23")
      * ``|v| < 1`` → tantos decimales como sean necesarios para 3 cifras
        significativas ("0.123", "0.0123", "0.00123", …)
      * cero → "0"
    """
    import math

    if v is None:
        return "0"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if not math.isfinite(v):
        return str(v)
    if v == 0:
        return "0"
    abs_v = abs(v)
    if abs_v >= 100:
        decimals = 0
    elif abs_v >= 10:
        decimals = 1
    elif abs_v >= 1:
        decimals = 2
    else:
        order = math.floor(math.log10(abs_v))  # negativo
        decimals = min(8, -order + 2)
    return f"{v:,.{decimals}f}"


def _year_keep_indices(
    categories: list[Any],
    year_from: int | None,
    year_to: int | None,
) -> list[int]:
    """Índices de ``categories`` cuyo valor (parseable como año) cae en el rango.

    Categorías no parseables como entero se preservan (no son años).
    """
    if year_from is None and year_to is None:
        return list(range(len(categories)))
    keep: list[int] = []
    for i, c in enumerate(categories):
        try:
            y = int(str(c))
        except (TypeError, ValueError):
            keep.append(i)
            continue
        if year_from is not None and y < year_from:
            continue
        if year_to is not None and y > year_to:
            continue
        keep.append(i)
    return keep


def reorder_chart_series(chart: Any, order: list[str] | None) -> None:
    """Reordena ``chart.series`` in-place según ``order`` (lista de nombres).

    Series no listadas se mantienen al final en su orden natural. Series
    listadas pero no presentes en el chart se ignoran silenciosamente.

    El primer nombre del array queda arriba del stack — convención del
    proyecto (Highcharts ``yAxis.reversedStacks=true`` por defecto, y los
    renderers matplotlib iteran en ``reversed()``).
    """
    if not order:
        return
    series = getattr(chart, "series", None)
    if not series:
        return
    by_name = {s.name: s for s in series}
    used: set[str] = set()
    new_series: list[Any] = []
    for name in order:
        s = by_name.get(name)
        if s is not None and name not in used:
            new_series.append(s)
            used.add(name)
    for s in series:
        if s.name not in used:
            new_series.append(s)
    chart.series = new_series


def apply_period_years(chart: Any, period: int | None) -> None:
    """Filtra ``categories`` (años) tomando uno cada ``period``, in-place.

    El primer año visible es el primer índice donde el offset (idx desde el
    primer año-categoría) es múltiplo de ``period``. El último año siempre
    se preserva (aunque rompa la cadencia) para que la tabla cierre en el
    horizonte real del modelo.

    Categorías no parseables como año se preservan tal cual.
    Si ``period`` es ``None`` o ``< 2``, no se hace nada.
    """
    if period is None or period < 2:
        return
    cats = getattr(chart, "categories", None)
    series = getattr(chart, "series", None)
    if not cats or series is None:
        return
    # Encontrar índices de años parseables.
    year_indices: list[int] = []
    year_values: list[int] = []
    for i, c in enumerate(cats):
        try:
            y = int(str(c))
            year_indices.append(i)
            year_values.append(y)
        except (TypeError, ValueError):
            continue
    if not year_indices:
        return
    base = year_values[0]
    keep = set()
    # No-año (otros): siempre conservar.
    for i in range(len(cats)):
        if i not in year_indices:
            keep.add(i)
    # Cada ``period`` años desde el primer año visible.
    for idx, year in zip(year_indices, year_values):
        if (year - base) % period == 0:
            keep.add(idx)
    # Garantizar que el último año esté presente.
    keep.add(year_indices[-1])
    keep_sorted = sorted(keep)
    chart.categories = [cats[i] for i in keep_sorted]
    for s in series:
        data = getattr(s, "data", None)
        if data is None:
            continue
        s.data = [data[i] for i in keep_sorted if i < len(data)]


def apply_cumulative_series(chart: Any) -> None:
    """Reemplaza cada ``series.data`` por su suma acumulada, in-place.

    Útil para tablas de "capacidad acumulada", "emisiones acumuladas", etc.
    NaN/None se trata como 0 para no destruir el acumulador.
    """
    import math

    series = getattr(chart, "series", None)
    if not series:
        return
    for s in series:
        data = getattr(s, "data", None)
        if data is None:
            continue
        cum: list[float] = []
        running = 0.0
        for v in data:
            try:
                f = float(v)
                if not math.isfinite(f):
                    f = 0.0
            except (TypeError, ValueError):
                f = 0.0
            running += f
            cum.append(running)
        s.data = cum


def filter_chart_by_year_range(
    chart: Any,
    year_from: int | None,
    year_to: int | None,
) -> None:
    """Filtra in-place ``ChartDataResponse`` o ``CompareChartFacetResponse``.

    - ``ChartDataResponse``: corta ``categories`` y cada ``series.data``.
    - ``CompareChartFacetResponse``: aplica el corte por cada faceta.

    Si ambos extremos son ``None``, no hace nada. Categorías no-año se
    preservan tal cual.
    """
    if year_from is None and year_to is None:
        return
    # Single chart / line-total (categories + series)
    cats = getattr(chart, "categories", None)
    series = getattr(chart, "series", None)
    if cats is not None and series is not None:
        keep = _year_keep_indices(cats, year_from, year_to)
        chart.categories = [cats[i] for i in keep]
        for s in series:
            data = getattr(s, "data", None)
            if data is None:
                continue
            s.data = [data[i] for i in keep if i < len(data)]
    # Facet (per-facet categories + series)
    facets = getattr(chart, "facets", None)
    if facets is not None:
        for f in facets:
            f_cats = getattr(f, "categories", None)
            f_series = getattr(f, "series", None)
            if f_cats is None or f_series is None:
                continue
            keep = _year_keep_indices(f_cats, year_from, year_to)
            f.categories = [f_cats[i] for i in keep]
            for s in f_series:
                data = getattr(s, "data", None)
                if data is None:
                    continue
                s.data = [data[i] for i in keep if i < len(data)]


# ═══════════════════════════════════════════════════════════════════════════
# 2. HELPERS DE TRANSFORMACIÓN (ports de graficas_comparacion.py)
# ═══════════════════════════════════════════════════════════════════════════


def _fuel_to_group(row) -> str:
    """Normaliza una fila (FUEL + TECHNOLOGY) al código base de grupo.

    Códigos FUEL con sufijos numéricos (ELC002, NGS002…) se mapean a su
    clave base (ELC, NGS…) usando asignar_grupo, para que coincidan con
    COLORES_GRUPOS y DISPLAY_NAMES.  OIL se desambigua por tecnología.
    """
    fuel = row.get("FUEL", "")
    tech = str(row.get("TECHNOLOGY", ""))
    if fuel == "OIL":
        if "MINOIL_3PES" in tech:
            return "MINOIL_3PES"
        if "MINOIL_2MID" in tech:
            return "MINOIL_2MID"
        if "MINOIL_1LIV" in tech:
            return "MINOIL_1LIV"
        if "MINOIL" in tech:
            return "MINOIL"
    return asignar_grupo(fuel) if fuel else "OTRO"


def _filtrar_df(
    df: pd.DataFrame,
    prefijo: str | tuple[str, ...],
    sub_filtro: str | None,
    loc: str | None,
) -> pd.DataFrame:
    """Aplica filtro por prefijo de TECHNOLOGY, sub_filtro y localización.

    Port directo de ``graficas_comparacion._filtrar_df``.
    """
    if df.empty:
        return df

    df = df[df["TECHNOLOGY"].str.startswith(prefijo)].copy()

    if sub_filtro:
        df = df[df["TECHNOLOGY"].str.contains(sub_filtro)]

    if loc == "URB":
        df = df[~df["TECHNOLOGY"].str.contains("RUR")]
        df = df[~df["TECHNOLOGY"].str.contains("ZNI")]
    elif loc == "RUR":
        df = df[df["TECHNOLOGY"].str.contains("RUR")]
        df = df[~df["TECHNOLOGY"].str.contains("ZNI")]
    elif loc == "ZNI":
        df = df[~df["TECHNOLOGY"].str.contains("RUR")]
        df = df[df["TECHNOLOGY"].str.contains("ZNI")]

    return df


def _sector_labels(tech_series: pd.Series) -> pd.Series:
    """Asignación vectorizada de sector, incluyendo PWR → Generación Electricidad.

    Usa ``MAPA_SECTOR`` desde BD (sembrado al arrancar el API; fallback al
    dict hardcoded sólo si la BD aún no está poblada).
    """
    labels = tech_series.str[:6].map(get_mapa_sector())
    pwr_mask = labels.isna() & tech_series.str.startswith("PWR")
    labels = labels.where(~pwr_mask, "Generación Electricidad")
    return labels.fillna("Otros")


def _asignar_categoria(
    df: pd.DataFrame,
    agrupacion: str,
) -> pd.DataFrame:
    """Crea columna CATEGORIA según el tipo de agrupación.

    Port de ``graficas_comparacion._asignar_categoria``.
    """
    df = df.copy()

    if agrupacion == "TECNOLOGIA":
        df["CATEGORIA"] = df["TECHNOLOGY"]

    elif agrupacion == "COMBUSTIBLE":
        if "FUEL" in df.columns:
            df["_TECH_FUEL"] = (
                df["TECHNOLOGY"].astype(str) + "_" + df["FUEL"].astype(str)
            )
        else:
            df["_TECH_FUEL"] = df["TECHNOLOGY"].astype(str)

        df["CATEGORIA"] = df["_TECH_FUEL"].apply(asignar_grupo)
        df = df.drop(columns="_TECH_FUEL")

    elif agrupacion == "SECTOR":
        df["CATEGORIA"] = _sector_labels(df["TECHNOLOGY"])

    elif agrupacion == "EMISION":
        # Para AnnualTechnologyEmission, FUEL contiene el tipo de emisión
        df["CATEGORIA"] = df["FUEL"] if "FUEL" in df.columns else "?"

    return df


def _convertir_unidades(df: pd.DataFrame, un: str) -> pd.DataFrame:
    """Convierte columna VALUE a las unidades solicitadas.

    Port de ``graficas_comparacion._convertir_unidades``.
    """
    df = df.copy()
    if un == "GW":
        df["VALUE"] /= 31.536
    elif un == "MW":
        df["VALUE"] /= 0.031536
    elif un == "TWh":
        df["VALUE"] /= 3.6
    elif un == "Gpc":
        df["VALUE"] /= 1.0095581216
    # PJ: unidad base, sin conversión
    return df


def _convertir_unidades_emision(df: pd.DataFrame, un: str) -> pd.DataFrame:
    """Convierte emisiones GEI entre MtCO₂eq y ktCO₂eq.

    Base de datos: MtCO₂eq. Multiplicar × 1000 para obtener ktCO₂eq.
    """
    df = df.copy()
    if un == "ktCO2eq":
        df["VALUE"] *= 1000.0
    return df


def _emision_unit_label(un: str, es_emision_kt: bool) -> str:
    """Devuelve la etiqueta de unidad correcta para gráficas de emisión."""
    if es_emision_kt:
        return "kt"
    return "ktCO₂eq" if un == "ktCO2eq" else "MtCO₂eq"


def _color_map_comparison(
    agrupacion: str,
    categorias_unicas: list[str],
) -> dict[str, str]:
    """Devuelve ``{categoria: color_hex}`` para gráficas de comparación.

    Port de ``graficas_comparacion._color_map``.
    """
    if agrupacion == "COMBUSTIBLE":
        palette = get_colores_grupos()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    if agrupacion == "SECTOR":
        palette = get_colores_sector()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    if agrupacion == "EMISION":
        palette = get_colores_emisiones()
        return {c: palette.get(c, "#999999") for c in categorias_unicas}

    # TECNOLOGIA: reutiliza generar_colores_tecnologias de colors.py
    df_tmp = pd.DataFrame({"COLOR": list(categorias_unicas)})
    colores_lista, orden_lista = generar_colores_tecnologias(df_tmp, "COLOR")
    return dict(zip(orden_lista, colores_lista))


def _build_factor_planta_data(
    db: Session,
    job_id: int,
    cfg: dict,
    title: str,
    sub_filtro: str | None,
    loc: str | None,
) -> ChartDataResponse:
    """CF = Producción[PJ] / TotalCapacityAnnual[PJ] × 100 %

    Ambas variables están en PJ (baseline del modelo).
    TotalCapacityAnnual[PJ] = capacidad[GW] × 31.536 = energía máxima anual posible.
    CF = Energía real / Energía máxima posible ∈ [0 %, 100 %].
    """
    filtro_fn = cfg.get("filtro")

    df_cap = _load_variable_data(db, job_id, "TotalCapacityAnnual")
    df_prd = _load_variable_data(db, job_id, "ProductionByTechnology")

    if filtro_fn is not None:
        df_cap = filtro_fn(df_cap, sub_filtro=sub_filtro, loc=loc)
        df_prd = filtro_fn(df_prd, sub_filtro=sub_filtro, loc=loc)

    if df_cap.empty or df_prd.empty:
        return ChartDataResponse(categories=[], series=[], title=title, yAxisLabel="%")

    cap_agg = df_cap.groupby(["TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    prd_agg = df_prd.groupby(["TECHNOLOGY", "YEAR"], as_index=False)["VALUE"].sum()
    prd_agg = prd_agg.rename(columns={"VALUE": "PRODUCTION"})

    df = cap_agg.merge(prd_agg, on=["TECHNOLOGY", "YEAR"], how="inner")
    df = df[df["VALUE"] > 1e-6].copy()

    if df.empty:
        return ChartDataResponse(categories=[], series=[], title=title, yAxisLabel="%")

    # Ambos en PJ → ratio directo, sin conversión adicional
    df["CF"] = (df["PRODUCTION"] / df["VALUE"] * 100.0).clip(0, 100)
    df["COLOR"] = df["TECHNOLOGY"]

    color_fn = cfg.get("color_fn")
    if color_fn is not None:
        colores_ordenados, orden_color = color_fn(df, "COLOR")
    else:
        orden_color = sorted(df["COLOR"].unique())
        colores_ordenados = ["#999999"] * len(orden_color)

    color_dict = dict(zip(orden_color, colores_ordenados))
    años = sorted(df["YEAR"].unique())
    categories = [str(a) for a in años]

    series: list[ChartSeries] = []
    for tech in orden_color:
        df_tech = df[df["COLOR"] == tech]
        valor_por_año = {int(row["YEAR"]): row["CF"] for _, row in df_tech.iterrows()}
        data = [round(valor_por_año.get(a, 0.0), 4) for a in años]
        series.append(
            ChartSeries(
                name=get_label(str(tech)),
                data=data,
                color=color_dict.get(tech, "#999999"),
                stack="default",
            )
        )

    return ChartDataResponse(
        categories=categories, series=series, title=title, yAxisLabel="%"
    )


# ═══════════════════════════════════════════════════════════════════════════
# 3. build_chart_data — SINGLE ESCENARIO
# ═══════════════════════════════════════════════════════════════════════════


def build_chart_data(
    db: Session,
    job_id: int,
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    variable: str | None = None,
    agrupar_por: str | None = None,
    es_porcentaje_override: bool = False,
) -> ChartDataResponse:
    """Construye la respuesta de gráfica para un solo escenario.

    Parámetros
    ----------
    db : Session
    job_id : int
    tipo : str
        Clave en ``CONFIGS`` (ej: ``'cap_electricidad'``, ``'gas_consumo'``).
    un : str
        Unidades de salida (PJ, GW, MW, TWh, Gpc).
    sub_filtro : str | None
        Filtro adicional dentro del sector.
    loc : str | None
        Localización (URB, RUR, ZNI).
    variable : str | None
        Override de variable para configs de capacidad.
    """
    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no existe en CONFIGS.")

    cfg = CONFIGS[tipo]
    es_capacidad = cfg.get("es_capacidad", False)
    es_porcentaje = cfg.get("es_porcentaje", False) or es_porcentaje_override
    es_factor_planta = cfg.get("es_factor_planta", False)

    # Variable a consultar
    variable_name = variable if (variable and es_capacidad) else cfg["variable_default"]

    # ── Título ───────────────────────────────────────────────────────────
    if es_capacidad:
        titulo_var = TITULOS_VARIABLES_CAPACIDAD.get(variable_name, variable_name)
        title = f"{cfg['titulo_base']} — {titulo_var}"
    elif es_porcentaje or es_factor_planta:
        title = cfg.get("titulo_base", cfg.get("titulo", tipo))
    else:
        title = cfg.get("titulo", tipo)

    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"

    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)

    if es_porcentaje or es_factor_planta:
        title += " (%)"
    elif es_emision:
        title += f" ({_emision_unit_label(un, es_emision_kt)})"
    else:
        title += f" ({un})"

    # ── Factor de Planta: pipeline propio ────────────────────────────────
    if es_factor_planta:
        return _build_factor_planta_data(db, job_id, cfg, title, sub_filtro, loc)

    # ── Cargar datos ─────────────────────────────────────────────────────
    df = _load_variable_data(db, job_id, variable_name)

    if df.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Filtrar ──────────────────────────────────────────────────────────
    filtro_fn = cfg.get("filtro")
    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)

    if df.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Agrupación ───────────────────────────────────────────────────────
    agrupar_col = agrupar_por if agrupar_por is not None else cfg["agrupar_por"]

    if agrupar_col == "TECNOLOGIA":
        # Algunas configs piden separar las refinerías en (refinería × combustible)
        # mientras dejan el resto agrupado por tecnología (típicamente imports).
        if cfg.get("split_refineries_by_fuel") and "FUEL" in df.columns:
            df["COLOR"] = df.apply(
                lambda r: (
                    f"{r['TECHNOLOGY']}::{r['FUEL']}"
                    if str(r.get("TECHNOLOGY", "")).startswith("UPSREF")
                    and str(r.get("FUEL", "")).strip() != ""
                    else r["TECHNOLOGY"]
                ),
                axis=1,
            )
        else:
            df["COLOR"] = df["TECHNOLOGY"]
    elif agrupar_col == "GROUP":
        if "FUEL" in df.columns:
            df["COLOR"] = (df["TECHNOLOGY"] + "_" + df["FUEL"]).apply(asignar_grupo)
        else:
            df["COLOR"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "FUEL":
        if "FUEL" in df.columns:
            df["COLOR"] = df.apply(_fuel_to_group, axis=1)
        else:
            df["COLOR"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "SECTOR":
        df["COLOR"] = _sector_labels(df["TECHNOLOGY"])
    elif agrupar_col == "EMISION":
        df["COLOR"] = df["FUEL"] if "FUEL" in df.columns else "?"
    elif agrupar_col == "H2_PRODUCCION":
        df["COLOR"] = df["TECHNOLOGY"].apply(_map_h2_verde_azul_gris)
    elif agrupar_col == "YEAR":
        # emisiones_total: solo agrupa por año
        df["COLOR"] = "Total"
    else:
        df["COLOR"] = df["TECHNOLOGY"]

    # ── Agregar ──────────────────────────────────────────────────────────
    df_agg = df.groupby(["COLOR", "YEAR"], as_index=False)["VALUE"].sum()

    # Descartar grupos insignificantes
    df_agg = df_agg[df_agg.groupby("COLOR")["VALUE"].transform("sum") > 1e-5]

    if df_agg.empty:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=un,
        )

    # ── Conversión de unidades ───────────────────────────────────────────
    if es_emision:
        if not es_emision_kt:
            df_agg = _convertir_unidades_emision(df_agg, un)
        # es_emision_kt: base = kt, sin conversión
    else:
        df_agg = _convertir_unidades(df_agg, un)

    # ── Porcentaje (prd_electricidad) ────────────────────────────────────
    if es_porcentaje:
        total_por_año = df_agg.groupby("YEAR")["VALUE"].transform("sum")
        df_agg["VALUE"] = df_agg["VALUE"] / total_por_año * 100.0

    # ── Colores ──────────────────────────────────────────────────────────
    # Si agrupar_por fue overridden, ajustar color_fn según agrupación
    if agrupar_por is not None and agrupar_por != cfg.get("agrupar_por"):
        if agrupar_col in ("FUEL", "GROUP"):
            color_fn = _color_por_grupo_fijo
        elif agrupar_col == "SECTOR":
            color_fn = _color_por_sector
        elif agrupar_col == "EMISION":
            color_fn = _color_por_emision
        else:
            color_fn = (
                cfg.get("color_fn")
                if cfg.get("color_fn") == _color_electricidad
                else generar_colores_tecnologias
            )
    else:
        color_fn = cfg.get("color_fn")
    if color_fn is not None:
        colores_ordenados, orden_color = color_fn(df_agg, "COLOR")
    else:
        orden_color = sorted(df_agg["COLOR"].unique())
        _palette = get_colores_grupos()
        colores_ordenados = [_palette.get(c, "#999999") for c in orden_color]

    color_dict = dict(zip(orden_color, colores_ordenados))

    # ── Construir respuesta ──────────────────────────────────────────────
    años = sorted(df_agg["YEAR"].unique())
    categories = [str(a) for a in años]

    def _composite_label(code: str) -> str:
        # COLOR de la forma "UPSREF_XXX::FUEL" → "Refinería ... — Combustible"
        if "::" in code:
            left, right = code.split("::", 1)
            return f"{get_label(left)} — {get_label(right)}"
        return get_label(code)

    series: list[ChartSeries] = []
    for tech in orden_color:
        df_tech = df_agg[df_agg["COLOR"] == tech]
        valor_por_año = {
            int(row["YEAR"]): row["VALUE"] for _, row in df_tech.iterrows()
        }
        data = [round(valor_por_año.get(a, 0.0), 6) for a in años]
        series.append(
            ChartSeries(
                name=_composite_label(str(tech)),
                data=data,
                color=color_dict.get(tech, "#999999"),
                stack="default",
            )
        )

    return ChartDataResponse(
        categories=categories,
        series=series,
        title=title,
        yAxisLabel="%"
        if es_porcentaje
        else (_emision_unit_label(un, es_emision_kt) if es_emision else un),
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4. build_comparison_data — MULTI-ESCENARIO
# ═══════════════════════════════════════════════════════════════════════════
def build_comparison_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    years_to_plot: list[int] | None = None,
    agrupacion: str | None = None,
    sub_filtro: str | None = None,
    loc: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
    es_porcentaje_override: bool = False,
) -> CompareChartResponse:
    """Construye la respuesta de comparación multi-escenario.

    Genera subplots por año clave, con barras apiladas por categoría.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en ``CONFIGS_COMPARACION``.
    un : str
        Unidades.
    years_to_plot : list[int] | None
        Años clave (default: [2024, 2030, 2050]).
    agrupacion : str | None
        Override de agrupación (TECNOLOGIA, COMBUSTIBLE, SECTOR).
    sub_filtro, loc : str | None
        Filtros opcionales.
    """
    # MAPEO DE SECTORES A COMPARACIÓN (Si el usuario escoge la tabla normal, la mappeamos a su configuración de comparación)
    MAPEO_COMPARACION = {
        "tra_total": "tra_comparacion",
        "ind_total": "ind_comparacion",
        "res_total": "res_comparacion",
        "ter_total": "ter_comparacion",
    }

    es_generico = False
    if tipo in MAPEO_COMPARACION:
        tipo = MAPEO_COMPARACION[tipo]

    if tipo not in CONFIGS_COMPARACION:
        if tipo in CONFIGS:
            es_generico = True
            cfg = CONFIGS[tipo]
            es_emision = cfg.get("es_emision", False)
        else:
            raise ValueError(
                f"tipo='{tipo}' no existe ni en CONFIGS ni en CONFIGS_COMPARACION."
            )
    else:
        cfg = CONFIGS_COMPARACION[tipo]

    if years_to_plot is None:
        years_to_plot = [2024, 2030, 2050]

    # Resolver agrupación y año histórico
    if not es_generico:
        prefijo = cfg["prefijo"]
        agrupacion_fija = cfg.get("agrupacion_fija")
        if agrupacion_fija is not None:
            agrupacion_usar = agrupacion_fija
        elif agrupacion is not None:
            agrupacion_usar = agrupacion
        else:
            agrupacion_usar = cfg["agrupacion_default"]

        usa_historico = cfg["año_historico_unico"]
        año_historico = years_to_plot[0] if years_to_plot else 2024

        label_agrupacion = {
            "TECNOLOGIA": "Tecnologías",
            "COMBUSTIBLE": "Combustibles",
            "SECTOR": "Sectores",
        }.get(agrupacion_usar, agrupacion_usar)
        title_base = f"{cfg['titulo_base']} por {label_agrupacion}"
    else:
        usa_historico = False
        año_historico = years_to_plot[0] if years_to_plot else 2024
        agrupacion_usar = "TECNOLOGIA"  # Fallback a agrupar por tecnología para cualquier otra gráfica
        title_base = cfg.get("titulo", cfg.get("titulo_base", tipo)) + " (Comparación)"

    title = title_base
    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"
    title += f" ({un})"

    # ── Cargar nombres de escenarios ─────────────────────────────────────
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            from app.models import Scenario

            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"
        # Override por alias del reporte, si se proporcionó.
        ov = (job_display_overrides or {}).get(jid)
        if isinstance(ov, str) and ov.strip():
            scenario_names[jid] = ov.strip()

    variable_name = cfg["variable_default"]

    # ── Procesar datos ───────────────────────────────────────────────────
    all_data: list[pd.DataFrame] = []

    # Paso 1: Año histórico (solo del primer escenario)
    if usa_historico and año_historico in years_to_plot and job_ids:
        first_job_id = job_ids[0]
        df_var = _load_variable_data(db, first_job_id, variable_name)

        if not df_var.empty:
            df_hist = _procesar_bloque_comparacion(
                df_var,
                prefijo,
                sub_filtro,
                loc,
                agrupacion_usar,
                [año_historico],
                un,
            )
            if df_hist is not None and not df_hist.empty:
                df_hist["SCENARIO"] = str(año_historico)
                all_data.append(df_hist)

    # Paso 2: Años proyectados (todos los escenarios)
    años_a_procesar = (
        [y for y in years_to_plot if y != año_historico]
        if usa_historico
        else years_to_plot
    )

    for jid in job_ids:
        df_var = _load_variable_data(db, jid, variable_name)
        if df_var.empty:
            continue

        if not es_generico:
            df = _procesar_bloque_comparacion(
                df_var,
                prefijo,
                sub_filtro,
                loc,
                agrupacion_usar,
                años_a_procesar,
                un,
            )
        else:
            df = _procesar_bloque_single(
                df_var, cfg, sub_filtro, loc, años_a_procesar, un
            )

        if df is None or df.empty:
            continue

        df["SCENARIO"] = scenario_names.get(jid, f"Job {jid}")
        all_data.append(df)

    if not all_data:
        return CompareChartResponse(title=title, subplots=[], yAxisLabel=un)

    df_final = pd.concat(all_data, ignore_index=True)

    # ── Porcentaje override ──────────────────────────────────────────────
    if es_porcentaje_override:
        total_por_año_escenario = df_final.groupby(["YEAR", "SCENARIO"])[
            "VALUE"
        ].transform("sum")
        df_final["VALUE"] = df_final["VALUE"] / total_por_año_escenario * 100.0

    # ── Colores ──────────────────────────────────────────────────────────
    categorias_unicas = sorted(df_final["CATEGORIA"].dropna().unique())
    if not es_generico:
        mapa_colores = _color_map_comparison(agrupacion_usar, categorias_unicas)
    else:
        # Reutilizar el sistema de colores original de la gráfica base
        color_fn = cfg.get("color_fn")
        if color_fn is not None:
            # Fake dataframe for generic coloring
            df_tmp = pd.DataFrame({"COLOR": list(categorias_unicas)})
            colores_lista, orden_lista = color_fn(df_tmp, "COLOR")
            mapa_colores = dict(zip(orden_lista, colores_lista))
        else:
            _palette = get_colores_grupos()
            mapa_colores = {c: _palette.get(c, "#999999") for c in categorias_unicas}

    # ── Construir subplots por año ───────────────────────────────────────
    años_ordenados = sorted(df_final["YEAR"].unique())
    subplots: list[SubplotData] = []

    for año in años_ordenados:
        df_año = df_final[df_final["YEAR"] == año]
        escenarios_en_año = sorted(df_año["SCENARIO"].unique())

        series: list[ChartSeries] = []
        for categoria in categorias_unicas:
            df_cat = df_año[df_año["CATEGORIA"] == categoria]
            if df_cat.empty:
                continue

            valor_por_escenario = {
                row["SCENARIO"]: row["VALUE"]
                for _, row in df_cat.groupby("SCENARIO", as_index=False)["VALUE"]
                .sum()
                .iterrows()
            }
            data = [
                round(valor_por_escenario.get(esc, 0.0), 6) for esc in escenarios_en_año
            ]

            series.append(
                ChartSeries(
                    name=get_label(str(categoria)),
                    data=data,
                    color=mapa_colores.get(categoria, "#999999"),
                    stack="default",
                )
            )

        subplots.append(
            SubplotData(
                year=int(año),
                categories=escenarios_en_año,
                series=series,
            )
        )

    return CompareChartResponse(
        title=title, subplots=subplots, yAxisLabel="%" if es_porcentaje_override else un
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4b. build_comparison_facet_data — ESCENARIOS COMPLETOS (FACETS)
# ═══════════════════════════════════════════════════════════════════════════


def build_comparison_facet_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    variable: str | None = None,
    agrupar_por: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
    es_porcentaje_override: bool = False,
) -> CompareChartFacetResponse:
    """Construye datos para comparación por escenarios completos (facets).

    Cada facet muestra la evolución temporal completa de un escenario.
    Usa CONFIGS (no CONFIGS_COMPARACION). Una query por job_id.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en CONFIGS (ej: 'cap_electricidad', 'gas_consumo').
    un : str
        Unidades de salida (PJ, GW, etc.).
    sub_filtro, loc, variable : str | None
        Filtros y override de variable.
    """
    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no encontrado en CONFIGS.")

    cfg = CONFIGS[tipo]
    title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))
    title = title_base
    if sub_filtro:
        sub_label = NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)
        title += f" — {sub_label}"
    if loc:
        title += f" ({loc})"
    title += f" ({un})"

    facets: list[FacetData] = []
    y_label = un
    from app.models import Scenario

    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if not job:
            continue
        scenario = None
        if job.scenario_id is not None:
            scenario = db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
        scenario_name = scenario.name if scenario else (job.input_name or f"Job {jid}")
        tag_name = None
        if scenario is not None:
            from app.services.simulation_service import SimulationService as _SS

            primary = _SS._batch_scenario_tags_by_scenario_ids(
                db, {int(scenario.id)}
            ).get(int(scenario.id))
            if primary:
                tag_name = primary.get("name")
        override = (
            (job_display_overrides or {}).get(jid) if job_display_overrides else None
        )
        has_alias_override = isinstance(override, str) and bool(override.strip())
        job_display = (
            override.strip()
            if has_alias_override
            else (getattr(job, "display_name", None) or None)
        )
        # Cuando el alias reemplaza el nombre del escenario, no queremos
        # concatenar la etiqueta al subtítulo (quedaría "Alias — Tag").
        effective_tag_name = None if has_alias_override else tag_name
        effective_scenario_name = (
            override.strip() if has_alias_override else scenario_name
        )

        chart = build_chart_data(
            db=db,
            job_id=jid,
            tipo=tipo,
            un=un,
            sub_filtro=sub_filtro,
            loc=loc,
            variable=variable,
            agrupar_por=agrupar_por,
            es_porcentaje_override=es_porcentaje_override,
        )

        if not facets:
            y_label = chart.yAxisLabel
        facets.append(
            FacetData(
                scenario_name=effective_scenario_name,
                job_id=jid,
                display_name=job_display,
                scenario_tag_name=effective_tag_name,
                categories=chart.categories,
                series=chart.series,
            )
        )
    # Unifica el eje X entre todos los facets: si un escenario llega hasta
    # 2030 y otro hasta 2033, ambos paneles muestran 2022-2033 con 0 en los
    # años faltantes para el escenario corto. Mantener ejes idénticos hace
    # que las comparaciones visuales sean directas.
    _align_facet_x_axis(facets)
    return CompareChartFacetResponse(
        title=title,
        facets=facets,
        yAxisLabel=y_label,
    )


def _align_facet_x_axis(facets: list[FacetData]) -> None:
    """Unifica el eje X entre todos los facets, in-place.

    1. Construye la unión ordenada de todas las categorías (ordena como int
       si todas son parseables como año, sino orden lexicográfico).
    2. Para cada facet, reescribe ``categories`` con la unión y rellena
       ``series.data`` con **None** en las posiciones que ese escenario no
       tenía.

    ``None`` (serializado como ``null`` en JSON) es importante porque:
      - **Líneas**: ``null`` crea un *gap* (Highcharts/matplotlib no traza
        el punto) en vez de hacer caer la línea a 0.
      - **Barras apiladas**: ``null`` no se dibuja como barra ni contribuye
        al total apilado — los totales (``StackItemObject.total``) no se
        contaminan.

    Si solo hay 0 o 1 facet, o todos comparten exactamente el mismo eje, no
    hace nada.
    """
    if not facets or len(facets) < 2:
        return
    # Cortocircuito: si todos los facets ya comparten exactamente las mismas
    # categorías en el mismo orden, no hacemos nada.
    first_cats = list(facets[0].categories)
    if all(list(f.categories) == first_cats for f in facets):
        return

    # 1) Construir la unión.
    union_set: set[str] = set()
    for f in facets:
        for c in f.categories:
            union_set.add(str(c))
    # ¿Todas las categorías son parseables como entero (año)?
    try:
        as_ints = sorted({int(c) for c in union_set})
        union: list[str] = [str(y) for y in as_ints]
    except ValueError:
        union = sorted(union_set)

    union_index: dict[str, int] = {c: i for i, c in enumerate(union)}
    n = len(union)

    # 2) Reescribir cada facet con la unión. Posiciones faltantes → None.
    for f in facets:
        old_cats = [str(c) for c in f.categories]
        old_index_in_union = [union_index.get(c) for c in old_cats]
        for s in f.series:
            new_data: list[float | None] = [None] * n
            for i, target in enumerate(old_index_in_union):
                if target is None or i >= len(s.data):
                    continue
                v = s.data[i]
                if v is None:
                    new_data[target] = None
                    continue
                try:
                    fv = float(v)
                    # NaN/Infinity también se representan como None.
                    import math as _math

                    new_data[target] = None if not _math.isfinite(fv) else fv
                except (TypeError, ValueError):
                    new_data[target] = None
            s.data = new_data
        f.categories = list(union)


def _procesar_bloque_comparacion(
    df_var: pd.DataFrame,
    prefijo: str | tuple[str, ...],
    sub_filtro: str | None,
    loc: str | None,
    agrupacion: str,
    años: list[int],
    un: str,
    es_emision: bool = False,
) -> pd.DataFrame | None:
    """Pipeline para un bloque de datos de comparación.

    filtrar → filtrar años → asignar categorías → agregar → convertir.
    Port de ``graficas_comparacion._procesar_bloque``.
    """
    if df_var is None or df_var.empty:
        return None

    if "TECHNOLOGY" not in df_var.columns or "YEAR" not in df_var.columns:
        return None

    df = _filtrar_df(df_var, prefijo, sub_filtro, loc)
    if df.empty:
        return None

    df = df[df["YEAR"].isin(años)]
    if df.empty:
        return None

    df = _asignar_categoria(df, agrupacion)

    df = df.groupby(["CATEGORIA", "YEAR"], as_index=False)["VALUE"].sum()

    # Descartar categorías insignificantes
    df = df[df.groupby("CATEGORIA")["VALUE"].transform("sum") > 1e-5]
    if df.empty:
        return None

    # df = _convertir_unidades(df, un)

    if not es_emision:
        df = _convertir_unidades(df, un)

    return df


def _procesar_bloque_single(
    df_var: pd.DataFrame,
    cfg: dict,
    sub_filtro: str | None,
    loc: str | None,
    años: list[int],
    un: str,
) -> pd.DataFrame | None:
    """Procesador genérico que emula la agrupación de build_chart_data para comparación."""
    if df_var is None or df_var.empty:
        return None

    if "TECHNOLOGY" not in df_var.columns or "YEAR" not in df_var.columns:
        return None

    df = df_var.copy()

    filtro_fn = cfg.get("filtro")
    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)

    if df.empty:
        return None

    df = df[df["YEAR"].isin(años)]
    if df.empty:
        return None

    agrupar_col = cfg["agrupar_por"]

    if agrupar_col == "TECNOLOGIA":
        df["CATEGORIA"] = df["TECHNOLOGY"]
    elif agrupar_col == "GROUP":
        if "FUEL" in df.columns:
            df["CATEGORIA"] = (df["TECHNOLOGY"] + "_" + df["FUEL"]).apply(asignar_grupo)
        else:
            df["CATEGORIA"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "FUEL":
        if "FUEL" in df.columns:
            df["CATEGORIA"] = df.apply(_fuel_to_group, axis=1)
        else:
            df["CATEGORIA"] = df["TECHNOLOGY"].apply(asignar_grupo)
    elif agrupar_col == "SECTOR":
        df["CATEGORIA"] = _sector_labels(df["TECHNOLOGY"])
    elif agrupar_col == "EMISION":
        df["CATEGORIA"] = df["FUEL"] if "FUEL" in df.columns else "?"
    elif agrupar_col == "YEAR":
        df["CATEGORIA"] = "Total"
    else:
        df["CATEGORIA"] = df["TECHNOLOGY"]

    df = df.groupby(["CATEGORIA", "YEAR"], as_index=False)["VALUE"].sum()
    df = df[df.groupby("CATEGORIA")["VALUE"].transform("sum") > 1e-5]

    if df.empty:
        return None

    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)
    if es_emision:
        if not es_emision_kt:
            df = _convertir_unidades_emision(df, un)
    else:
        df = _convertir_unidades(df, un)

    return df


# ═══════════════════════════════════════════════════════════════════════════
# 4c. build_comparison_line_data — LÍNEAS MULTI-ESCENARIO CONSOLIDADAS
# ═══════════════════════════════════════════════════════════════════════════

_SCENARIO_LINE_COLORS = [
    "#3b82f6",
    "#f59e0b",
    "#10b981",
    "#ef4444",
    "#8b5cf6",
    "#06b6d4",
    "#f97316",
    "#84cc16",
    "#ec4899",
    "#6366f1",
]


def build_comparison_line_data(
    db: Session,
    job_ids: list[int],
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
    job_display_overrides: dict[int, str] | None = None,
) -> ChartDataResponse:
    """Construye líneas totales multi-escenario sobre el mismo eje.

    Todos los escenarios se trazan en la misma figura (sin subplots).
    X = años, Y = total agregado (suma de todas las tecnologías), una línea por escenario.

    Parámetros
    ----------
    db : Session
    job_ids : list[int]
        IDs de simulation_job a comparar (max 10).
    tipo : str
        Clave en CONFIGS_COMPARACION o CONFIGS.
    un : str
        Unidades de salida (PJ, GW, etc.).
    sub_filtro, loc : str | None
        Filtros opcionales de tecnología/localización.
    """
    MAPEO_COMPARACION = {
        "tra_total": "tra_comparacion",
        "ind_total": "ind_comparacion",
        "res_total": "res_comparacion",
        "ter_total": "ter_comparacion",
    }
    if tipo in MAPEO_COMPARACION:
        tipo = MAPEO_COMPARACION[tipo]

    if tipo in CONFIGS_COMPARACION:
        cfg = CONFIGS_COMPARACION[tipo]
        variable_name: str = cfg["variable_default"]
        prefijo = cfg["prefijo"]
        es_emision = False
        es_emision_kt = False
        title_base = cfg["titulo_base"]

        def _apply_filter(df: pd.DataFrame) -> pd.DataFrame:
            return _filtrar_df(df, prefijo, sub_filtro, loc)

    elif tipo in CONFIGS:
        cfg = CONFIGS[tipo]
        variable_name = cfg["variable_default"]
        es_emision = cfg.get("es_emision", False)
        es_emision_kt = cfg.get("es_emision_kt", False)
        title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))
        filtro_fn = cfg.get("filtro")

        def _apply_filter(df: pd.DataFrame) -> pd.DataFrame:
            if filtro_fn is not None:
                return filtro_fn(df, sub_filtro=sub_filtro, loc=loc)
            return df
    else:
        raise ValueError(
            f"tipo='{tipo}' no existe en CONFIGS ni en CONFIGS_COMPARACION."
        )

    title = title_base
    if sub_filtro:
        title += f" — {NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)}"
    if loc:
        title += f" ({loc})"
    title += f" — Total ({un})"

    # Cargar nombres de escenarios
    scenario_names: dict[int, str] = {}
    for jid in job_ids:
        job = db.query(SimulationJob).filter(SimulationJob.id == jid).first()
        if job:
            from app.models import Scenario

            scenario = (
                db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
                if job.scenario_id
                else None
            )
            base = scenario.name if scenario else (job.input_name or f"Job {jid}")
            disp = (getattr(job, "display_name", None) or "").strip()
            scenario_names[jid] = disp if disp else base
        else:
            scenario_names[jid] = f"Job {jid}"
        ov = (job_display_overrides or {}).get(jid)
        if isinstance(ov, str) and ov.strip():
            scenario_names[jid] = ov.strip()

    # Agregar total por año para cada job
    all_years: set[int] = set()
    totals_per_job: dict[int, dict[int, float]] = {}

    for jid in job_ids:
        df = _load_variable_data(db, jid, variable_name)
        if df.empty:
            totals_per_job[jid] = {}
            continue
        df = _apply_filter(df)
        if df.empty:
            totals_per_job[jid] = {}
            continue
        if not es_emision:
            df = _convertir_unidades(df, un)
        year_totals = df.groupby("YEAR")["VALUE"].sum()
        totals_per_job[jid] = {
            int(y): round(float(v), 6) for y, v in year_totals.items()
        }
        all_years.update(totals_per_job[jid].keys())

    if not all_years:
        return ChartDataResponse(
            categories=[],
            series=[],
            title=title,
            yAxisLabel=_emision_unit_label(un, es_emision_kt) if es_emision else un,
        )

    years_sorted = sorted(all_years)
    categories = [str(y) for y in years_sorted]

    series: list[ChartSeries] = []
    for idx, jid in enumerate(job_ids):
        year_data = totals_per_job.get(jid, {})
        if not year_data:
            continue
        data = [year_data.get(y, 0.0) for y in years_sorted]
        series.append(
            ChartSeries(
                name=scenario_names.get(jid, f"Job {jid}"),
                data=data,
                color=_SCENARIO_LINE_COLORS[idx % len(_SCENARIO_LINE_COLORS)],
            )
        )

    es_emision_kt_line = (
        CONFIGS.get(tipo, {}).get("es_emision_kt", False) if tipo in CONFIGS else False
    )
    y_label = _emision_unit_label(un, es_emision_kt_line) if es_emision else un
    return ChartDataResponse(
        categories=categories, series=series, title=title, yAxisLabel=y_label
    )


# ═══════════════════════════════════════════════════════════════════════════
# 4d. build_pareto_data — PARETO POR TECNOLOGÍA
# ═══════════════════════════════════════════════════════════════════════════


def build_pareto_data(
    db: Session,
    job_id: int,
    tipo: str,
    un: str = "PJ",
    sub_filtro: str | None = None,
    loc: str | None = None,
) -> ParetoChartResponse:
    """Construye los datos para un gráfico de Pareto por tecnología.

    Agrega valores de todas las tecnologías sumando sobre todos los años,
    ordena de mayor a menor y calcula el porcentaje acumulado.

    Parámetros
    ----------
    db : Session
    job_id : int
    tipo : str
        Clave en CONFIGS con agrupar_por='TECNOLOGIA' y no es_capacidad.
    un : str
        Unidades de salida.
    sub_filtro, loc : str | None
        Filtros opcionales.
    """
    if tipo not in CONFIGS:
        raise ValueError(f"tipo='{tipo}' no encontrado en CONFIGS.")

    cfg = CONFIGS[tipo]
    es_capacidad = cfg.get("es_capacidad", False)
    es_emision = cfg.get("es_emision", False)
    es_emision_kt = cfg.get("es_emision_kt", False)
    variable_name: str = cfg["variable_default"]
    filtro_fn = cfg.get("filtro")
    title_base = cfg.get("titulo", cfg.get("titulo_base", tipo))

    _emi_label = _emision_unit_label(un, es_emision_kt)
    title = f"{title_base} — Pareto por Tecnología ({_emi_label if es_emision else un})"
    if sub_filtro:
        title += f" [{NOMBRES_COMBUSTIBLES.get(sub_filtro, sub_filtro)}]"

    df = _load_variable_data(db, job_id, variable_name)
    if df.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel=_emi_label if es_emision else un,
        )

    if filtro_fn is not None:
        df = filtro_fn(df, sub_filtro=sub_filtro, loc=loc)
    if df.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel=_emi_label if es_emision else un,
        )

    if not es_emision and not es_capacidad:
        df = _convertir_unidades(df, un)

    # Agregar por tecnología (suma sobre todos los años)
    tech_totals = df.groupby("TECHNOLOGY")["VALUE"].sum().reset_index()
    tech_totals = tech_totals[tech_totals["VALUE"] > 1e-5]
    tech_totals = tech_totals.sort_values("VALUE", ascending=False).reset_index(
        drop=True
    )

    if tech_totals.empty:
        return ParetoChartResponse(
            categories=[],
            values=[],
            cumulative_percent=[],
            title=title,
            yAxisLabel="MtCO₂eq" if es_emision else un,
        )

    total = tech_totals["VALUE"].sum()
    tech_totals["CUMSUM"] = tech_totals["VALUE"].cumsum()
    tech_totals["CUM_PCT"] = (tech_totals["CUMSUM"] / total * 100).round(2)

    categories = [get_label(str(t)) for t in tech_totals["TECHNOLOGY"]]
    values = [round(float(v), 6) for v in tech_totals["VALUE"]]
    cumulative_percent = [float(p) for p in tech_totals["CUM_PCT"]]

    y_label = _emi_label if es_emision else un
    return ParetoChartResponse(
        categories=categories,
        values=values,
        cumulative_percent=cumulative_percent,
        title=title,
        yAxisLabel=y_label,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 5. get_result_summary — KPIs
# ═══════════════════════════════════════════════════════════════════════════


def get_result_summary(
    db: Session,
    job_id: int,
    current_user_id=None,
) -> ResultSummaryResponse:
    """Devuelve resumen de KPIs para el header de visualización.

    Si se provee ``current_user_id`` (UUID), se completan los flags
    ``is_favorite`` para ese usuario; ``is_public``/``is_infeasible_result``
    se derivan del job mismo.
    """
    from app.models import Scenario, SimulationJobFavorite, User
    from app.services.simulation_service import SimulationService

    job = db.query(SimulationJob).filter(SimulationJob.id == job_id).first()

    if not job:
        raise ValueError(f"Job {job_id} no encontrado.")

    scenario = None
    if job.scenario_id is not None:
        scenario = db.query(Scenario).filter(Scenario.id == job.scenario_id).first()
    scenario_name = scenario.name if scenario else job.input_name
    scenario_tag = None
    scenario_tags_list: list[ScenarioTagPublic] = []
    if scenario is not None:
        all_tags = SimulationService._batch_all_scenario_tags_by_scenario_ids(
            db, {int(scenario.id)}
        ).get(int(scenario.id), [])
        scenario_tags_list = [ScenarioTagPublic.model_validate(t) for t in all_tags]
        if scenario_tags_list:
            scenario_tag = scenario_tags_list[0]

    solver_status = (job.model_timings_json or {}).get("solver_status", "unknown")

    total_co2 = (
        db.query(func.coalesce(func.sum(OsemosysOutputParamValue.value), 0))
        .filter(
            OsemosysOutputParamValue.id_simulation_job == job_id,
            OsemosysOutputParamValue.variable_name == "AnnualEmissions",
        )
        .scalar()
    ) or 0.0

    is_favorite = False
    if current_user_id is not None:
        is_favorite = (
            db.query(SimulationJobFavorite)
            .filter(
                SimulationJobFavorite.user_id == current_user_id,
                SimulationJobFavorite.job_id == job_id,
            )
            .first()
            is not None
        )

    owner = (
        db.query(User.username).filter(User.id == job.user_id).scalar()
        if job.user_id
        else None
    )

    return ResultSummaryResponse(
        job_id=job.id,
        scenario_id=job.scenario_id,
        scenario_name=scenario_name,
        scenario_tag=scenario_tag,
        scenario_tags=scenario_tags_list,
        display_name=getattr(job, "display_name", None) or None,
        solver_name=job.solver_name,
        solver_status=solver_status,
        objective_value=job.objective_value or 0.0,
        coverage_ratio=job.coverage_ratio or 0.0,
        total_demand=job.total_demand or 0.0,
        total_dispatch=job.total_dispatch or 0.0,
        total_unmet=job.total_unmet or 0.0,
        total_co2=float(total_co2),
        is_public=bool(getattr(job, "is_public", True)),
        is_favorite=bool(is_favorite),
        is_infeasible_result=SimulationService._is_infeasible_succeeded_job(job),
        owner_username=owner,
    )


# ═══════════════════════════════════════════════════════════════════════════
# 6. get_chart_catalog — CATÁLOGO DE GRÁFICAS
# ═══════════════════════════════════════════════════════════════════════════


def get_chart_catalog() -> list[ChartCatalogItem]:
    """Devuelve la lista de gráficas disponibles para el selector del frontend."""
    from app.schemas.visualization import DataExplorerFilters
    from app.visualization.data_explorer_filters import get_data_explorer_filters

    items: list[ChartCatalogItem] = []

    for config_id, cfg in CONFIGS.items():
        label = cfg.get("titulo", cfg.get("titulo_base", config_id))
        de_raw = get_data_explorer_filters(config_id, cfg.get("variable_default"))
        de_filters = (
            DataExplorerFilters(**{k: v for k, v in de_raw.items() if v})
            if de_raw
            else None
        )
        items.append(
            ChartCatalogItem(
                id=config_id,
                label=label,
                variable_default=cfg["variable_default"],
                has_sub_filtro=_config_has_sub_filtro(cfg),
                has_loc=_config_has_loc(cfg),
                sub_filtros=_config_sub_filtros(cfg),
                es_capacidad=cfg.get("es_capacidad", False),
                soporta_pareto=_config_soporta_pareto(cfg),
                data_explorer_filters=de_filters,
            )
        )

    return items


def _config_has_sub_filtro(cfg: dict) -> bool:
    """Determina si un config single-scenario soporta sub_filtro.

    Los configs de residencial, industrial, transporte y terciario
    aceptan sub_filtro a través de su función filtro.
    """
    filtro = cfg.get("filtro")
    if filtro is None:
        return False
    # Funciones que soportan sub_filtro por su signature
    filtro_name = getattr(filtro, "__name__", "")
    return filtro_name in (
        "_filtro_residencial",
        "_filtro_industrial",
        "_filtro_transporte",
        "_filtro_terciario",
        "_filtro_prefijo_con_sub",
        "_filtro_construccion",
        "_filtro_agroforestal",
        "_filtro_mineria",
        "_filtro_coquerias",
    )


def _config_has_loc(cfg: dict) -> bool:
    """Determina si un config single-scenario soporta localización."""
    filtro = cfg.get("filtro")
    if filtro is None:
        return False
    filtro_name = getattr(filtro, "__name__", "")
    return filtro_name == "_filtro_residencial"


def _config_soporta_pareto(cfg: dict) -> bool:
    """Pareto disponible para configs con agrupación TECNOLOGIA, no capacidad, no ratios."""
    return (
        cfg.get("agrupar_por") == "TECNOLOGIA"
        and not cfg.get("es_capacidad", False)
        and not cfg.get("es_factor_planta", False)
        and not cfg.get("es_porcentaje", False)
        and cfg.get("variable_default") not in ("AnnualEmissions",)
    )


def _config_sub_filtros(cfg: dict) -> list[str] | None:
    """Devuelve la lista de sub_filtros conocidos para un config, o None."""
    filtro = cfg.get("filtro")
    if filtro is None:
        return None
    filtro_name = getattr(filtro, "__name__", "")
    if filtro_name == "_filtro_residencial":
        return ["CKN", "WHT", "AIR", "REF", "ILU", "TV", "OTH"]
    if filtro_name == "_filtro_industrial":
        return ["BOI", "FUR", "MPW", "AIR", "REF", "ILU", "OTH"]
    if filtro_name == "_filtro_transporte":
        return [
            "AVI",
            "BOT",
            "SHP",
            "LDV",
            "FWD",
            "BUS",
            "TCK_C2P",
            "TCK_CSG",
            "MOT",
            "MIC",
            "TAX",
            "STT",
            "MET",
        ]
    if filtro_name == "_filtro_terciario":
        return ["AIR", "ILU", "OTH"]
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 7. EXPORT ALL — ZIP con gráficas como imágenes
# ═══════════════════════════════════════════════════════════════════════════


def export_all_charts_zip(
    db: Session,
    job_id: int,
    un: str = "PJ",
    fmt: str = "svg",
) -> "io.BytesIO":
    """Genera un ZIP con todas las gráficas renderizadas como SVG o PNG.

    Para configs de capacidad genera 3 figuras (Total, Nueva, Acumulada).
    Retorna un BytesIO listo para streaming.
    """
    import io
    import zipfile

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    CAPACITY_VARIABLES = [
        ("TotalCapacityAnnual", "Cap_Total"),
        ("NewCapacity", "Cap_Nueva"),
        ("AccumulatedNewCapacity", "Cap_Acumulada"),
    ]

    ext = "svg" if fmt == "svg" else "png"
    output = io.BytesIO()

    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        file_count = 0

        for config_id, cfg in CONFIGS.items():
            es_capacidad = cfg.get("es_capacidad", False)
            label = cfg.get("titulo", cfg.get("titulo_base", config_id))

            charts_to_render: list[tuple[str, ChartDataResponse]] = []

            if es_capacidad:
                for var_name, var_suffix in CAPACITY_VARIABLES:
                    chart = build_chart_data(
                        db,
                        job_id,
                        config_id,
                        un=un,
                        variable=var_name,
                    )
                    if chart.series:
                        charts_to_render.append((f"{label} — {var_suffix}", chart))
            else:
                chart = build_chart_data(db, job_id, config_id, un=un)
                if chart.series:
                    charts_to_render.append((label, chart))

            for chart_label, chart_data in charts_to_render:
                img_buf = _render_stacked_bar(
                    chart_data,
                    chart_label,
                    fmt=ext,
                )
                safe_name = _safe_filename(chart_label)
                zf.writestr(f"{safe_name}.{ext}", img_buf.getvalue())
                file_count += 1

    output.seek(0)
    return output


def _legend_ncols_for_labels(labels: list[str], hard_cap: int = 5) -> int:
    """Cap de columnas de leyenda según el largo máximo de las etiquetas.

    Evita que leyendas con etiquetas muy largas se expandan más allá del ancho
    de la figura — lo que con ``bbox_inches="tight"`` haría que matplotlib
    estire la imagen final horizontalmente, dejando el panel pequeño y con
    huecos blancos a los lados.
    """
    n = len(labels)
    if n == 0:
        return 1
    max_len = max(len(lab) for lab in labels)
    if max_len > 48:
        cap = 2
    elif max_len > 34:
        cap = 3
    elif max_len > 22:
        cap = 4
    else:
        cap = hard_cap
    return max(1, min(cap, hard_cap, n))


def _render_stacked_bar(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como gráfica de barras apiladas con matplotlib."""
    import io
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    bottom = np.zeros(n_cats)

    # Convención visual igual a Highcharts: la PRIMERA serie de
    # ``chart.series`` queda en la parte de ARRIBA del stack. Iteramos en
    # orden inverso para acumular.
    # NaN/None → 0 antes de acumular para no contaminar ``bottom`` (los
    # ``None`` pueden venir de synthetic series con huecos o de filtros).
    for s in reversed(chart.series):
        raw = np.array(s.data, dtype=float)
        values = np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0)
        ax.bar(x, values, bottom=bottom, color=s.color, width=0.7)
        bottom += values
    # Markers circulares estilo Highcharts para la leyenda (orden top→bottom).
    from matplotlib.lines import Line2D as _Line2D

    bar_handles = [
        _Line2D(
            [0],
            [0],
            marker="o",
            color=s.color,
            linestyle="None",
            markersize=10,
            markerfacecolor=s.color,
            markeredgecolor=s.color,
        )
        for s in chart.series
    ]
    bar_labels = [s.name for s in chart.series]

    # Stack totals on top — 1 decimal máx, cada 2 categorías (0, 2, 4, …).
    for i, total in enumerate(bottom):
        if i % 2 != 0:
            continue
        if total > 0:
            ax.text(
                i,
                total,
                f"{total:,.1f}",
                ha="center",
                va="bottom",
                fontsize=11,
                color="#333",
            )

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=12)
    ax.set_ylabel(chart.yAxisLabel, fontsize=14, fontweight="bold")
    ax.set_title(title, fontsize=17, fontweight="bold", pad=12)
    # Leyenda invertida respecto al stack: la primera serie (top del stack)
    # aparece al final de la leyenda → lectura abajo→arriba como las barras.
    ax.legend(
        list(reversed(bar_handles)),
        list(reversed(bar_labels)),
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels(bar_labels),
        fontsize=14,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.tick_params(axis="y", labelsize=10)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if y_axis_min is not None or y_axis_max is not None:
        cur_lo, cur_hi = ax.get_ylim()
        ax.set_ylim(
            float(y_axis_min) if y_axis_min is not None else cur_lo,
            float(y_axis_max) if y_axis_max is not None else cur_hi,
        )

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


_SYNTH_LINESTYLE_MAP: dict[str, Any] = {
    "Solid": "-",
    "Dash": "--",
    "Dot": ":",
    "DashDot": "-.",
    "ShortDash": (0, (3, 3)),
}

_SYNTH_MARKER_MAP: dict[str, str] = {
    "circle": "o",
    "diamond": "D",
    "square": "s",
    "triangle": "^",
    "triangle-down": "v",
    "none": "",
}


def _series_style_for_render(s: Any) -> dict[str, Any]:
    """Resuelve los kwargs de ``ax.plot`` para una serie (sintética o no).

    Para series no-sintéticas: defaults del renderer (`o`, ms=4, lw=2, sólida).
    Para sintéticas: usa los campos opcionales de la serie con fallback.
    """
    is_synth = bool(getattr(s, "is_synthetic", False))
    if not is_synth:
        return {"marker": "o", "markersize": 4, "linewidth": 2, "linestyle": "-"}

    raw_marker = getattr(s, "markerSymbol", None) or "diamond"
    marker = _SYNTH_MARKER_MAP.get(raw_marker, "D")
    raw_ls = getattr(s, "lineStyle", None) or "ShortDash"
    linestyle = _SYNTH_LINESTYLE_MAP.get(raw_ls, (0, (3, 3)))
    radius = getattr(s, "markerRadius", None)
    markersize = float(radius) if radius is not None else 5.0
    width = getattr(s, "lineWidth", None)
    linewidth = float(width) if width is not None else 2.0
    style: dict[str, Any] = {
        "marker": marker,
        "markersize": markersize,
        "linewidth": linewidth,
        "linestyle": linestyle,
    }
    if marker == "":
        style.pop("marker")
        style["markersize"] = 0
    return style


def _render_line_chart(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como gráfica de líneas con matplotlib."""
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    for s in chart.series:
        values = np.array(s.data, dtype=float)
        style = _series_style_for_render(s)
        ax.plot(
            x,
            values,
            label=s.name,
            color=s.color,
            **style,
        )

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=12)
    ax.set_ylabel(chart.yAxisLabel, fontsize=14, fontweight="bold")
    ax.set_title(title, fontsize=17, fontweight="bold", pad=12)
    # Orden de leyenda:
    #   1) Series naturales en orden invertido (lectura abajo→arriba como
    #      en las columnas apiladas — convención del proyecto).
    #   2) Series manuales (sintéticas) SIEMPRE al final, en su orden natural.
    _line_handles, _line_labels = ax.get_legend_handles_labels()
    _synth_flags = [bool(getattr(s, "is_synthetic", False)) for s in chart.series]
    _natural = [
        (h, l) for (h, l, f) in zip(_line_handles, _line_labels, _synth_flags) if not f
    ]
    _synth = [
        (h, l) for (h, l, f) in zip(_line_handles, _line_labels, _synth_flags) if f
    ]
    _ordered = list(reversed(_natural)) + _synth
    ax.legend(
        [h for h, _ in _ordered],
        [l for _, l in _ordered],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels([s.name for s in chart.series]),
        fontsize=14,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.tick_params(axis="y", labelsize=10)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if y_axis_min is not None or y_axis_max is not None:
        cur_lo, cur_hi = ax.get_ylim()
        ax.set_ylim(
            float(y_axis_min) if y_axis_min is not None else cur_lo,
            float(y_axis_max) if y_axis_max is not None else cur_hi,
        )

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def render_chart_visualization_bytes(
    chart: ChartDataResponse,
    fmt: str,
    view_mode: str = "column",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> bytes:
    """Genera PNG o SVG con Matplotlib.

    ``view_mode``: ``column`` | ``line`` | ``area`` | ``table``.
    ``y_axis_min`` / ``y_axis_max``: override del rango del eje Y. ``None`` = auto.
    """
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    title = chart.title
    if view_mode == "line":
        buf = _render_line_chart(
            chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max,
        )
    elif view_mode == "area":
        buf = _render_stacked_area(
            chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max,
        )
    elif view_mode == "table":
        buf = _render_table_image(chart, title, fmt=fmt)
    else:
        buf = _render_stacked_bar(
            chart, title, fmt=fmt, y_axis_min=y_axis_min, y_axis_max=y_axis_max,
        )
    return buf.getvalue()


def _render_table_image(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "png",
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como **tabla** (matplotlib ``ax.table``).

    Layout:
      * Header: categorías (años o periodos) en columnas.
      * Primera columna: nombre de serie.
      * Cuerpo: valores formateados con ``format_axis_3sig`` (≥ 3 cifras sig).
      * Última fila: "Total" con suma vertical por columna.

    El swatch de color por serie se aplica como ``cellColours`` en la
    primera columna (celda con el color de la serie + texto blanco).
    """
    import io
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    categories = list(chart.categories)
    series = list(chart.series)
    n_cols = 1 + len(categories)  # 1 columna para "Tecnología/Categoría"
    n_rows_body = len(series)

    # Cuerpo de la tabla
    cell_text: list[list[str]] = []
    for s in series:
        row = [s.name]
        for i in range(len(categories)):
            v = s.data[i] if i < len(s.data) else None
            row.append(format_axis_3sig(v))
        cell_text.append(row)

    # Fila Total (suma vertical por columna)
    totals: list[float] = []
    for i in range(len(categories)):
        col_total = 0.0
        for s in series:
            try:
                f = float(s.data[i]) if i < len(s.data) else 0.0
                if f != f:  # NaN
                    f = 0.0
            except (TypeError, ValueError):
                f = 0.0
            col_total += f
        totals.append(col_total)
    if n_rows_body > 0:
        cell_text.append(["Total"] + [format_axis_3sig(t) for t in totals])

    # Cabecera
    col_labels = ["Tecnología"] + [str(c) for c in categories]

    # Colores de celdas
    header_color = "#1e293b"  # fondo cabecera (slate)
    header_text_color = "#ffffff"
    alt_row = "#f8fafc"
    base_row = "#ffffff"
    total_row = "#e2e8f0"

    # Cell colours: misma forma que cell_text (filas × columnas).
    n_total_rows = len(cell_text)
    cell_colours: list[list[str]] = []
    for r_idx in range(n_total_rows):
        is_total = r_idx == n_total_rows - 1 and n_rows_body > 0
        if is_total:
            row_colors = [total_row] * n_cols
        else:
            base = alt_row if r_idx % 2 == 1 else base_row
            row_colors = [base] * n_cols
            # Primera columna con color de la serie.
            if r_idx < n_rows_body:
                row_colors[0] = series[r_idx].color or "#94a3b8"
        cell_colours.append(row_colors)

    # Tamaño dinámico de figura.
    fig_w = max(8.0, min(24.0, 1.6 + 1.4 * n_cols))
    fig_h = max(2.0, min(20.0, 1.6 + 0.55 * (n_total_rows + 1)))

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")
    ax.set_title(title, fontsize=17, fontweight="bold", pad=14)

    table = ax.table(
        cellText=cell_text if cell_text else [[""]],
        colLabels=col_labels,
        cellColours=cell_colours if cell_colours else None,
        colColours=[header_color] * n_cols,
        cellLoc="center",
        loc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(12)
    # Ancho mínimo razonable: matplotlib ajusta auto pero forzamos pad.
    table.scale(1.0, 1.4)

    # Estilo: cabecera bold, texto blanco; primera columna con texto blanco
    # contra el color de la serie. Si el color es muy claro, matplotlib
    # mostrará el texto en negro — pero los colores de serie tienden a ser
    # saturados así que blanco suele leerse bien.
    for (row, col), cell in table.get_celld().items():
        # row=0 → cabecera (porque colLabels existe).
        if row == 0:
            cell.set_text_props(color=header_text_color, fontweight="bold")
            cell.set_edgecolor("#0f172a")
        else:
            cell.set_edgecolor("#cbd5e1")
        # Primera columna del cuerpo (no la cabecera): texto blanco bold sobre
        # el color de la serie.
        if col == 0 and 0 < row <= n_rows_body:
            cell.set_text_props(color="#ffffff", fontweight="bold")
        # Fila Total: texto bold.
        if n_rows_body > 0 and row == n_total_rows:
            cell.set_text_props(fontweight="bold")

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=200, bbox_inches="tight", facecolor="#ffffff")
    plt.close(fig)
    buf.seek(0)
    return buf


def chart_data_to_xlsx_bytes(chart: ChartDataResponse) -> bytes:
    """Serializa ``ChartDataResponse`` a un workbook XLSX (wide format).

    Hoja única "Datos" con:
      * cabecera = ``["Categoría"] + [serie.name]``
      * filas    = una por categoría, valores numéricos.
      * fila final "Total" con suma vertical.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    headers = ["Categoría"] + [s.name for s in chart.series]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1e293b")
    header_font = Font(bold=True, color="ffffff")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas
    for i, cat in enumerate(chart.categories):
        row = [cat]
        for s in chart.series:
            v = s.data[i] if i < len(s.data) else None
            row.append(v)
        ws.append(row)

    # Fila Total
    if chart.categories and chart.series:
        totals_row: list = ["Total"]
        for s in chart.series:
            total = 0.0
            for v in s.data:
                try:
                    f = float(v)
                    if f != f:
                        f = 0.0
                except (TypeError, ValueError):
                    f = 0.0
                total += f
            totals_row.append(total)
        ws.append(totals_row)
        last_row_idx = ws.max_row
        bold = Font(bold=True)
        total_fill = PatternFill("solid", fgColor="e2e8f0")
        for col_idx in range(1, len(totals_row) + 1):
            c = ws.cell(row=last_row_idx, column=col_idx)
            c.font = bold
            c.fill = total_fill

    # Auto-ajuste de ancho de columnas
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if isinstance(v, float):
                    s = f"{v:,.2f}"
                else:
                    s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _render_stacked_area(
    chart: ChartDataResponse,
    title: str,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> "io.BytesIO":
    """Renderiza un ChartDataResponse como áreas apiladas."""
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = chart.categories
    n_cats = len(categories)
    x = np.arange(n_cats)

    fig, ax = plt.subplots(figsize=(max(12, n_cats * 0.5), 7))

    if chart.series:
        # stackplot dibuja la primera serie al fondo. Para que la convención
        # coincida con Highcharts (primera serie del array → arriba),
        # invertimos el orden antes de pasarlo a stackplot.
        # ``nan_to_num`` evita que NaN propague y rompa el stackplot.
        rev_series = list(reversed(chart.series))
        ys = [
            np.nan_to_num(
                np.array(s.data, dtype=float), nan=0.0, posinf=0.0, neginf=0.0,
            )
            for s in rev_series
        ]
        labels = [s.name for s in rev_series]
        colors = [getattr(s, "color", None) for s in rev_series]
        ax.stackplot(
            x,
            np.vstack(ys) if ys else np.zeros((0, n_cats)),
            labels=labels,
            colors=[c if c else None for c in colors],
            alpha=0.9,
            linewidth=0.5,
            edgecolor="white",
        )

    ax.set_xticks(x)
    ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=12)
    ax.tick_params(axis="y", labelsize=10)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax.set_ylabel(chart.yAxisLabel, fontsize=14, fontweight="bold")
    ax.set_title(title, fontsize=17, fontweight="bold", pad=12)
    # Markers circulares en la leyenda (estilo Highcharts), orden top→bottom.
    from matplotlib.lines import Line2D as _Line2D

    legend_handles = [
        _Line2D(
            [0],
            [0],
            marker="o",
            color=s.color or "#999999",
            linestyle="None",
            markersize=10,
            markerfacecolor=s.color or "#999999",
            markeredgecolor=s.color or "#999999",
        )
        for s in chart.series
    ]
    legend_labels = [s.name for s in chart.series]
    # Orden de leyenda: naturales en orden invertido (lectura abajo→arriba)
    # y series manuales (sintéticas) SIEMPRE al final.
    _synth_flags = [bool(getattr(s, "is_synthetic", False)) for s in chart.series]
    _natural = [
        (h, l)
        for (h, l, f) in zip(legend_handles, legend_labels, _synth_flags)
        if not f
    ]
    _synth = [
        (h, l) for (h, l, f) in zip(legend_handles, legend_labels, _synth_flags) if f
    ]
    _ordered = list(reversed(_natural)) + _synth
    ax.legend(
        [h for h, _ in _ordered],
        [l for _, l in _ordered],
        loc="upper center",
        bbox_to_anchor=(0.5, -0.15),
        ncol=_legend_ncols_for_labels(legend_labels),
        fontsize=14,
        frameon=False,
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.55,
    )
    ax.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlim(x.min() if n_cats > 0 else 0, x.max() if n_cats > 0 else 0)
    if y_axis_min is not None or y_axis_max is not None:
        cur_lo, cur_hi = ax.get_ylim()
        ax.set_ylim(
            float(y_axis_min) if y_axis_min is not None else cur_lo,
            float(y_axis_max) if y_axis_max is not None else cur_hi,
        )

    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def render_comparison_by_year_bytes(
    data: CompareChartResponse,
    fmt: str = "svg",
    *,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> bytes:
    """Renderiza una comparación por año (subplots, un panel por año).

    Cada subplot muestra barras agrupadas (una barra por escenario).
    """
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    subplots = [sp for sp in data.subplots if sp.series]
    if not subplots:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    n = len(subplots)
    cols = min(3, n)
    rows = (n + cols - 1) // cols
    fig, axes = plt.subplots(
        rows, cols, figsize=(max(6, cols * 5), max(4, rows * 4)), squeeze=False
    )

    # Paleta consistente por nombre de serie a través de subplots.
    all_names: list[str] = []
    for sp in subplots:
        for s in sp.series:
            if s.name not in all_names:
                all_names.append(s.name)
    name_to_color: dict[str, str] = {}
    for sp in subplots:
        for s in sp.series:
            if s.name not in name_to_color and getattr(s, "color", None):
                name_to_color[s.name] = s.color  # type: ignore[assignment]

    for idx, sp in enumerate(subplots):
        ax = axes[idx // cols][idx % cols]
        categories = list(sp.categories)
        nc = len(categories)
        ns = len(sp.series)
        if nc == 0 or ns == 0:
            ax.set_axis_off()
            continue
        x = np.arange(nc)
        width = 0.8 / ns
        for si, s in enumerate(sp.series):
            offset = (si - (ns - 1) / 2) * width
            ax.bar(
                x + offset,
                s.data,
                width=width,
                label=s.name,
                color=name_to_color.get(s.name) or getattr(s, "color", None),
            )
        ax.set_title(f"Año {sp.year}", fontsize=15, fontweight="bold")
        ax.set_xticks(x)
        ax.set_xticklabels(categories, rotation=90, ha="center", fontsize=12)
        ax.set_ylabel(data.yAxisLabel, fontsize=14, fontweight="bold")
        ax.tick_params(axis="y", labelsize=10)
        from matplotlib.ticker import FuncFormatter as _FuncFormatter

        ax.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
        ax.grid(axis="y", alpha=0.3, linewidth=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        if idx == 0:
            ax.legend(
                loc="upper center",
                bbox_to_anchor=(0.5, -0.2),
                ncol=_legend_ncols_for_labels([s.name for s in sp.series], hard_cap=4),
                fontsize=14,
                frameon=False,
                handlelength=1.0,
                handletextpad=0.6,
                columnspacing=1.85,
                labelspacing=0.55,
            )

    # Override del rango Y (aplica a todos los subplots por consistencia).
    if y_axis_min is not None or y_axis_max is not None:
        for j in range(n):
            ax = axes[j // cols][j % cols]
            cur_lo, cur_hi = ax.get_ylim()
            ax.set_ylim(
                float(y_axis_min) if y_axis_min is not None else cur_lo,
                float(y_axis_max) if y_axis_max is not None else cur_hi,
            )

    # Ocultar axes sobrantes
    for j in range(n, rows * cols):
        axes[j // cols][j % cols].set_axis_off()

    fig.suptitle(data.title, fontsize=17, fontweight="bold")
    fig.tight_layout(rect=[0, 0.02, 1, 0.96])

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def render_pareto_chart_bytes(
    pareto: ParetoChartResponse,
    fmt: str = "svg",
) -> bytes:
    """Renderiza un ParetoChartResponse (barras descendentes + % acumulado)."""
    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    categories = list(pareto.categories)
    values = list(pareto.values)
    cum_pct = list(pareto.cumulative_percent)
    n = len(categories)
    if n == 0:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center")
        ax.set_axis_off()
        buf = io.BytesIO()
        fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
        plt.close(fig)
        buf.seek(0)
        return buf.getvalue()

    x = np.arange(n)
    fig, ax1 = plt.subplots(figsize=(max(12, n * 0.5), 7))
    ax1.bar(x, values, color="#60a5fa", edgecolor="#1e3a8a", linewidth=0.5)
    ax1.set_ylabel(pareto.yAxisLabel, fontsize=14, fontweight="bold", color="#1e3a8a")
    ax1.set_xticks(x)
    # Eje X a 45° (más legible que vertical para etiquetas largas tipo
    # tecnología/sector). ``ha="right"`` ancla el final de la etiqueta al tick
    # para que no se solape con la barra siguiente.
    ax1.set_xticklabels(
        categories,
        rotation=45,
        ha="right",
        rotation_mode="anchor",
        fontsize=12,
    )
    ax1.tick_params(axis="y", labelsize=10)
    from matplotlib.ticker import FuncFormatter as _FuncFormatter

    ax1.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax1.grid(axis="y", alpha=0.3, linewidth=0.5)
    ax1.spines["top"].set_visible(False)

    ax2 = ax1.twinx()
    ax2.plot(x, cum_pct, color="#dc2626", marker="o", linewidth=2)
    ax2.set_ylabel("% acumulado", fontsize=14, fontweight="bold", color="#dc2626")
    ax2.tick_params(axis="y", labelsize=10)
    ax2.yaxis.set_major_formatter(_FuncFormatter(lambda v, _p: format_axis_3sig(v)))
    ax2.set_ylim(0, 110)
    ax2.spines["top"].set_visible(False)

    ax1.set_title(pareto.title, fontsize=17, fontweight="bold", pad=12)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format=fmt, dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def _facet_x_axis_label_step(categories: list[Any]) -> int:
    """Paso entre etiquetas visibles del eje X (responsive).

    Si las categorías son años (1800–2200): como mínimo cada **2 años** y a lo sumo
    unas **12** marcas con texto para horizontes largos.
    En otro caso, limita a ~12 etiquetas sobre el total de categorías.
    """
    n = len(categories)
    if n <= 1:
        return 1

    try:
        years = [int(str(c).strip()) for c in categories]
    except (ValueError, TypeError):
        years = None
    else:
        if not years or not all(1800 <= y <= 2200 for y in years):
            years = None

    target_visible = 12
    if years is not None:
        return max(2, (n + target_visible - 1) // target_visible)

    return max(1, (n + target_visible - 1) // target_visible)


def _facet_x_ticklabels_thinned(categories: list[Any], step: int) -> list[str]:
    """Etiquetas con cadena vacía en índices omitidos; asegura inicio y fin legibles."""
    n = len(categories)
    if step < 1:
        step = 1
    out: list[str] = [""] * n
    for i in range(0, n, step):
        out[i] = str(categories[i])
    if n > 1:
        if not out[0]:
            out[0] = str(categories[0])
        if not out[-1]:
            out[-1] = str(categories[-1])
    return out


def render_comparison_facet_figure_bytes(
    data: CompareChartFacetResponse,
    fmt: str = "png",
    *,
    legend_title: str | None = None,
    y_axis_min: float | None = None,
    y_axis_max: float | None = None,
) -> bytes:
    """Una sola figura: facetas en fila, título global, leyenda inferior (Matplotlib).

    Prioriza **legibilidad**: misma escala Y entre escenarios, tipografía clara, leyenda
    con marco y números formateados en ejes y totales de barra cuando aportan.
    """
    import io

    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from matplotlib import patheffects as pe
    from matplotlib.patches import Rectangle
    from matplotlib.ticker import AutoMinorLocator, FuncFormatter, MaxNLocator

    if fmt not in ("png", "svg"):
        raise ValueError("fmt debe ser 'png' o 'svg'")

    facets = [f for f in data.facets if f.series]
    if not facets:
        raise ValueError("Sin facetas con series para exportar")

    n = len(facets)
    legend_order: list[tuple[str, str]] = []
    seen_names: set[str] = set()
    for facet in facets:
        for s in facet.series:
            if s.name not in seen_names:
                seen_names.add(s.name)
                legend_order.append((s.name, s.color))

    n_leg_items = len(legend_order)
    legend_labels_full = [name for name, _c in legend_order]
    max_leg_label_len = max((len(lab) for lab in legend_labels_full), default=0)

    # Leyenda: **siempre** texto completo; menos columnas si los nombres son largos.
    # Cap duro de **5 columnas** para evitar que figuras con muchas series
    # corten demasiado ancho horizontal con leyendas extremadamente anchas.
    LEG_NCOL_MAX = 5
    if max_leg_label_len > 48:
        leg_ncol = max(1, min(2, n_leg_items))
    elif max_leg_label_len > 34:
        leg_ncol = max(1, min(3, n_leg_items))
    elif max_leg_label_len > 22:
        leg_ncol = max(1, min(4, n_leg_items))
    else:
        leg_ncol = max(1, min(LEG_NCOL_MAX, n_leg_items))
    n_leg_rows = max(1, (n_leg_items + leg_ncol - 1) // leg_ncol)

    # Ancho por panel: aprovechamos un "presupuesto" total horizontal de ~26″
    # repartido entre los `n` paneles, con un mínimo de 5.0″ y un máximo de
    # 9.0″ por panel. Así con pocos escenarios cada panel es bien ancho, y
    # con muchos escenarios mantenemos legibilidad sin estirar la figura
    # más de lo necesario.
    inter_panel = 0.55
    fig_w_target = 26.0
    w_per_facet = max(
        5.0,
        min(9.0, (fig_w_target - inter_panel * max(0, n - 1)) / n),
    )
    fig_w = w_per_facet * n + inter_panel * max(0, n - 1)
    fig_w = max(9.0, fig_w)
    # Leyenda más ancha en figuras anchas — pero respetando el cap de 5 cols.
    if max_leg_label_len <= 40:
        leg_ncol = min(
            n_leg_items,
            max(leg_ncol, min(LEG_NCOL_MAX, 3 + max(0, int(fig_w // 2.6)))),
        )
        n_leg_rows = max(1, (n_leg_items + leg_ncol - 1) // leg_ncol)

    # `leg_font` se usa abajo para el cálculo de altura; lo derivamos aquí.
    # Bump perceptual: la figura del facet es ~3× más ancha que un single,
    # así que la leyenda visualmente se siente pequeña aunque su tamaño en pt
    # sea mayor. Subimos +4pt sobre el cálculo previo para que en perspectiva
    # se vea similar a la leyenda de los single-chart.
    leg_font_estimate = (
        21.0
        if max_leg_label_len > 52 or n_leg_items > 14
        else (21.6 if max_leg_label_len > 36 or n_leg_items > 10 else 22.8)
    )
    leg_font_estimate = float(min(leg_font_estimate, 23.5))

    # Geometría en PULGADAS — las fracciones de figura se derivan de aquí
    # para que paneles, x-labels y leyenda no se superpongan nunca.
    title_band_inch = 1.05  # suptitle (24pt) + aire respecto a títulos de panel (20pt)
    x_label_inch = 0.88  # años rotados 90° (≈17pt) + padding
    gap_inch = 0.24  # separación x-labels ↔ leyenda
    legend_pad_inch = 0.12  # padding leyenda ↔ borde inferior figura
    line_h_inch = leg_font_estimate * 1.35 / 72.0
    legend_h_inch = line_h_inch * n_leg_rows + 0.12

    bottom_margin_inch = legend_pad_inch + legend_h_inch + gap_inch + x_label_inch

    # Altura objetivo del panel (axes): ~55% del ancho del panel ⇒ ratio
    # ancho:alto ≈ 1.8:1 (claramente más ancho que alto). Floor de 4.0″ para
    # mantener legibilidad cuando hay muchos escenarios.
    panel_h_inch = max(4.0, w_per_facet * 0.55)

    fig_h = panel_h_inch + title_band_inch + bottom_margin_inch
    fig_h = float(min(max(fig_h, 6.0), 14.0))

    fig, axes = plt.subplots(1, n, figsize=(fig_w, fig_h), squeeze=False)
    row_axes = axes[0]

    y_label = data.yAxisLabel or "Valor"
    stack_tops: list[np.ndarray] = []

    for ax, facet in zip(row_axes, facets):
        categories = list(facet.categories)
        n_cats = len(categories)
        x = np.arange(n_cats, dtype=float)
        bottom = np.zeros(n_cats, dtype=float)

        # Iteramos en reverso: primer elemento de ``facet.series`` queda
        # arriba del stack (igual a Highcharts).
        #
        # NOTA: ``s.data`` puede contener ``None`` (NaN al convertir a float)
        # cuando ``_align_facet_x_axis`` rellenó años faltantes en facets de
        # rango distinto. Si propagamos NaN al ``bottom`` cumulado, ``np.max``
        # del stack devuelve NaN, ``global_max`` queda NaN, ``y_top = NaN`` y
        # ``set_ylim(0, NaN)`` deja que matplotlib auto-ajuste a una altura
        # menor — cortando visualmente los datos. Convertimos NaN→0 ANTES
        # del stack (zero-height bar = invisible, equivalente a "sin dato").
        for s in reversed(facet.series):
            raw = np.array(s.data, dtype=float)
            if raw.size < n_cats:
                raw = np.pad(raw, (0, n_cats - int(raw.size)))
            elif raw.size > n_cats:
                raw = raw[:n_cats]
            values = np.nan_to_num(raw, nan=0.0, posinf=0.0, neginf=0.0)
            ax.bar(
                x,
                values,
                bottom=bottom,
                color=s.color,
                width=0.74,
                edgecolor="#ffffff",
                linewidth=0.45,
            )
            bottom = bottom + values

        stack_tops.append(bottom.copy())

        ax.set_xticks(x)
        x_step = _facet_x_axis_label_step(categories)
        x_labels = _facet_x_ticklabels_thinned(categories, x_step)
        n_labeled = sum(1 for lb in x_labels if lb)
        x_fs = (
            15
            if n_labeled > 14 or n_cats > 36
            else (16 if n_cats > 22 or n_labeled > 11 else 17)
        )
        ax.set_xticklabels(
            x_labels,
            rotation=90,
            ha="center",
            fontsize=x_fs,
            color="#1e293b",
        )
        ax.set_ylabel(
            y_label,
            fontsize=19,
            color="#0f172a",
            fontweight="bold",
            labelpad=8,
        )
        sim_lbl = (
            facet.display_name or facet.scenario_name or f"Job {facet.job_id}"
        ).strip()
        tag_lbl = (facet.scenario_tag_name or "").strip()
        facet_title = f"{sim_lbl} — {tag_lbl}" if tag_lbl else sim_lbl
        ax.set_title(
            facet_title,
            fontsize=20,
            fontweight="bold",
            color="#0f172a",
            pad=10,
        )
        ax.set_axisbelow(True)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        for _side in ("left", "bottom"):
            ax.spines[_side].set_visible(True)
            ax.spines[_side].set_color("#1e293b")
            ax.spines[_side].set_linewidth(1.35)
        ax.tick_params(
            axis="y",
            labelsize=18,
            colors="#0f172a",
            width=1.15,
            length=6,
            labelcolor="#0f172a",
        )
        ax.tick_params(
            axis="x",
            colors="#0f172a",
            width=1.15,
            length=5,
            labelcolor="#1e293b",
        )
        ax.set_facecolor("#ffffff")

    global_max = 0.0
    # Línea ~2809
    global_max = 0.0
    for b in stack_tops:
        if b.size:
            b_clean = b[np.isfinite(b)]  # ← ignorar nan/inf
            if b_clean.size:
                global_max = max(global_max, float(b_clean.max()))
    if global_max <= 0:
        global_max = 1.0
    y_top = global_max * 1.12

    show_stack_totals = all(len(b) <= 18 for b in stack_tops)

    # Override del rango Y por usuario (aplica a TODOS los facets para que
    # mantengan misma escala — la idea del facet es comparar visualmente).
    effective_y_lo = float(y_axis_min) if y_axis_min is not None else 0.0
    effective_y_hi = float(y_axis_max) if y_axis_max is not None else y_top

    for ax, bottom in zip(row_axes, stack_tops):
        ax.set_ylim(effective_y_lo, effective_y_hi)
        ax.yaxis.set_major_formatter(
            FuncFormatter(lambda v, _p: format_axis_3sig(v))
        )
        ax.yaxis.set_major_locator(
            MaxNLocator(nbins=7, min_n_ticks=5, steps=[1, 2, 2.5, 5, 10]),
        )
        # Líneas horizontales intermedias entre marcas principales
        ax.yaxis.set_minor_locator(AutoMinorLocator(2))

        # Rejilla: lectura de valores (horizontales mayor + menor; verticales por categoría)
        ax.grid(
            which="major",
            axis="y",
            linestyle="-",
            linewidth=0.95,
            color="#64748b",
            alpha=0.55,
            zorder=0,
        )
        ax.grid(
            which="minor",
            axis="y",
            linestyle=":",
            linewidth=0.7,
            color="#94a3b8",
            alpha=0.5,
            zorder=0,
        )
        ax.grid(
            which="major",
            axis="x",
            linestyle="-",
            linewidth=0.75,
            color="#94a3b8",
            alpha=0.42,
            zorder=0,
        )

        if not show_stack_totals:
            continue
        n_cats = len(bottom)
        for i in range(n_cats):
            # Mostrar el total solo cada 2 categorías (0, 2, 4, …).
            if i % 2 != 0:
                continue
            total = float(bottom[i])
            if total <= 0 or total < global_max * 0.018:
                continue
            t = ax.text(
                i,
                total,
                f"{total:,.1f}",
                ha="center",
                va="bottom",
                fontsize=11.5,
                color="#0f172a",
                fontweight="600",
            )
            t.set_path_effects([pe.withStroke(linewidth=2.5, foreground="white")])

    fig.patch.set_facecolor("#ffffff")
    # Suptitle: lo posicionamos ~0.30" desde el borde superior, lo que deja
    # un aire claro respecto a los títulos por panel (que viven dentro de la
    # banda reservada arriba).
    suptitle_y = 1.0 - 0.28 / fig_h
    fig.suptitle(
        data.title,
        fontsize=24,
        fontweight="bold",
        color="#020617",
        y=suptitle_y,
    )

    # Markers circulares (estilo Highcharts) para coincidir con el chart
    # individual y evitar la barra/cuadrado.
    from matplotlib.lines import Line2D as _Line2D

    handles = [
        _Line2D(
            [0],
            [0],
            marker="o",
            color=c,
            linestyle="None",
            markersize=10,
            markerfacecolor=c,
            markeredgecolor=c,
        )
        for _name, c in legend_order
    ]
    leg_font = leg_font_estimate

    # Las fracciones de figura derivan directo de las pulgadas calculadas
    # arriba — así el cálculo es coherente y no quedan superposiciones.
    bottom_margin = bottom_margin_inch / fig_h
    top_margin = 1.0 - title_band_inch / fig_h
    legend_anchor_y = legend_pad_inch / fig_h

    # Leyenda al estilo de las gráficas individuales: sin marco, sin título.
    # Leyenda invertida respecto al stack (lectura abajo→arriba como las barras).
    fig.legend(
        handles=list(reversed(handles)),
        labels=list(reversed(legend_labels_full)),
        loc="lower center",
        bbox_to_anchor=(0.5, legend_anchor_y),
        ncol=leg_ncol,
        fontsize=leg_font,
        frameon=False,
        labelcolor="#0f172a",
        handlelength=1.0,
        handletextpad=0.6,
        columnspacing=1.85,
        labelspacing=0.45,
    )

    # `wspace` se reduce con menos paneles para que la separación entre
    # paneles no domine la figura (el ojo percibe huecos enormes con n=2-3).
    wspace = 0.20 if n >= 4 else 0.16
    plt.subplots_adjust(
        left=0.07,
        right=0.995,
        top=top_margin,
        bottom=bottom_margin,
        wspace=wspace,
    )

    buf = io.BytesIO()
    fig.savefig(
        buf,
        format=fmt,
        dpi=200,
        facecolor="#ffffff",
        edgecolor="none",
        bbox_inches="tight",
        pad_inches=0.18,
    )
    plt.close(fig)
    return buf.getvalue()


def chart_data_to_csv_bytes(chart: ChartDataResponse) -> bytes:
    """Tabla categorías × series como CSV UTF-8 con BOM (compatible con Excel)."""
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Categoria"] + [s.name for s in chart.series])
    for i, cat in enumerate(chart.categories):
        row: list[Any] = [cat]
        for s in chart.series:
            val = s.data[i] if i < len(s.data) else None
            row.append(val)
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def pareto_data_to_csv_bytes(pareto: ParetoChartResponse) -> bytes:
    """CSV UTF-8 con BOM: categoría, valor, % acumulado."""
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Categoria", pareto.yAxisLabel, "% Acumulado"])
    for i, cat in enumerate(pareto.categories):
        val = pareto.values[i] if i < len(pareto.values) else None
        cum = (
            pareto.cumulative_percent[i] if i < len(pareto.cumulative_percent) else None
        )
        writer.writerow([cat, val, cum])
    return buffer.getvalue().encode("utf-8-sig")


def _safe_filename(name: str) -> str:
    """Genera un nombre de archivo seguro."""
    clean = "".join(c if c.isalnum() or c in (" ", "_", "-") else "_" for c in name)
    return clean.strip()[:80]


# ═══════════════════════════════════════════════════════════════════════════
# 8. EXPORT RAW DATA — Excel
# ═══════════════════════════════════════════════════════════════════════════


def export_raw_data_excel(
    db: Session,
    job_id: int,
) -> "io.BytesIO":
    """Exporta todos los datos crudos del job a un archivo Excel (.xlsx)."""
    import io

    rows = (
        db.query(OsemosysOutputParamValue)
        .filter(OsemosysOutputParamValue.id_simulation_job == job_id)
        .all()
    )

    if not rows:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=404,
            detail="No hay datos crudos disponibles para este escenario. La simulación puede no haber guardado resultados en la base de datos.",
        )

    records = []
    for r in rows:
        records.append(
            {
                "VariableName": r.variable_name,
                "Technology": r.technology_name or "",
                "Fuel": r.fuel_name or "",
                "Emission": r.emission_name or "",
                "Year": r.year,
                "Value": float(r.value),
                "IndexJSON": str(r.index_json) if r.index_json else "",
            }
        )

    df = pd.DataFrame(records)

    output = io.BytesIO()
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        worksheet = writer.sheets["Raw Data"]
        # Autofit columns without depending on xlsxwriter.
        for idx, col in enumerate(df):
            series = df[col]
            value_len_max = int(
                series.apply(lambda x: len(str(x)) if pd.notna(x) else 0).max()
            )
            max_len = max(value_len_max, len(str(series.name))) + 2
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_len

    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════════════════════════
# 9. EXPORT RESULTS — ZIP de CSVs por variable (formato OSeMOSYS estándar)
# ═══════════════════════════════════════════════════════════════════════════

# Dimensiones por variable. Extiende VARIABLE_INDEX_NAMES con las legacy
# tipadas (Dispatch / UnmetDemand) que se persisten directamente vía
# pipeline._build_output_rows pero no aparecen en el registry.
_LEGACY_TYPED_INDEX_NAMES: dict[str, tuple[str, ...]] = {
    "Dispatch": ("REGION", "TECHNOLOGY", "FUEL", "YEAR"),
    "UnmetDemand": ("REGION", "YEAR"),
}

# Overrides de orden de columnas para el CSV exportado, donde el orden
# OSeMOSYS estándar difiere del usado en VARIABLE_INDEX_NAMES.
# (No modificar VARIABLE_INDEX_NAMES: se usa para interpretar índices al
# persistir resultados.)
_EXPORT_INDEX_OVERRIDES: dict[str, tuple[str, ...]] = {
    "ProductionByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
    "UseByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
    "RateOfProductionByTechnology": (
        "REGION",
        "TIMESLICE",
        "TECHNOLOGY",
        "FUEL",
        "YEAR",
    ),
    "RateOfUseByTechnology": ("REGION", "TIMESLICE", "TECHNOLOGY", "FUEL", "YEAR"),
}


def export_results_csv_zip(
    db: Session,
    job_id: int,
) -> "io.BytesIO":
    """Exporta resultados de un job a un ZIP con un CSV por variable.

    Cada CSV usa el formato estándar OSeMOSYS: columnas de dimensión en el
    orden declarado por ``VARIABLE_INDEX_NAMES``, seguidas de ``VALUE``.
    Un archivo por ``variable_name`` presente en BD.
    """
    import io
    import zipfile

    from app.models import (
        Dailytimebracket,
        Daytype,
        Emission,
        Fuel,
        ModeOfOperation,
        Region,
        Season,
        StorageSet,
        Technology,
        Timeslice,
    )
    from app.simulation.core.results_processing import VARIABLE_INDEX_NAMES

    rows = (
        db.query(
            OsemosysOutputParamValue.variable_name,
            Region.name.label("region"),
            Technology.name.label("technology"),
            Fuel.name.label("fuel"),
            Emission.name.label("emission"),
            Timeslice.code.label("timeslice"),
            ModeOfOperation.code.label("mode_of_operation"),
            StorageSet.code.label("storage"),
            Season.code.label("season"),
            Daytype.code.label("daytype"),
            Dailytimebracket.code.label("dailytimebracket"),
            OsemosysOutputParamValue.year,
            OsemosysOutputParamValue.value,
            OsemosysOutputParamValue.technology_name,
            OsemosysOutputParamValue.fuel_name,
            OsemosysOutputParamValue.emission_name,
        )
        .outerjoin(Region, OsemosysOutputParamValue.id_region == Region.id)
        .outerjoin(Technology, OsemosysOutputParamValue.id_technology == Technology.id)
        .outerjoin(Fuel, OsemosysOutputParamValue.id_fuel == Fuel.id)
        .outerjoin(Emission, OsemosysOutputParamValue.id_emission == Emission.id)
        .outerjoin(Timeslice, OsemosysOutputParamValue.id_timeslice == Timeslice.id)
        .outerjoin(
            ModeOfOperation,
            OsemosysOutputParamValue.id_mode_of_operation == ModeOfOperation.id,
        )
        .outerjoin(StorageSet, OsemosysOutputParamValue.id_storage == StorageSet.id)
        .outerjoin(Season, OsemosysOutputParamValue.id_season == Season.id)
        .outerjoin(Daytype, OsemosysOutputParamValue.id_daytype == Daytype.id)
        .outerjoin(
            Dailytimebracket,
            OsemosysOutputParamValue.id_dailytimebracket == Dailytimebracket.id,
        )
        .filter(OsemosysOutputParamValue.id_simulation_job == job_id)
        .all()
    )

    if not rows:
        from fastapi import HTTPException

        raise HTTPException(
            status_code=404,
            detail="No hay resultados para este escenario.",
        )

    _DIM_TO_ROW_ATTR = {
        "REGION": "region",
        "TECHNOLOGY": "technology",
        "FUEL": "fuel",
        "EMISSION": "emission",
        "TIMESLICE": "timeslice",
        "MODE_OF_OPERATION": "mode_of_operation",
        "STORAGE": "storage",
        "SEASON": "season",
        "DAYTYPE": "daytype",
        "DAILYTIMEBRACKET": "dailytimebracket",
    }

    def _index_names_for(var_name: str) -> tuple[str, ...]:
        if var_name in _EXPORT_INDEX_OVERRIDES:
            return _EXPORT_INDEX_OVERRIDES[var_name]
        if var_name in VARIABLE_INDEX_NAMES:
            return VARIABLE_INDEX_NAMES[var_name]
        if var_name in _LEGACY_TYPED_INDEX_NAMES:
            return _LEGACY_TYPED_INDEX_NAMES[var_name]
        return ()

    def _fallback_dims_from_row(r) -> list[str]:
        """Para variables desconocidas, deduce columnas no nulas observadas."""
        dims: list[str] = []
        for dim, attr in _DIM_TO_ROW_ATTR.items():
            if getattr(r, attr, None) not in (None, ""):
                dims.append(dim)
        if r.year is not None:
            dims.append("YEAR")
        return dims

    by_var: dict[str, list] = {}
    for r in rows:
        by_var.setdefault(r.variable_name, []).append(r)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for var_name, var_rows in sorted(by_var.items()):
            dims = list(_index_names_for(var_name))
            if not dims:
                dims = _fallback_dims_from_row(var_rows[0])

            header = list(dims) + ["VALUE"]
            records: list[dict[str, object]] = []
            for r in var_rows:
                rec: dict[str, object] = {}
                for dim in dims:
                    if dim == "YEAR":
                        rec["YEAR"] = r.year if r.year is not None else ""
                        continue
                    attr = _DIM_TO_ROW_ATTR.get(dim)
                    val = getattr(r, attr, None) if attr else None
                    if val is None:
                        if dim == "TECHNOLOGY":
                            val = r.technology_name
                        elif dim == "FUEL":
                            val = r.fuel_name
                        elif dim == "EMISSION":
                            val = r.emission_name
                    rec[dim] = val if val is not None else ""
                rec["VALUE"] = float(r.value)
                records.append(rec)

            df = pd.DataFrame(records, columns=header)
            csv_bytes = df.to_csv(index=False).encode("utf-8")
            zf.writestr(f"{var_name}.csv", csv_bytes)

    buffer.seek(0)
    return buffer
