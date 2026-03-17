"""
Exporta un escenario a Excel en formato SAND (hoja Parameters).

Estructura idéntica a la esperada por el importador: Parameter, dimensiones
(REGION, TECHNOLOGY, FUEL, etc.), Time indipendent variables (opcional),
y columnas por año. Permite descargar, editar en Excel y volver a subir.
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.simulation.core.data_processing import PARAM_INDEX, _resolved_query

# Cabeceras SAND para export (mismo orden que reconoce el importador).
# _resolved_query devuelve: param_name, region, technology, fuel, emission, timeslice, mode, season, daytype, dailytimebracket, storage, udc, year, value
SAND_DIMENSION_HEADERS = [
    "Parameter",
    "REGION",
    "TECHNOLOGY",
    "FUEL",
    "EMISSION",
    "TIMESLICE",
    "MODE_OF_OPERATION",
    "Storage",
    "Season",
    "Daytype",
    "Dailytimebracket",
    "UDC",
]
TIME_INDEPENDENT_HEADER = "Time indipendent variables"
RAW_HEADERS = [
    "Parameter",
    "REGION",
    "TECHNOLOGY",
    "FUEL",
    "EMISSION",
    "TIMESLICE",
    "MODE_OF_OPERATION",
    "Storage",
    "Season",
    "Daytype",
    "Dailytimebracket",
    "UDC",
    "YEAR",
    "VALUE",
]


def _row_to_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def export_scenario_to_excel(db: Session, *, scenario_id: int, scenario_name: str) -> bytes:
    """
    Genera un Excel con una hoja "Parameters" en formato SAND a partir de
    osemosys_param_value del escenario dado.

    - Agrupa por (param_name, REGION, TECHNOLOGY, ...) sin YEAR.
    - Parámetros con año: columnas por año (1900–2200).
    - Parámetros sin año: valor en columna "Time indipendent variables".
    """
    result_proxy = db.execute(_resolved_query(), {"scenario_id": scenario_id})

    # key = (pname, region, technology, fuel, emission, timeslice, mode, season, daytype, dtb, storage, udc)
    # value = {year: value} o {None: value} para time-independent
    grouped: dict[tuple, dict[int | None, float]] = defaultdict(dict)
    years_used: set[int] = set()

    for row in result_proxy.yield_per(50_000):
        pname = row[0]
        if PARAM_INDEX.get(pname) is None:
            continue

        value = float(row[13])
        year_raw = row[12]
        year_val: int | None = int(year_raw) if year_raw is not None else None
        if year_val is not None:
            years_used.add(year_val)

        key = (pname,) + tuple(_row_to_str(row[i]) for i in range(1, 12))
        grouped[key][year_val] = value

    years_sorted = sorted(years_used) if years_used else []

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "Parameters"

    # Fila 1: cabeceras (dimensiones + Time indipendent variables + columnas de año)
    headers = list(SAND_DIMENSION_HEADERS) + [TIME_INDEPENDENT_HEADER]
    if years_sorted:
        headers.extend(str(y) for y in years_sorted)

    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    row_data_list = list(grouped.items())
    row_data_list.sort(key=lambda x: x[0])

    col_parameter = 1
    col_region = 2
    col_technology = 3
    col_fuel = 4
    col_emission = 5
    col_timeslice = 6
    col_mode = 7
    col_storage = 8
    col_season = 9
    col_daytype = 10
    col_dtb = 11
    col_udc = 12
    col_time_indep = 13
    first_year_col = 14

    for excel_row, (key, year_to_val) in enumerate(row_data_list, start=2):
        pname = key[0]
        dim_vals = list(key[1:])
        ws.cell(row=excel_row, column=col_parameter, value=pname)
        for c, v in enumerate(dim_vals, start=col_region):
            ws.cell(row=excel_row, column=c, value=v or None)

        ti = year_to_val.get(None)
        if ti is not None:
            ws.cell(row=excel_row, column=col_time_indep, value=ti)
        if years_sorted:
            for i, yr in enumerate(years_sorted):
                val = year_to_val.get(yr)
                ws.cell(row=excel_row, column=first_year_col + i, value=val)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def export_scenario_raw_to_excel(db: Session, *, scenario_id: int, scenario_name: str) -> bytes:
    """Genera un Excel RAW (1 fila por registro) de `osemosys_param_value`.

    A diferencia del formato SAND, no agrupa por dimensiones y no filtra por PARAM_INDEX.
    """
    result_proxy = db.execute(_resolved_query(), {"scenario_id": scenario_id})

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active sheet")
    ws.title = "RawParameters"

    for col, h in enumerate(RAW_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)

    for excel_row, row in enumerate(result_proxy.yield_per(50_000), start=2):
        ws.cell(row=excel_row, column=1, value=_row_to_str(row[0]) or None)
        for idx, col in enumerate(range(1, 12), start=2):
            ws.cell(row=excel_row, column=idx, value=_row_to_str(row[col]) or None)
        ws.cell(row=excel_row, column=13, value=int(row[12]) if row[12] is not None else None)
        ws.cell(row=excel_row, column=14, value=float(row[13]) if row[13] is not None else None)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
