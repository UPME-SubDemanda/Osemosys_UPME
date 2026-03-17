"""
Servicio de importación de datos oficiales desde Excel (.xlsm).

Lee hojas SAND del workbook, mapea columnas a dimensiones OSeMOSYS
(region, technology, fuel, emission, timeslice, mode, season, daytype, dtb, storage, udc, year)
y realiza upsert en la base de datos. Soporta catálogos por nombre (Parameter, Region, Technology,
Fuel, Emission, Solver) y por código (Timeslice, ModeOfOperation, Season, Daytype, Dailytimebracket,
StorageSet, UdcSet). Tras la importación ejecuta preprocesamiento tipo notebook para paridad
con el flujo Jupyter UPME.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import and_, bindparam, delete, insert, select, update as sa_update
from sqlalchemy.orm import Session

from app.models import (
    ChangeRequest,
    ChangeRequestValue,
    Daytype,
    Dailytimebracket,
    Emission,
    Fuel,
    ModeOfOperation,
    OsemosysParamValue,
    Parameter,
    ParameterStorage,
    ParameterValue,
    ParameterValueAudit,
    Region,
    Scenario,
    ScenarioPermission,
    Season,
    Solver,
    StorageSet,
    Technology,
    Timeslice,
    UdcSet,
)
from app.services.sand_notebook_preprocess import run_notebook_preprocess


def _normalize_key(value: str) -> str:
    """Normaliza un string para comparación: quita acentos, minúsculas, solo alfanuméricos y '_'."""
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return "".join(ch for ch in ascii_only.lower().strip() if ch.isalnum() or ch == "_")


def _clean_str(value: object) -> str | None:
    """Convierte valor a string limpio (strip) o None si está vacío."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _dimension_str(value: object) -> str | None:
    """Convierte valor de dimensión a string replicando pandas read_excel.

    - Números → str(float(v)):  int 1 → "1.0"  (pandas lee como float64 cuando hay NaN)
    - Strings → se aplica strip() para eliminar newlines/espacios accidentales del SAND.
    - Vacío/None → None
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(float(value))
    text = str(value).strip()
    if not text:
        return None
    return text


def _to_int(value: object) -> int | None:
    """Parsea valor a entero; acepta coma como separador decimal. Retorna None si no es parseable."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _to_float(value: object) -> float | None:
    """Parsea valor a float; acepta coma como separador decimal. Retorna None si no es parseable."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _to_float_or_zero(value: object) -> float:
    """Parsea a float; retorna 0.0 si el valor es None o no parseable."""
    parsed = _to_float(value)
    return 0.0 if parsed is None else float(parsed)


# Valores por defecto para parámetros OSeMOSYS cuando el Excel no trae valor explícito.
OSEMOSYS_PARAM_DEFAULTS: dict[str, float] = {
    "yearsplit": 0.0,
    "discountrate": 0.05,
    "discountrateidv": 0.05,
    "operationallife": 1.0,
    "depreciationmethod": 1.0,
    "accumulatedannualdemand": 0.0,
    "specifiedannualdemand": 0.0,
    "specifieddemandprofile": 0.0,
    "capacitytoactivityunit": 1.0,
    "capacityfactor": 1.0,
    "availabilityfactor": 1.0,
    "residualcapacity": 0.0,
    "inputactivityratio": 0.0,
    "outputactivityratio": 0.0,
    "capitalcost": 0.000001,
    "variablecost": 0.000001,
    "fixedcost": 0.0,
    "capacityofonetechnologyunit": 0.0,
    "totalannualmaxcapacity": 9999999.0,
    "totalannualmincapacity": 0.0,
    "totalannualmaxcapacityinvestment": 9999999.0,
    "totalannualmincapacityinvestment": 0.0,
    "totaltechnologyannualactivityupperlimit": 9999999.0,
    "totaltechnologyannualactivitylowerlimit": 0.0,
    "totaltechnologymodelperiodactivityupperlimit": 9999999.0,
    "totaltechnologymodelperiodactivitylowerlimit": 0.0,
    "reservemargintagtechnology": 0.0,
    "reservemargintagfuel": 0.0,
    "reservemargin": 1.0,
    "retagtechnology": 0.0,
    "retagfuel": 0.0,
    "reminproductiontarget": 0.0,
    "emissionactivityratio": 0.0,
    "emissionspenalty": 0.0,
    "annualexogenousemission": 0.0,
    "annualemissionlimit": 9999999.0,
    "modelperiodexogenousemission": 0.0,
    "modelperiodemissionlimit": 9999999.0,
    "inputtonewcapacityratio": 0.0,
    "inputtototalcapacityratio": 0.0,
    "technologyactivitybymodelowerlimit": 0.0,
    "technologyactivitybymodeupperlimit": 0.0,
    "technologyactivitydecreasebymodeLimit": 0.0,
    "technologyactivityincreasebymodeLimit": 0.0,
    "emissiontoactivitychangeratio": 0.0,
    "udcmultipliertotalcapacity": 0.0,
    "udcmultipliernewcapacity": 0.0,
    "udcmultiplieractivity": 0.0,
    "udcconstant": 0.0,
    "udctag": 2.0,
    "daysplit": 0.00137,
    "conversionls": 0.0,
    "conversionld": 0.0,
    "conversionlh": 0.0,
    "daysindaytype": 7.0,
    "technologytostorage": 0.0,
    "technologyfromstorage": 0.0,
    "storagelevelstart": 0.0000001,
    "storagemaxchargerate": 9999999.0,
    "storagemaxdischargerate": 9999999.0,
    "minstoragecharge": 0.0,
    "operationallifestorage": 0.0,
    "capitalcoststorage": 0.0,
    "residualstoragecapacity": 0.0,
    "disposalcostpercapacity": 0.0,
    "recoveryvaluepercapacity": 0.0,
}


def _get_param_default(param_name: str | None) -> float:
    """Retorna el valor por defecto del parámetro OSeMOSYS por nombre normalizado."""
    if not param_name:
        return 0.0
    normalized = "".join(ch for ch in param_name.strip().lower() if ch.isalnum())
    return OSEMOSYS_PARAM_DEFAULTS.get(normalized, 0.0)


def _to_float_with_param_default(value: object, param_name: str | None) -> float:
    """Parsea a float; si falla, usa el valor por defecto del parámetro OSeMOSYS."""
    parsed = _to_float(value)
    if parsed is not None:
        return float(parsed)
    return _get_param_default(param_name)


def _to_bool(value: object) -> bool:
    """Parsea valor a booleano; acepta 1/0, true/false, si/no, etc."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "t", "yes", "y", "si", "sí"}:
        return True
    if text in {"0", "false", "f", "no", "n"}:
        return False
    return False


@dataclass
class ImportStats:
    """Estadísticas de una importación: filas leídas, insertadas, actualizadas, omitidas y advertencias."""
    total_rows_read: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    warnings: list[str] | None = None

    def __post_init__(self) -> None:
        if self.warnings is None:
            self.warnings = []

    def warn(self, message: str) -> None:
        if len(self.warnings) < 200:
            self.warnings.append(message)


def _sheet_by_alias(workbook, aliases: set[str]):
    """Busca la primera hoja cuyo nombre normalizado coincida con alguno de los alias."""
    alias_set = {_normalize_key(a) for a in aliases}
    for sheet_name in workbook.sheetnames:
        if _normalize_key(sheet_name) in alias_set:
            return workbook[sheet_name]
    return None


def _iter_rows(sheet):
    """Itera filas de la hoja como dicts {header_normalizado: valor}; detecta fila de encabezado."""
    rows = sheet.iter_rows(values_only=True)
    header_row = None
    for row in rows:
        if row and any(cell is not None and str(cell).strip() != "" for cell in row):
            header_row = row
            break
    if header_row is None:
        return
    headers = [_normalize_key(str(cell or "")) for cell in header_row]
    for row in rows:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        payload: dict[str, object] = {}
        for idx, cell in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
            payload[key] = cell
        yield payload


def _get_or_create_name_catalog(
    db: Session,
    *,
    model,
    row: dict[str, object],
    stats: ImportStats,
    section: str,
) -> None:
    """Crea o reactiva catálogo por nombre (Parameter, Region, Technology, etc.) desde fila Excel."""
    name = _clean_str(row.get("name") or row.get("nombre") or row.get("value") or row.get("valor"))
    if not name:
        non_empty_values = [str(v).strip() for v in row.values() if v is not None and str(v).strip()]
        if len(non_empty_values) == 1:
            name = non_empty_values[0]
    if not name:
        stats.skipped += 1
        stats.warn(f"[{section}] fila omitida por nombre vacío.")
        return
    existing = db.execute(select(model).where(model.name == name)).scalar_one_or_none()
    if existing:
        if hasattr(existing, "is_active") and not existing.is_active:
            existing.is_active = True
            stats.updated += 1
        else:
            stats.skipped += 1
        return
    db.add(model(name=name))
    stats.inserted += 1


def _get_or_create_code_catalog(
    db: Session,
    *,
    model,
    row: dict[str, object],
    stats: ImportStats,
    section: str,
) -> None:
    """Crea o actualiza catálogo por código (Timeslice, ModeOfOperation, etc.) desde fila Excel."""
    code = _clean_str(row.get("code") or row.get("codigo") or row.get("cod"))
    if not code:
        stats.skipped += 1
        stats.warn(f"[{section}] fila omitida por código vacío.")
        return
    description = _clean_str(row.get("description") or row.get("descripcion"))
    existing = db.execute(select(model).where(model.code == code)).scalar_one_or_none()
    if existing:
        if description and existing.description != description:
            existing.description = description
            stats.updated += 1
        else:
            stats.skipped += 1
        return
    db.add(model(code=code, description=description))
    stats.inserted += 1


def _load_name_map(db: Session, model) -> dict[str, int]:
    """Carga mapa nombre -> id para un modelo de catálogo por nombre."""
    rows = db.execute(select(model)).scalars().all()
    return {item.name: item.id for item in rows}


def _load_code_map(db: Session, model) -> dict[str, int]:
    """Carga mapa código -> id para un modelo de catálogo por código."""
    rows = db.execute(select(model)).scalars().all()
    return {item.code: item.id for item in rows}


def _resolve_ref(
    row: dict[str, object],
    *,
    id_keys: list[str],
    name_keys: list[str],
    ref_map: dict[str, int],
) -> int | None:
    """Resuelve referencia: primero por id_keys (id_scenario, etc.), luego por name_keys y ref_map."""
    for key in id_keys:
        if key in row:
            parsed = _to_int(row.get(key))
            if parsed is not None:
                return parsed
    for key in name_keys:
        raw = _clean_str(row.get(key))
        if raw:
            return ref_map.get(raw)
    return None


def _is_sheet_selected(sheet_name: str, selected_sheet_name: str | None) -> bool:
    """Retorna True si no hay hoja seleccionada o si el nombre de la hoja coincide (normalizado)."""
    if not selected_sheet_name:
        return True
    return _normalize_key(sheet_name) == _normalize_key(selected_sheet_name)


def _get_or_create_default_scenario(db: Session, *, owner: str) -> Scenario:
    """Obtiene o crea el escenario 'Escenario de defecto' para carga oficial."""
    scenario = db.execute(
        select(Scenario).where(Scenario.name == "Escenario de defecto", Scenario.is_template.is_(False))
    ).scalar_one_or_none()
    if scenario is None:
        scenario = Scenario(
            name="Escenario de defecto",
            description="Escenario base sincronizado desde carga oficial",
            owner=owner,
            edit_policy="OWNER_ONLY",
            is_template=False,
        )
        db.add(scenario)
        db.flush()
    perm = db.execute(
        select(ScenarioPermission).where(
            ScenarioPermission.id_scenario == scenario.id,
            ScenarioPermission.user_identifier == f"user:{owner}",
        )
    ).scalar_one_or_none()
    if perm is None:
        db.add(
            ScenarioPermission(
                id_scenario=scenario.id,
                user_identifier=f"user:{owner}",
                user_id=None,
                can_edit_direct=True,
                can_propose=True,
                can_manage_values=True,
            )
        )
    return scenario


def _reset_scenario_data(db: Session, *, scenario_id: int) -> None:
    """Elimina datos del escenario (OsemosysParamValue, ChangeRequest) y defaults globales (ParameterValue).

    Usa subconsultas SQL para evitar cargar millones de IDs en Python.
    """
    ov_subq = select(OsemosysParamValue.id).where(
        OsemosysParamValue.id_scenario == scenario_id
    )
    cr_subq = select(ChangeRequest.id).where(
        ChangeRequest.id_osemosys_param_value.in_(ov_subq)
    )

    db.execute(delete(ChangeRequestValue).where(
        ChangeRequestValue.id_change_request.in_(cr_subq)
    ))
    db.execute(delete(ChangeRequest).where(
        ChangeRequest.id_osemosys_param_value.in_(ov_subq)
    ))
    db.execute(delete(OsemosysParamValue).where(
        OsemosysParamValue.id_scenario == scenario_id
    ))

    db.execute(delete(ParameterValueAudit))
    db.execute(delete(ParameterStorage))
    db.execute(delete(ParameterValue))


BATCH_SIZE = 5000


def _bulk_insert_osemosys_params(
    db: Session, rows: list[dict], stats: "ImportStats"
) -> None:
    """Inserta un lote de OsemosysParamValue en una sola sentencia INSERT masiva.

    Ambos flujos de importación trabajan con datos frescos (escenario nuevo o
    ``replace_scenario_data=True``), por lo que no se esperan conflictos.

    Cada dict en ``rows`` debe tener las claves: id_scenario, param_name, value, year,
    id_region, id_technology, id_fuel, id_emission, id_timeslice, id_mode_of_operation,
    id_season, id_daytype, id_dailytimebracket, id_storage_set, id_udc_set.
    """
    if not rows:
        return
    db.execute(insert(OsemosysParamValue), rows)
    stats.inserted += len(rows)
    db.flush()


def _bulk_insert_parameter_values(
    db: Session, rows: list[dict], stats: "ImportStats"
) -> None:
    """Inserta un lote de ParameterValue en una sola sentencia INSERT masiva."""
    if not rows:
        return
    db.execute(insert(ParameterValue), rows)
    stats.inserted += len(rows)
    db.flush()


def _bulk_insert_parameter_values_with_storage(
    db: Session, rows: list[dict], stats: "ImportStats"
) -> None:
    """Inserta lote de defaults y su metadata storage (si existe)."""
    if not rows:
        return

    pv_objects: list[ParameterValue] = []
    storage_payloads: list[dict[str, int | None] | None] = []

    for row in rows:
        row_copy = dict(row)
        storage_payloads.append(row_copy.pop("_storage", None))
        pv_objects.append(ParameterValue(**row_copy))

    db.add_all(pv_objects)
    db.flush()

    storage_objects: list[ParameterStorage] = []
    for pv_obj, storage in zip(pv_objects, storage_payloads):
        if not storage:
            continue
        if all(v is None for v in storage.values()):
            continue
        storage_objects.append(
            ParameterStorage(
                id_parameter_value=int(pv_obj.id),
                timesline=storage.get("timesline"),
                daytype=storage.get("daytype"),
                season=storage.get("season"),
                dailytimebracket=storage.get("dailytimebracket"),
                id_storage_set=storage.get("id_storage_set"),
            )
        )

    if storage_objects:
        db.add_all(storage_objects)
        db.flush()

    stats.inserted += len(pv_objects)


def _upsert_parameter_storage_for_pv(
    db: Session,
    *,
    parameter_value_id: int,
    timesline: int | None,
    daytype: int | None,
    season: int | None,
    dailytimebracket: int | None,
    id_storage_set: int | None,
) -> None:
    existing = db.execute(
        select(ParameterStorage).where(ParameterStorage.id_parameter_value == parameter_value_id)
    ).scalar_one_or_none()

    has_any_value = any(
        v is not None for v in (timesline, daytype, season, dailytimebracket, id_storage_set)
    )
    if not has_any_value:
        if existing is not None:
            db.delete(existing)
        return

    if existing is None:
        db.add(
            ParameterStorage(
                id_parameter_value=parameter_value_id,
                timesline=timesline,
                daytype=daytype,
                season=season,
                dailytimebracket=dailytimebracket,
                id_storage_set=id_storage_set,
            )
        )
        return

    existing.timesline = timesline
    existing.daytype = daytype
    existing.season = season
    existing.dailytimebracket = dailytimebracket
    existing.id_storage_set = id_storage_set


def _upsert_osemosys_param(
    db: Session,
    *,
    scenario_id: int,
    param_name: str,
    value: float,
    year: int | None = None,
    id_region: int | None = None,
    id_technology: int | None = None,
    id_fuel: int | None = None,
    id_emission: int | None = None,
    id_timeslice: int | None = None,
    id_mode_of_operation: int | None = None,
    id_season: int | None = None,
    id_daytype: int | None = None,
    id_dailytimebracket: int | None = None,
    id_storage_set: int | None = None,
    id_udc_set: int | None = None,
) -> bool:
    """Inserta o actualiza un OsemosysParamValue; retorna True si se actualizó, False si se insertó."""
    existing = db.execute(
        select(OsemosysParamValue).where(
            and_(
                OsemosysParamValue.id_scenario == scenario_id,
                OsemosysParamValue.param_name == param_name,
                OsemosysParamValue.id_region.is_(None)
                if id_region is None
                else OsemosysParamValue.id_region == id_region,
                OsemosysParamValue.id_technology.is_(None)
                if id_technology is None
                else OsemosysParamValue.id_technology == id_technology,
                OsemosysParamValue.id_fuel.is_(None)
                if id_fuel is None
                else OsemosysParamValue.id_fuel == id_fuel,
                OsemosysParamValue.id_emission.is_(None)
                if id_emission is None
                else OsemosysParamValue.id_emission == id_emission,
                OsemosysParamValue.id_timeslice.is_(None)
                if id_timeslice is None
                else OsemosysParamValue.id_timeslice == id_timeslice,
                OsemosysParamValue.id_mode_of_operation.is_(None)
                if id_mode_of_operation is None
                else OsemosysParamValue.id_mode_of_operation == id_mode_of_operation,
                OsemosysParamValue.id_season.is_(None)
                if id_season is None
                else OsemosysParamValue.id_season == id_season,
                OsemosysParamValue.id_daytype.is_(None)
                if id_daytype is None
                else OsemosysParamValue.id_daytype == id_daytype,
                OsemosysParamValue.id_dailytimebracket.is_(None)
                if id_dailytimebracket is None
                else OsemosysParamValue.id_dailytimebracket == id_dailytimebracket,
                OsemosysParamValue.id_storage_set.is_(None)
                if id_storage_set is None
                else OsemosysParamValue.id_storage_set == id_storage_set,
                OsemosysParamValue.id_udc_set.is_(None)
                if id_udc_set is None
                else OsemosysParamValue.id_udc_set == id_udc_set,
                OsemosysParamValue.year.is_(None)
                if year is None
                else OsemosysParamValue.year == year,
            )
        )
    ).scalar_one_or_none()
    if existing:
        existing.value = float(value)
        return True
    db.add(
        OsemosysParamValue(
            id_scenario=scenario_id,
            param_name=param_name,
            id_region=id_region,
            id_technology=id_technology,
            id_fuel=id_fuel,
            id_emission=id_emission,
            id_timeslice=id_timeslice,
            id_mode_of_operation=id_mode_of_operation,
            id_season=id_season,
            id_daytype=id_daytype,
            id_dailytimebracket=id_dailytimebracket,
            id_storage_set=id_storage_set,
            id_udc_set=id_udc_set,
            year=year,
            value=float(value),
        )
    )
    return False


@dataclass
class ParsedRow:
    """Fila SAND parseada con dimensiones resueltas a IDs y valores por año."""
    param_name: str
    ids: dict[str, int | None]
    year_values: dict[int, float]
    time_indep_val: float | None
    timeslice_code: str | None
    group_key: tuple | None


def _parse_sand_rows(
    db: Session,
    *,
    sheet,
    stats: ImportStats,
    region_map: dict[str, int],
    tech_map: dict[str, int],
    fuel_map: dict[str, int],
    emission_map: dict[str, int],
    timeslice_map: dict[str, int],
    mode_map: dict[str, int],
    storage_map: dict[str, int],
    season_map: dict[str, int] | None = None,
    daytype_map: dict[str, int] | None = None,
    dtb_map: dict[str, int] | None = None,
    udc_map: dict[str, int] | None = None,
    create_missing_catalogs: bool = True,
):
    """Generador que parsea filas SAND (hoja Parameters) y yield ParsedRow.

    Si ``create_missing_catalogs`` es True (import), crea catálogos faltantes.
    Si es False (update), solo resuelve IDs existentes.

    Yields: ParsedRow por cada fila válida del sheet.
    """
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return

    headers = [_normalize_key(str(cell or "")) for cell in header_row]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    if "parameter" not in header_idx:
        return

    year_cols: list[tuple[int, int]] = []
    for idx, raw in enumerate(header_row):
        parsed_year = _to_int(raw)
        if parsed_year is not None and 1900 <= parsed_year <= 2200:
            year_cols.append((idx, parsed_year))

    has_timeslice_col = "timeslice" in header_idx
    time_independent_idx = header_idx.get("timeindipendentvariables")

    _resolve_name = _get_or_create_name_id if create_missing_catalogs else _lookup_name_id
    _resolve_code = _get_or_create_code_id if create_missing_catalogs else _lookup_code_id

    for row_values in rows_iter:
        if not row_values or all(v is None or str(v).strip() == "" for v in row_values):
            continue

        def _cell(name: str, _rv=row_values):
            idx = header_idx.get(name)
            return _rv[idx] if idx is not None and idx < len(_rv) else None

        param_name = _clean_str(_cell("parameter"))
        if not param_name:
            stats.total_rows_read += 1
            stats.skipped += 1
            continue

        region_name = _dimension_str(_cell("region"))
        technology_name = _dimension_str(_cell("technology"))
        fuel_name = _dimension_str(_cell("fuel"))
        emission_name = _dimension_str(_cell("emission"))
        timeslice_code = _dimension_str(_cell("timeslice"))
        mode_code = _dimension_str(_cell("mode_of_operation"))
        storage_code = _dimension_str(_cell("storage"))
        season_code = _dimension_str(_cell("season"))
        daytype_code = _dimension_str(_cell("daytype"))
        dtb_code = _dimension_str(_cell("dailytimebracket"))
        udc_code = _dimension_str(_cell("udc") or _cell("udc_set"))

        id_region = _resolve_name(db, model=Region, name=region_name, ref_map=region_map)
        id_technology = _resolve_name(db, model=Technology, name=technology_name, ref_map=tech_map)
        id_fuel = _resolve_name(db, model=Fuel, name=fuel_name, ref_map=fuel_map)
        id_emission = _resolve_name(db, model=Emission, name=emission_name, ref_map=emission_map)
        id_timeslice = _resolve_code(db, model=Timeslice, code=timeslice_code, ref_map=timeslice_map)
        id_mode = _resolve_code(db, model=ModeOfOperation, code=mode_code, ref_map=mode_map)
        id_storage = _resolve_code(db, model=StorageSet, code=storage_code, ref_map=storage_map)
        id_season = (
            _resolve_code(db, model=Season, code=season_code, ref_map=season_map)
            if season_map is not None and season_code else None
        )
        id_daytype = (
            _resolve_code(db, model=Daytype, code=daytype_code, ref_map=daytype_map)
            if daytype_map is not None and daytype_code else None
        )
        id_dtb = (
            _resolve_code(db, model=Dailytimebracket, code=dtb_code, ref_map=dtb_map)
            if dtb_map is not None and dtb_code else None
        )
        id_udc = (
            _resolve_code(db, model=UdcSet, code=udc_code, ref_map=udc_map)
            if udc_map is not None and udc_code else None
        )

        ids = dict(
            id_region=id_region,
            id_technology=id_technology,
            id_fuel=id_fuel,
            id_emission=id_emission,
            id_timeslice=id_timeslice,
            id_mode_of_operation=id_mode,
            id_season=id_season,
            id_daytype=id_daytype,
            id_dailytimebracket=id_dtb,
            id_storage_set=id_storage,
            id_udc_set=id_udc,
        )

        year_values: dict[int, float] = {}
        time_indep_val: float | None = None

        if time_independent_idx is not None and time_independent_idx < len(row_values):
            raw_ti = _to_float(row_values[time_independent_idx])
            if raw_ti is not None:
                time_indep_val = float(raw_ti)

        if time_indep_val is None:
            for col_idx, yr in year_cols:
                if col_idx < len(row_values):
                    year_values[yr] = _to_float_with_param_default(row_values[col_idx], param_name)

        group_key: tuple | None = None
        if has_timeslice_col and timeslice_code:
            group_key = (
                _normalize_key(param_name), region_name, technology_name,
                fuel_name, emission_name, mode_code, storage_code,
                season_code, daytype_code, dtb_code, udc_code,
            )

        yield ParsedRow(
            param_name=param_name,
            ids=ids,
            year_values=year_values,
            time_indep_val=time_indep_val,
            timeslice_code=timeslice_code,
            group_key=group_key,
        )


def _lookup_name_id(
    db: Session,
    *,
    model,
    name: str | None,
    ref_map: dict[str, int],
) -> int | None:
    """Busca id por nombre en ref_map sin crear registros faltantes."""
    if not name:
        return None
    return ref_map.get(name)


def _lookup_code_id(
    db: Session,
    *,
    model,
    code: str | None,
    ref_map: dict[str, int],
) -> int | None:
    """Busca id por código en ref_map sin crear registros faltantes."""
    if not code:
        return None
    return ref_map.get(code)


def _import_sand_matrix_sheet(
    db: Session,
    *,
    sheet,
    stats: ImportStats,
    imported_by: str,
    fallback_scenario: Scenario | None,
    region_map: dict[str, int],
    tech_map: dict[str, int],
    fuel_map: dict[str, int],
    emission_map: dict[str, int],
    timeslice_map: dict[str, int],
    mode_map: dict[str, int],
    storage_map: dict[str, int],
    season_map: dict[str, int] | None = None,
    daytype_map: dict[str, int] | None = None,
    dtb_map: dict[str, int] | None = None,
    udc_map: dict[str, int] | None = None,
) -> Scenario | None:
    """Importa hoja SAND (formato Parameters) usando inserción por lotes.

    Siempre agrega todos los timeslices a 1 solo: promedia para
    CapacityFactor y suma para los demás parámetros.
    """
    if fallback_scenario is None:
        fallback_scenario = Scenario(
            name="EscenarioImportado",
            description="Escenario creado por importación oficial",
            owner=imported_by,
            edit_policy="OWNER_ONLY",
            is_template=False,
        )
        db.add(fallback_scenario)
        db.flush()

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return fallback_scenario

    headers = [_normalize_key(str(cell or "")) for cell in header_row]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    if "parameter" not in header_idx:
        return fallback_scenario

    year_cols: list[tuple[int, int]] = []
    for idx, raw in enumerate(header_row):
        parsed_year = _to_int(raw)
        if parsed_year is not None and 1900 <= parsed_year <= 2200:
            year_cols.append((idx, parsed_year))

    has_timeslice_col = "timeslice" in header_idx
    need_aggregation = has_timeslice_col

    pending_batch: list[dict] = []

    def _flush_batch() -> None:
        if not pending_batch:
            return
        _bulk_insert_osemosys_params(db, pending_batch, stats)
        pending_batch.clear()

    def _add_to_batch(param_name: str, ids: dict, value: float, year: int | None) -> None:
        pending_batch.append({
            "id_scenario": fallback_scenario.id,
            "param_name": param_name,
            "value": value,
            "year": year,
            "id_region": ids.get("id_region"),
            "id_technology": ids.get("id_technology"),
            "id_fuel": ids.get("id_fuel"),
            "id_emission": ids.get("id_emission"),
            "id_timeslice": ids.get("id_timeslice"),
            "id_mode_of_operation": ids.get("id_mode_of_operation"),
            "id_season": ids.get("id_season"),
            "id_daytype": ids.get("id_daytype"),
            "id_dailytimebracket": ids.get("id_dailytimebracket"),
            "id_storage_set": ids.get("id_storage_set"),
            "id_udc_set": ids.get("id_udc_set"),
        })
        if len(pending_batch) >= BATCH_SIZE:
            _flush_batch()

    _AggKey = tuple
    _agg_buffer: dict[_AggKey, list[dict]] = {}

    def _flush_agg_group(group_key: _AggKey, rows_in_group: list[dict]) -> None:
        if not rows_in_group:
            return
        param_name = rows_in_group[0]["param_name"]
        use_mean = _normalize_key(param_name) == "capacityfactor"
        n = len(rows_in_group)
        representative = rows_in_group[0]

        aggregated_years: dict[int, float] = {}
        for _yi, yr in year_cols:
            total = sum(r["year_values"].get(yr, 0.0) for r in rows_in_group)
            aggregated_years[yr] = total / n if use_mean else total

        if not use_mean and all(abs(v) < 1e-15 for v in aggregated_years.values()):
            stats.skipped += n
            stats.total_rows_read += n
            return

        ids = representative["ids"]
        wrote_any = False
        if representative.get("time_indep_val") is not None:
            _add_to_batch(param_name, ids, representative["time_indep_val"], None)
            wrote_any = True
        else:
            for yr, val in aggregated_years.items():
                _add_to_batch(param_name, ids, val, yr)
                wrote_any = True

        if not wrote_any:
            stats.skipped += n
        stats.total_rows_read += n

    for parsed in _parse_sand_rows(
        db,
        sheet=sheet,
        stats=stats,
        region_map=region_map,
        tech_map=tech_map,
        fuel_map=fuel_map,
        emission_map=emission_map,
        timeslice_map=timeslice_map,
        mode_map=mode_map,
        storage_map=storage_map,
        season_map=season_map,
        daytype_map=daytype_map,
        dtb_map=dtb_map,
        udc_map=udc_map,
        create_missing_catalogs=True,
    ):
        if parsed.group_key is not None:
            row_entry = {
                "param_name": parsed.param_name,
                "timeslice_code": parsed.timeslice_code,
                "year_values": parsed.year_values,
                "time_indep_val": parsed.time_indep_val,
                "ids": parsed.ids,
            }
            if parsed.group_key not in _agg_buffer:
                _agg_buffer[parsed.group_key] = []
            _agg_buffer[parsed.group_key].append(row_entry)
        else:
            stats.total_rows_read += 1
            wrote_any = False
            if parsed.time_indep_val is not None:
                _add_to_batch(parsed.param_name, parsed.ids, parsed.time_indep_val, None)
                wrote_any = True
            else:
                for yr, val in parsed.year_values.items():
                    _add_to_batch(parsed.param_name, parsed.ids, val, yr)
                    wrote_any = True
            if not wrote_any:
                stats.skipped += 1

    for group_key, group_rows in _agg_buffer.items():
        _flush_agg_group(group_key, group_rows)

    _flush_batch()

    return fallback_scenario


def _update_sand_matrix_sheet(
    db: Session,
    *,
    sheet,
    stats: ImportStats,
    scenario_id: int,
    region_map: dict[str, int],
    tech_map: dict[str, int],
    fuel_map: dict[str, int],
    emission_map: dict[str, int],
    timeslice_map: dict[str, int],
    mode_map: dict[str, int],
    storage_map: dict[str, int],
    season_map: dict[str, int] | None = None,
    daytype_map: dict[str, int] | None = None,
    dtb_map: dict[str, int] | None = None,
    udc_map: dict[str, int] | None = None,
) -> dict:
    """Actualiza valores existentes en un escenario a partir de una hoja SAND.

    Aplica la misma lógica de parsing y agregación de timeslices que la
    importación, pero en lugar de insertar ejecuta UPDATE-only. Los registros
    no encontrados se registran como advertencias.

    Retorna dict con contadores ``updated`` y ``not_found``.
    """
    existing_rows = db.execute(
        select(
            OsemosysParamValue.id,
            OsemosysParamValue.param_name,
            OsemosysParamValue.id_region,
            OsemosysParamValue.id_technology,
            OsemosysParamValue.id_fuel,
            OsemosysParamValue.id_emission,
            OsemosysParamValue.id_timeslice,
            OsemosysParamValue.id_mode_of_operation,
            OsemosysParamValue.id_season,
            OsemosysParamValue.id_daytype,
            OsemosysParamValue.id_dailytimebracket,
            OsemosysParamValue.id_storage_set,
            OsemosysParamValue.id_udc_set,
            OsemosysParamValue.year,
        ).where(OsemosysParamValue.id_scenario == scenario_id)
    ).all()

    def _none_key(v):
        return v if v is not None else None

    existing_index: dict[tuple, int] = {}
    for row in existing_rows:
        key = (
            row.param_name,
            _none_key(row.id_region),
            _none_key(row.id_technology),
            _none_key(row.id_fuel),
            _none_key(row.id_emission),
            _none_key(row.id_timeslice),
            _none_key(row.id_mode_of_operation),
            _none_key(row.id_season),
            _none_key(row.id_daytype),
            _none_key(row.id_dailytimebracket),
            _none_key(row.id_storage_set),
            _none_key(row.id_udc_set),
            _none_key(row.year),
        )
        existing_index[key] = row.id

    _AggKey = tuple
    _agg_buffer: dict[_AggKey, list[dict]] = {}

    rows_iter = sheet.iter_rows(values_only=True)
    header_row_check = next(rows_iter, None)
    if not header_row_check:
        return {"updated": 0, "not_found": 0}

    year_cols: list[tuple[int, int]] = []
    for idx_c, raw_c in enumerate(header_row_check):
        parsed_yr = _to_int(raw_c)
        if parsed_yr is not None and 1900 <= parsed_yr <= 2200:
            year_cols.append((idx_c, parsed_yr))

    update_count = 0
    not_found_count = 0
    pending_updates: list[dict] = []

    def _flush_updates() -> None:
        if not pending_updates:
            return
        tbl = OsemosysParamValue.__table__
        stmt = (
            sa_update(tbl)
            .where(tbl.c.id == bindparam("_id"))
            .values(value=bindparam("value"))
        )
        db.execute(stmt, pending_updates)
        pending_updates.clear()

    def _try_update(param_name: str, ids: dict, value: float, year: int | None) -> None:
        nonlocal update_count, not_found_count
        key = (
            param_name,
            ids.get("id_region"),
            ids.get("id_technology"),
            ids.get("id_fuel"),
            ids.get("id_emission"),
            ids.get("id_timeslice"),
            ids.get("id_mode_of_operation"),
            ids.get("id_season"),
            ids.get("id_daytype"),
            ids.get("id_dailytimebracket"),
            ids.get("id_storage_set"),
            ids.get("id_udc_set"),
            year,
        )
        row_id = existing_index.get(key)
        if row_id is not None:
            pending_updates.append({"_id": row_id, "value": float(value)})
            update_count += 1
            if len(pending_updates) >= BATCH_SIZE:
                _flush_updates()
        else:
            not_found_count += 1
            dims = f"param={param_name}, year={year}"
            for dim_label, dim_key in [
                ("region", "id_region"), ("tech", "id_technology"),
                ("fuel", "id_fuel"), ("emission", "id_emission"),
            ]:
                v = ids.get(dim_key)
                if v is not None:
                    dims += f", {dim_label}={v}"
            stats.warn(f"[update] No encontrado: {dims}")

    def _flush_agg_group_update(group_key: _AggKey, rows_in_group: list[dict]) -> None:
        if not rows_in_group:
            return
        param_name = rows_in_group[0]["param_name"]
        use_mean = _normalize_key(param_name) == "capacityfactor"
        n = len(rows_in_group)
        representative = rows_in_group[0]

        aggregated_years: dict[int, float] = {}
        for _yi, yr in year_cols:
            total = sum(r["year_values"].get(yr, 0.0) for r in rows_in_group)
            aggregated_years[yr] = total / n if use_mean else total

        if not use_mean and all(abs(v) < 1e-15 for v in aggregated_years.values()):
            stats.skipped += n
            stats.total_rows_read += n
            return

        ids = representative["ids"]
        stats.total_rows_read += n
        if representative.get("time_indep_val") is not None:
            _try_update(param_name, ids, representative["time_indep_val"], None)
        else:
            for yr, val in aggregated_years.items():
                _try_update(param_name, ids, val, yr)

    for parsed in _parse_sand_rows(
        db,
        sheet=sheet,
        stats=stats,
        region_map=region_map,
        tech_map=tech_map,
        fuel_map=fuel_map,
        emission_map=emission_map,
        timeslice_map=timeslice_map,
        mode_map=mode_map,
        storage_map=storage_map,
        season_map=season_map,
        daytype_map=daytype_map,
        dtb_map=dtb_map,
        udc_map=udc_map,
        create_missing_catalogs=False,
    ):
        if parsed.group_key is not None:
            row_entry = {
                "param_name": parsed.param_name,
                "timeslice_code": parsed.timeslice_code,
                "year_values": parsed.year_values,
                "time_indep_val": parsed.time_indep_val,
                "ids": parsed.ids,
            }
            if parsed.group_key not in _agg_buffer:
                _agg_buffer[parsed.group_key] = []
            _agg_buffer[parsed.group_key].append(row_entry)
        else:
            stats.total_rows_read += 1
            if parsed.time_indep_val is not None:
                _try_update(parsed.param_name, parsed.ids, parsed.time_indep_val, None)
            else:
                for yr, val in parsed.year_values.items():
                    _try_update(parsed.param_name, parsed.ids, val, yr)

    for gk, group_rows in _agg_buffer.items():
        _flush_agg_group_update(gk, group_rows)

    _flush_updates()

    return {"updated": update_count, "not_found": not_found_count}


def _preview_sand_matrix_sheet(
    db: Session,
    *,
    sheet,
    stats: ImportStats,
    scenario_id: int,
    region_map: dict[str, int],
    tech_map: dict[str, int],
    fuel_map: dict[str, int],
    emission_map: dict[str, int],
    timeslice_map: dict[str, int],
    mode_map: dict[str, int],
    storage_map: dict[str, int],
    season_map: dict[str, int] | None = None,
    daytype_map: dict[str, int] | None = None,
    dtb_map: dict[str, int] | None = None,
    udc_map: dict[str, int] | None = None,
) -> dict:
    """Genera preview de cambios sin modificar la base de datos.

    Misma logica de parsing/agregacion que ``_update_sand_matrix_sheet``
    pero recopila diffs ``(row_id, old_value, new_value, nombres)``
    en vez de ejecutar UPDATEs.
    """
    existing_rows = db.execute(
        select(
            OsemosysParamValue.id,
            OsemosysParamValue.param_name,
            OsemosysParamValue.id_region,
            OsemosysParamValue.id_technology,
            OsemosysParamValue.id_fuel,
            OsemosysParamValue.id_emission,
            OsemosysParamValue.id_timeslice,
            OsemosysParamValue.id_mode_of_operation,
            OsemosysParamValue.id_season,
            OsemosysParamValue.id_daytype,
            OsemosysParamValue.id_dailytimebracket,
            OsemosysParamValue.id_storage_set,
            OsemosysParamValue.id_udc_set,
            OsemosysParamValue.year,
            OsemosysParamValue.value,
        ).where(OsemosysParamValue.id_scenario == scenario_id)
    ).all()

    existing_index: dict[tuple, tuple[int, float]] = {}
    for row in existing_rows:
        key = (
            row.param_name,
            row.id_region,
            row.id_technology,
            row.id_fuel,
            row.id_emission,
            row.id_timeslice,
            row.id_mode_of_operation,
            row.id_season,
            row.id_daytype,
            row.id_dailytimebracket,
            row.id_storage_set,
            row.id_udc_set,
            row.year,
        )
        existing_index[key] = (row.id, float(row.value))

    rev_region = {v: k for k, v in region_map.items()}
    rev_tech = {v: k for k, v in tech_map.items()}
    rev_fuel = {v: k for k, v in fuel_map.items()}
    rev_emission = {v: k for k, v in emission_map.items()}

    _AggKey = tuple
    _agg_buffer: dict[_AggKey, list[dict]] = {}

    rows_iter = sheet.iter_rows(values_only=True)
    header_row_check = next(rows_iter, None)
    if not header_row_check:
        return {"changes": [], "not_found": 0}

    year_cols: list[tuple[int, int]] = []
    for idx_c, raw_c in enumerate(header_row_check):
        parsed_yr = _to_int(raw_c)
        if parsed_yr is not None and 1900 <= parsed_yr <= 2200:
            year_cols.append((idx_c, parsed_yr))

    not_found_count = 0
    changes: list[dict] = []

    def _try_preview(param_name: str, ids: dict, value: float, year: int | None) -> None:
        nonlocal not_found_count
        key = (
            param_name,
            ids.get("id_region"),
            ids.get("id_technology"),
            ids.get("id_fuel"),
            ids.get("id_emission"),
            ids.get("id_timeslice"),
            ids.get("id_mode_of_operation"),
            ids.get("id_season"),
            ids.get("id_daytype"),
            ids.get("id_dailytimebracket"),
            ids.get("id_storage_set"),
            ids.get("id_udc_set"),
            year,
        )
        match = existing_index.get(key)
        if match is not None:
            row_id, old_value = match
            new_val = float(value)
            # Solo incluir en el preview filas donde el valor realmente cambió (evita saturar UI con miles de filas iguales)
            if abs(old_value - new_val) > 1e-12:
                changes.append({
                    "row_id": row_id,
                    "param_name": param_name,
                    "region_name": rev_region.get(ids.get("id_region")) if ids.get("id_region") else None,
                    "technology_name": rev_tech.get(ids.get("id_technology")) if ids.get("id_technology") else None,
                    "fuel_name": rev_fuel.get(ids.get("id_fuel")) if ids.get("id_fuel") else None,
                    "emission_name": rev_emission.get(ids.get("id_emission")) if ids.get("id_emission") else None,
                    "year": year,
                    "old_value": old_value,
                    "new_value": new_val,
                })
        else:
            not_found_count += 1
            dims = f"param={param_name}, year={year}"
            for dim_label, dim_key in [
                ("region", "id_region"), ("tech", "id_technology"),
                ("fuel", "id_fuel"), ("emission", "id_emission"),
            ]:
                v = ids.get(dim_key)
                if v is not None:
                    dims += f", {dim_label}={v}"
            stats.warn(f"[update] No encontrado: {dims}")

    def _flush_agg_group_preview(group_key: _AggKey, rows_in_group: list[dict]) -> None:
        if not rows_in_group:
            return
        param_name = rows_in_group[0]["param_name"]
        use_mean = _normalize_key(param_name) == "capacityfactor"
        n = len(rows_in_group)
        representative = rows_in_group[0]

        aggregated_years: dict[int, float] = {}
        for _yi, yr in year_cols:
            total = sum(r["year_values"].get(yr, 0.0) for r in rows_in_group)
            aggregated_years[yr] = total / n if use_mean else total

        if not use_mean and all(abs(v) < 1e-15 for v in aggregated_years.values()):
            stats.skipped += n
            stats.total_rows_read += n
            return

        ids = representative["ids"]
        stats.total_rows_read += n
        if representative.get("time_indep_val") is not None:
            _try_preview(param_name, ids, representative["time_indep_val"], None)
        else:
            for yr, val in aggregated_years.items():
                _try_preview(param_name, ids, val, yr)

    for parsed in _parse_sand_rows(
        db,
        sheet=sheet,
        stats=stats,
        region_map=region_map,
        tech_map=tech_map,
        fuel_map=fuel_map,
        emission_map=emission_map,
        timeslice_map=timeslice_map,
        mode_map=mode_map,
        storage_map=storage_map,
        season_map=season_map,
        daytype_map=daytype_map,
        dtb_map=dtb_map,
        udc_map=udc_map,
        create_missing_catalogs=False,
    ):
        if parsed.group_key is not None:
            row_entry = {
                "param_name": parsed.param_name,
                "timeslice_code": parsed.timeslice_code,
                "year_values": parsed.year_values,
                "time_indep_val": parsed.time_indep_val,
                "ids": parsed.ids,
            }
            if parsed.group_key not in _agg_buffer:
                _agg_buffer[parsed.group_key] = []
            _agg_buffer[parsed.group_key].append(row_entry)
        else:
            stats.total_rows_read += 1
            if parsed.time_indep_val is not None:
                _try_preview(parsed.param_name, parsed.ids, parsed.time_indep_val, None)
            else:
                for yr, val in parsed.year_values.items():
                    _try_preview(parsed.param_name, parsed.ids, val, yr)
    for gk, group_rows in _agg_buffer.items():
        _flush_agg_group_preview(gk, group_rows)

    return {"changes": changes, "not_found": not_found_count}


def _import_sand_matrix_sheet_to_parameter_value(
    db: Session,
    *,
    sheet,
    stats: ImportStats,
    param_map: dict[str, int],
    region_map: dict[str, int],
    tech_map: dict[str, int],
    fuel_map: dict[str, int],
    emission_map: dict[str, int],
    timeslice_map: dict[str, int],
    season_map: dict[str, int],
    daytype_map: dict[str, int],
    dtb_map: dict[str, int],
    storage_map: dict[str, int],
    default_solver_id: int,
) -> None:
    """Importa hoja SAND (Parameters) a `parameter_value` en lotes.

    Reutiliza la lógica de agregación de timeslices:
    - CapacityFactor: promedio.
    - demás parámetros: suma.
    """
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return

    headers = [_normalize_key(str(cell or "")) for cell in header_row]
    header_idx = {name: idx for idx, name in enumerate(headers)}
    if "parameter" not in header_idx:
        return

    year_cols: list[tuple[int, int]] = []
    for idx, raw in enumerate(header_row):
        parsed_year = _to_int(raw)
        if parsed_year is not None and 1900 <= parsed_year <= 2200:
            year_cols.append((idx, parsed_year))

    has_timeslice_col = "timeslice" in header_idx
    need_aggregation = has_timeslice_col

    pending_batch: list[dict] = []

    def _flush_batch() -> None:
        if not pending_batch:
            return
        _bulk_insert_parameter_values_with_storage(db, pending_batch, stats)
        pending_batch.clear()

    def _add_to_batch(
        *,
        id_parameter: int,
        id_region: int,
        id_technology: int | None,
        id_fuel: int | None,
        id_emission: int | None,
        value: float,
        year: int,
        timesline: int | None,
        daytype: int | None,
        season: int | None,
        dailytimebracket: int | None,
        id_storage_set: int | None,
    ) -> None:
        pending_batch.append(
            {
                "id_parameter": id_parameter,
                "id_region": id_region,
                "id_technology": id_technology,
                "id_fuel": id_fuel,
                "id_emission": id_emission,
                "id_solver": default_solver_id,
                "mode_of_operation": False,
                "year": year,
                "value": float(value),
                "unit": None,
                "_storage": {
                    "timesline": timesline,
                    "daytype": daytype,
                    "season": season,
                    "dailytimebracket": dailytimebracket,
                    "id_storage_set": id_storage_set,
                },
            }
        )
        if len(pending_batch) >= BATCH_SIZE:
            _flush_batch()

    _AggKey = tuple
    _agg_buffer: dict[_AggKey, list[dict]] = {}

    def _flush_agg_group(rows_in_group: list[dict]) -> None:
        if not rows_in_group:
            return
        representative = rows_in_group[0]
        param_name = str(representative["param_name"])
        use_mean = _normalize_key(param_name) == "capacityfactor"
        n = len(rows_in_group)

        aggregated_years: dict[int, float] = {}
        for _idx, yr in year_cols:
            total = sum(float(r["year_values"].get(yr, 0.0)) for r in rows_in_group)
            aggregated_years[yr] = total / n if use_mean else total

        if not use_mean and aggregated_years and all(abs(v) < 1e-15 for v in aggregated_years.values()):
            stats.skipped += n
            stats.total_rows_read += n
            return

        id_parameter = int(representative["id_parameter"])
        id_region = int(representative["id_region"])
        id_technology = representative["id_technology"]
        id_fuel = representative["id_fuel"]
        id_emission = representative["id_emission"]
        timesline = representative.get("timesline")
        daytype = representative.get("daytype")
        season = representative.get("season")
        dailytimebracket = representative.get("dailytimebracket")
        id_storage_set = representative.get("id_storage_set")

        wrote_any = False
        if representative.get("time_indep_val") is not None and year_cols:
            ti_value = float(representative["time_indep_val"])
            for _col_idx, year in year_cols:
                _add_to_batch(
                    id_parameter=id_parameter,
                    id_region=id_region,
                    id_technology=id_technology,
                    id_fuel=id_fuel,
                    id_emission=id_emission,
                    value=ti_value,
                    year=year,
                    timesline=timesline,
                    daytype=daytype,
                    season=season,
                    dailytimebracket=dailytimebracket,
                    id_storage_set=id_storage_set,
                )
                wrote_any = True
        else:
            for year, val in aggregated_years.items():
                _add_to_batch(
                    id_parameter=id_parameter,
                    id_region=id_region,
                    id_technology=id_technology,
                    id_fuel=id_fuel,
                    id_emission=id_emission,
                    value=float(val),
                    year=year,
                    timesline=timesline,
                    daytype=daytype,
                    season=season,
                    dailytimebracket=dailytimebracket,
                    id_storage_set=id_storage_set,
                )
                wrote_any = True

        if not wrote_any:
            stats.skipped += n
        stats.total_rows_read += n

    for row_values in rows_iter:
        if not row_values or all(v is None or str(v).strip() == "" for v in row_values):
            continue

        def _cell(name: str):
            idx = header_idx.get(name)
            return row_values[idx] if idx is not None and idx < len(row_values) else None

        param_name = _clean_str(_cell("parameter"))
        if not param_name:
            stats.total_rows_read += 1
            stats.skipped += 1
            continue

        parameter_id = _get_or_create_name_id(db, model=Parameter, name=param_name, ref_map=param_map)
        region_name = _dimension_str(_cell("region"))
        id_region = _get_or_create_name_id(db, model=Region, name=region_name, ref_map=region_map)
        if id_region is None:
            stats.total_rows_read += 1
            stats.skipped += 1
            stats.warn("[parameters->parameter_value] fila omitida por región vacía.")
            continue

        technology_name = _dimension_str(_cell("technology"))
        fuel_name = _dimension_str(_cell("fuel"))
        emission_name = _dimension_str(_cell("emission"))
        timeslice_code = _dimension_str(_cell("timeslice"))
        season_code = _dimension_str(_cell("season"))
        daytype_code = _dimension_str(_cell("daytype"))
        dtb_code = _dimension_str(_cell("dailytimebracket"))
        storage_code = _dimension_str(_cell("storage"))

        id_technology = _get_or_create_name_id(db, model=Technology, name=technology_name, ref_map=tech_map)
        id_fuel = _get_or_create_name_id(db, model=Fuel, name=fuel_name, ref_map=fuel_map)
        id_emission = _get_or_create_name_id(db, model=Emission, name=emission_name, ref_map=emission_map)
        id_timeslice = _get_or_create_code_id(db, model=Timeslice, code=timeslice_code, ref_map=timeslice_map)
        id_season = _get_or_create_code_id(db, model=Season, code=season_code, ref_map=season_map)
        id_daytype = _get_or_create_code_id(db, model=Daytype, code=daytype_code, ref_map=daytype_map)
        id_dtb = _get_or_create_code_id(db, model=Dailytimebracket, code=dtb_code, ref_map=dtb_map)
        id_storage = _get_or_create_code_id(db, model=StorageSet, code=storage_code, ref_map=storage_map)
        time_independent_idx = header_idx.get("timeindipendentvariables")

        if not need_aggregation or not timeslice_code:
            stats.total_rows_read += 1
            wrote_any = False
            if time_independent_idx is not None and time_independent_idx < len(row_values) and year_cols:
                raw_ti = _to_float(row_values[time_independent_idx])
                if raw_ti is not None:
                    for _col_idx, year in year_cols:
                        _add_to_batch(
                            id_parameter=int(parameter_id),
                            id_region=int(id_region),
                            id_technology=id_technology,
                            id_fuel=id_fuel,
                            id_emission=id_emission,
                            value=float(raw_ti),
                            year=year,
                            timesline=id_timeslice,
                            daytype=id_daytype,
                            season=id_season,
                            dailytimebracket=id_dtb,
                            id_storage_set=id_storage,
                        )
                        wrote_any = True

            if not wrote_any:
                for col_idx, year in year_cols:
                    if col_idx >= len(row_values):
                        continue
                    year_value = _to_float_with_param_default(row_values[col_idx], param_name)
                    _add_to_batch(
                        id_parameter=int(parameter_id),
                        id_region=int(id_region),
                        id_technology=id_technology,
                        id_fuel=id_fuel,
                        id_emission=id_emission,
                        value=year_value,
                        year=year,
                        timesline=id_timeslice,
                        daytype=id_daytype,
                        season=id_season,
                        dailytimebracket=id_dtb,
                        id_storage_set=id_storage,
                    )
                    wrote_any = True
            if not wrote_any:
                stats.skipped += 1
            continue

        group_key: _AggKey = (
            int(parameter_id),
            int(id_region),
            id_technology,
            id_fuel,
            id_emission,
            id_timeslice,
            id_daytype,
            id_season,
            id_dtb,
            id_storage,
        )
        year_values: dict[int, float] = {}
        for col_idx, year in year_cols:
            if col_idx < len(row_values):
                year_values[year] = _to_float_with_param_default(row_values[col_idx], param_name)

        time_indep_val = None
        if time_independent_idx is not None and time_independent_idx < len(row_values):
            raw_ti = _to_float(row_values[time_independent_idx])
            if raw_ti is not None:
                time_indep_val = float(raw_ti)

        row_entry = {
            "param_name": param_name,
            "id_parameter": int(parameter_id),
            "id_region": int(id_region),
            "id_technology": id_technology,
            "id_fuel": id_fuel,
            "id_emission": id_emission,
            "timesline": id_timeslice,
            "daytype": id_daytype,
            "season": id_season,
            "dailytimebracket": id_dtb,
            "id_storage_set": id_storage,
            "year_values": year_values,
            "time_indep_val": time_indep_val,
        }
        if group_key not in _agg_buffer:
            _agg_buffer[group_key] = []
        _agg_buffer[group_key].append(row_entry)

    for group_rows in _agg_buffer.values():
        _flush_agg_group(group_rows)

    _flush_batch()


def _get_or_create_name_id(
    db: Session,
    *,
    model,
    name: str | None,
    ref_map: dict[str, int],
) -> int | None:
    """Obtiene id por nombre; si no existe, crea el registro y actualiza ref_map."""
    if not name:
        return None
    current = ref_map.get(name)
    if current is not None:
        return current
    kwargs: dict[str, object] = {"name": name}
    if getattr(model, "is_active", None) is not None:
        kwargs["is_active"] = True
    obj = model(**kwargs)
    db.add(obj)
    db.flush()
    ref_map[name] = obj.id
    return obj.id


def _get_or_create_code_id(
    db: Session,
    *,
    model,
    code: str | None,
    ref_map: dict[str, int],
) -> int | None:
    if not code:
        return None
    current = ref_map.get(code)
    if current is not None:
        return current
    kwargs: dict[str, object] = {"code": code}
    if getattr(model, "description", None) is not None:
        kwargs["description"] = None
    obj = model(**kwargs)
    db.add(obj)
    db.flush()
    ref_map[code] = obj.id
    return obj.id


class OfficialImportService:
    @staticmethod
    def list_workbook_sheets(*, filename: str, content: bytes) -> list[str]:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        return list(workbook.sheetnames)

    @staticmethod
    def import_xlsm(
        db: Session,
        *,
        filename: str,
        content: bytes,
        imported_by: str,
        selected_sheet_name: str | None = None,
        scenario_id_override: int | None = None,
        use_default_scenario: bool = False,
        replace_scenario_data: bool = False,
    ) -> dict[str, object]:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        stats = ImportStats()
        if selected_sheet_name:
            selected_names = {_normalize_key(n) for n in workbook.sheetnames}
            if _normalize_key(selected_sheet_name) not in selected_names:
                raise ValueError(f"La hoja seleccionada no existe en el archivo: {selected_sheet_name}")

        name_catalog_sheets = [
            (Parameter, {"parameter", "parameters", "parametro", "parametros"}),
            (Region, {"region", "regions", "regiones"}),
            (Technology, {"technology", "technologies", "tecnologia", "tecnologias"}),
            (Fuel, {"fuel", "fuels", "combustible", "combustibles"}),
            (Emission, {"emission", "emissions", "emision", "emisiones"}),
            (Solver, {"solver", "solvers"}),
        ]
        code_catalog_sheets = [
            (Timeslice, {"timeslice", "timeslices"}),
            (ModeOfOperation, {"mode_of_operation", "modeofoperation", "mode_operation", "modos_operacion"}),
            (Season, {"season", "seasons"}),
            (Daytype, {"daytype", "daytypes"}),
            (Dailytimebracket, {"dailytimebracket", "daily_time_bracket", "dailytimebrackets"}),
            (StorageSet, {"storage_set", "storageset", "storage"}),
            (UdcSet, {"udc_set", "udc", "udcset"}),
        ]

        for model, aliases in name_catalog_sheets:
            sheet = _sheet_by_alias(workbook, aliases)
            if not sheet:
                continue
            if not _is_sheet_selected(sheet.title, selected_sheet_name):
                continue
            for row in _iter_rows(sheet):
                stats.total_rows_read += 1
                _get_or_create_name_catalog(
                    db, model=model, row=row, stats=stats, section=model.__tablename__
                )

        for model, aliases in code_catalog_sheets:
            sheet = _sheet_by_alias(workbook, aliases)
            if not sheet:
                continue
            if not _is_sheet_selected(sheet.title, selected_sheet_name):
                continue
            for row in _iter_rows(sheet):
                stats.total_rows_read += 1
                _get_or_create_code_catalog(
                    db, model=model, row=row, stats=stats, section=model.__tablename__
                )

        if scenario_id_override is None and not use_default_scenario:
            scenario_sheet = _sheet_by_alias(workbook, {"scenario", "scenarios", "escenario", "escenarios"})
            if scenario_sheet and _is_sheet_selected(scenario_sheet.title, selected_sheet_name):
                for row in _iter_rows(scenario_sheet):
                    stats.total_rows_read += 1
                    name = _clean_str(row.get("name") or row.get("nombre"))
                    owner = _clean_str(row.get("owner") or row.get("propietario")) or imported_by
                    if not name:
                        stats.skipped += 1
                        stats.warn("[scenario] fila omitida por nombre vacío.")
                        continue
                    edit_policy = _clean_str(row.get("edit_policy") or row.get("politica")) or "OWNER_ONLY"
                    if edit_policy not in {"OWNER_ONLY", "OPEN", "RESTRICTED"}:
                        stats.warn(
                            f"[scenario:{name}] edit_policy '{edit_policy}' inválido; se ajusta a OWNER_ONLY."
                        )
                        edit_policy = "OWNER_ONLY"
                    is_template = _to_bool(row.get("is_template") or row.get("plantilla"))
                    description = _clean_str(row.get("description") or row.get("descripcion"))
                    existing = db.execute(
                        select(Scenario).where(Scenario.name == name, Scenario.owner == owner)
                    ).scalar_one_or_none()
                    if existing:
                        existing.description = description
                        existing.edit_policy = edit_policy
                        existing.is_template = is_template
                        stats.updated += 1
                    else:
                        db.add(
                            Scenario(
                                name=name,
                                description=description,
                                owner=owner,
                                edit_policy=edit_policy,
                                is_template=is_template,
                            )
                        )
                        stats.inserted += 1

        db.flush()

        scenario_map = {
            f"{row.owner}:{row.name}": row.id for row in db.execute(select(Scenario)).scalars().all()
        }
        effective_scenario_id_override = scenario_id_override
        if use_default_scenario:
            default_scenario = _get_or_create_default_scenario(db, owner=imported_by)
            effective_scenario_id_override = default_scenario.id
            if replace_scenario_data:
                _reset_scenario_data(db, scenario_id=default_scenario.id)
                db.flush()

        if effective_scenario_id_override is not None:
            fallback_scenario = db.execute(
                select(Scenario).where(Scenario.id == int(effective_scenario_id_override))
            ).scalar_one_or_none()
            if fallback_scenario is None:
                raise ValueError(f"Escenario no encontrado para importación: {effective_scenario_id_override}")
        else:
            fallback_scenario = db.execute(select(Scenario).order_by(Scenario.id.asc())).scalars().first()
        should_run_notebook_preprocess = False
        param_map = _load_name_map(db, Parameter)
        region_map = _load_name_map(db, Region)
        tech_map = _load_name_map(db, Technology)
        fuel_map = _load_name_map(db, Fuel)
        emission_map = _load_name_map(db, Emission)
        solver_map = _load_name_map(db, Solver)
        timeslice_map = _load_code_map(db, Timeslice)
        mode_map = _load_code_map(db, ModeOfOperation)
        season_map = _load_code_map(db, Season)
        daytype_map = _load_code_map(db, Daytype)
        dtb_map = _load_code_map(db, Dailytimebracket)
        storage_map = _load_code_map(db, StorageSet)
        udc_map = _load_code_map(db, UdcSet)
        default_solver = db.execute(select(Solver).where(Solver.name == "default")).scalar_one_or_none()
        if default_solver is None:
            default_solver = Solver(name="default", is_active=True)
            db.add(default_solver)
            db.flush()
        default_solver_id = default_solver.id

        parameter_value_sheet = _sheet_by_alias(
            workbook, {"parameter_value", "parameter_values", "valores_parametro", "parametervalue"}
        )
        if parameter_value_sheet and _is_sheet_selected(parameter_value_sheet.title, selected_sheet_name):
            for row in _iter_rows(parameter_value_sheet):
                stats.total_rows_read += 1
                parameter_id = _resolve_ref(
                    row,
                    id_keys=["id_parameter", "parameter_id"],
                    name_keys=["parameter", "parameter_name", "parametro"],
                    ref_map=param_map,
                )
                if parameter_id is None:
                    parameter_name = _clean_str(
                        row.get("parameter") or row.get("parameter_name") or row.get("parametro")
                    )
                    parameter_id = _get_or_create_name_id(
                        db, model=Parameter, name=parameter_name, ref_map=param_map
                    )
                region_id = _resolve_ref(
                    row,
                    id_keys=["id_region", "region_id"],
                    name_keys=["region", "region_name"],
                    ref_map=region_map,
                )
                if region_id is None:
                    region_name = _clean_str(row.get("region") or row.get("region_name"))
                    region_id = _get_or_create_name_id(
                        db, model=Region, name=region_name, ref_map=region_map
                    )
                solver_id = _resolve_ref(
                    row,
                    id_keys=["id_solver", "solver_id"],
                    name_keys=["solver", "solver_name"],
                    ref_map=solver_map,
                )
                if solver_id is None:
                    solver_name = _clean_str(row.get("solver") or row.get("solver_name"))
                    solver_id = _get_or_create_name_id(
                        db, model=Solver, name=solver_name, ref_map=solver_map
                    )
                if solver_id is None:
                    solver_id = default_solver_id
                year = _to_int(row.get("year") or row.get("anio") or row.get("ano"))
                pv_param_name = _clean_str(
                    row.get("parameter") or row.get("parameter_name") or row.get("parametro")
                )
                value = _to_float_with_param_default(row.get("value") or row.get("valor"), pv_param_name)
                if None in {parameter_id, region_id, solver_id, year}:
                    stats.skipped += 1
                    stats.warn("[parameter_value] fila omitida por campos requeridos faltantes.")
                    continue
                technology_id = _resolve_ref(
                    row,
                    id_keys=["id_technology", "technology_id"],
                    name_keys=["technology", "technology_name", "tecnologia"],
                    ref_map=tech_map,
                )
                if technology_id is None:
                    technology_name = _clean_str(
                        row.get("technology") or row.get("technology_name") or row.get("tecnologia")
                    )
                    technology_id = _get_or_create_name_id(
                        db, model=Technology, name=technology_name, ref_map=tech_map
                    )
                fuel_id = _resolve_ref(
                    row,
                    id_keys=["id_fuel", "fuel_id"],
                    name_keys=["fuel", "fuel_name", "combustible"],
                    ref_map=fuel_map,
                )
                if fuel_id is None:
                    fuel_name = _clean_str(row.get("fuel") or row.get("fuel_name") or row.get("combustible"))
                    fuel_id = _get_or_create_name_id(
                        db, model=Fuel, name=fuel_name, ref_map=fuel_map
                    )
                emission_id = _resolve_ref(
                    row,
                    id_keys=["id_emission", "emission_id"],
                    name_keys=["emission", "emission_name", "emision"],
                    ref_map=emission_map,
                )
                if emission_id is None:
                    emission_name = _clean_str(
                        row.get("emission") or row.get("emission_name") or row.get("emision")
                    )
                    emission_id = _get_or_create_name_id(
                        db, model=Emission, name=emission_name, ref_map=emission_map
                    )
                unit = _clean_str(row.get("unit") or row.get("unidad"))
                mode_of_operation = _to_bool(
                    row.get("mode_of_operation") or row.get("modo_operacion")
                )
                id_timeslice = _resolve_ref(
                    row,
                    id_keys=["id_timeslice", "timeslice_id"],
                    name_keys=["timeslice", "timeslice_code"],
                    ref_map=timeslice_map,
                )
                if id_timeslice is None:
                    timeslice_code = _clean_str(row.get("timeslice") or row.get("timeslice_code"))
                    id_timeslice = _get_or_create_code_id(
                        db, model=Timeslice, code=timeslice_code, ref_map=timeslice_map
                    )

                id_season = _resolve_ref(
                    row,
                    id_keys=["id_season", "season_id"],
                    name_keys=["season", "season_code"],
                    ref_map=season_map,
                )
                if id_season is None:
                    season_code = _clean_str(row.get("season") or row.get("season_code"))
                    id_season = _get_or_create_code_id(
                        db, model=Season, code=season_code, ref_map=season_map
                    )

                id_daytype = _resolve_ref(
                    row,
                    id_keys=["id_daytype", "daytype_id"],
                    name_keys=["daytype", "daytype_code"],
                    ref_map=daytype_map,
                )
                if id_daytype is None:
                    daytype_code = _clean_str(row.get("daytype") or row.get("daytype_code"))
                    id_daytype = _get_or_create_code_id(
                        db, model=Daytype, code=daytype_code, ref_map=daytype_map
                    )

                id_dtb = _resolve_ref(
                    row,
                    id_keys=["id_dailytimebracket", "dailytimebracket_id"],
                    name_keys=["dailytimebracket", "dailytimebracket_code"],
                    ref_map=dtb_map,
                )
                if id_dtb is None:
                    dtb_code = _clean_str(row.get("dailytimebracket") or row.get("dailytimebracket_code"))
                    id_dtb = _get_or_create_code_id(
                        db, model=Dailytimebracket, code=dtb_code, ref_map=dtb_map
                    )

                id_storage = _resolve_ref(
                    row,
                    id_keys=["id_storage_set", "storage_set_id"],
                    name_keys=["storage_set", "storage_set_code", "storage"],
                    ref_map=storage_map,
                )
                if id_storage is None:
                    storage_code = _clean_str(
                        row.get("storage_set") or row.get("storage_set_code") or row.get("storage")
                    )
                    id_storage = _get_or_create_code_id(
                        db, model=StorageSet, code=storage_code, ref_map=storage_map
                    )
                existing = db.execute(
                    select(ParameterValue).where(
                        and_(
                            ParameterValue.id_parameter == parameter_id,
                            ParameterValue.id_region == region_id,
                            ParameterValue.id_technology == technology_id,
                            ParameterValue.id_fuel == fuel_id,
                            ParameterValue.id_emission == emission_id,
                            ParameterValue.id_solver == solver_id,
                            ParameterValue.year == year,
                        )
                    )
                ).scalar_one_or_none()
                if existing:
                    existing.value = float(value)
                    existing.unit = unit
                    existing.mode_of_operation = mode_of_operation
                    _upsert_parameter_storage_for_pv(
                        db,
                        parameter_value_id=int(existing.id),
                        timesline=id_timeslice,
                        daytype=id_daytype,
                        season=id_season,
                        dailytimebracket=id_dtb,
                        id_storage_set=id_storage,
                    )
                    stats.updated += 1
                else:
                    pv = ParameterValue(
                        id_parameter=int(parameter_id),
                        id_region=int(region_id),
                        id_technology=technology_id,
                        id_fuel=fuel_id,
                        id_emission=emission_id,
                        id_solver=int(solver_id),
                        mode_of_operation=mode_of_operation,
                        year=int(year),
                        value=float(value),
                        unit=unit,
                    )
                    db.add(pv)
                    db.flush()
                    _upsert_parameter_storage_for_pv(
                        db,
                        parameter_value_id=int(pv.id),
                        timesline=id_timeslice,
                        daytype=id_daytype,
                        season=id_season,
                        dailytimebracket=id_dtb,
                        id_storage_set=id_storage,
                    )
                    stats.inserted += 1

        osemosys_param_sheet = _sheet_by_alias(
            workbook,
            {
                "osemosys_param_value",
                "osemosys_param_values",
                "osemosys",
                "parametros_osemosys",
            },
        )
        if osemosys_param_sheet and _is_sheet_selected(osemosys_param_sheet.title, selected_sheet_name):
            should_run_notebook_preprocess = True
            for row in _iter_rows(osemosys_param_sheet):
                stats.total_rows_read += 1
                scenario_id = (
                    int(effective_scenario_id_override)
                    if effective_scenario_id_override is not None
                    else _resolve_ref(
                        row,
                        id_keys=["id_scenario", "scenario_id"],
                        name_keys=["scenario", "scenario_name", "escenario"],
                        ref_map={k.split(":", maxsplit=1)[1]: v for k, v in scenario_map.items()},
                    )
                )
                if scenario_id is None and fallback_scenario is not None:
                    scenario_id = fallback_scenario.id
                    stats.warn(
                        "[osemosys_param_value] escenario no informado; se usa el primer escenario existente."
                    )
                param_name = _clean_str(row.get("param_name") or row.get("parameter") or row.get("parametro"))
                value = _to_float_with_param_default(row.get("value") or row.get("valor"), param_name)
                year = _to_int(row.get("year") or row.get("anio") or row.get("ano"))
                if scenario_id is None or not param_name:
                    stats.skipped += 1
                    stats.warn("[osemosys_param_value] fila omitida por campos requeridos faltantes.")
                    continue
                id_region = _resolve_ref(
                    row,
                    id_keys=["id_region", "region_id"],
                    name_keys=["region", "region_name"],
                    ref_map=region_map,
                )
                if id_region is None:
                    region_name = _clean_str(row.get("region") or row.get("region_name"))
                    id_region = _get_or_create_name_id(
                        db, model=Region, name=region_name, ref_map=region_map
                    )
                id_technology = _resolve_ref(
                    row,
                    id_keys=["id_technology", "technology_id"],
                    name_keys=["technology", "technology_name", "tecnologia"],
                    ref_map=tech_map,
                )
                if id_technology is None:
                    technology_name = _clean_str(
                        row.get("technology") or row.get("technology_name") or row.get("tecnologia")
                    )
                    id_technology = _get_or_create_name_id(
                        db, model=Technology, name=technology_name, ref_map=tech_map
                    )
                id_fuel = _resolve_ref(
                    row,
                    id_keys=["id_fuel", "fuel_id"],
                    name_keys=["fuel", "fuel_name", "combustible"],
                    ref_map=fuel_map,
                )
                if id_fuel is None:
                    fuel_name = _clean_str(row.get("fuel") or row.get("fuel_name") or row.get("combustible"))
                    id_fuel = _get_or_create_name_id(
                        db, model=Fuel, name=fuel_name, ref_map=fuel_map
                    )
                id_emission = _resolve_ref(
                    row,
                    id_keys=["id_emission", "emission_id"],
                    name_keys=["emission", "emission_name", "emision"],
                    ref_map=emission_map,
                )
                if id_emission is None:
                    emission_name = _clean_str(
                        row.get("emission") or row.get("emission_name") or row.get("emision")
                    )
                    id_emission = _get_or_create_name_id(
                        db, model=Emission, name=emission_name, ref_map=emission_map
                    )
                id_timeslice = _resolve_ref(
                    row,
                    id_keys=["id_timeslice", "timeslice_id"],
                    name_keys=["timeslice", "timeslice_code"],
                    ref_map=timeslice_map,
                )
                if id_timeslice is None:
                    timeslice_code = _clean_str(row.get("timeslice") or row.get("timeslice_code"))
                    id_timeslice = _get_or_create_code_id(
                        db, model=Timeslice, code=timeslice_code, ref_map=timeslice_map
                    )
                id_mode = _resolve_ref(
                    row,
                    id_keys=["id_mode_of_operation", "mode_of_operation_id"],
                    name_keys=["mode_of_operation", "mode_of_operation_code"],
                    ref_map=mode_map,
                )
                if id_mode is None:
                    mode_code = _clean_str(
                        row.get("mode_of_operation") or row.get("mode_of_operation_code")
                    )
                    id_mode = _get_or_create_code_id(
                        db, model=ModeOfOperation, code=mode_code, ref_map=mode_map
                    )
                id_season = _resolve_ref(
                    row,
                    id_keys=["id_season", "season_id"],
                    name_keys=["season", "season_code"],
                    ref_map=season_map,
                )
                id_daytype = _resolve_ref(
                    row,
                    id_keys=["id_daytype", "daytype_id"],
                    name_keys=["daytype", "daytype_code"],
                    ref_map=daytype_map,
                )
                id_dtb = _resolve_ref(
                    row,
                    id_keys=["id_dailytimebracket", "dailytimebracket_id"],
                    name_keys=["dailytimebracket", "dailytimebracket_code"],
                    ref_map=dtb_map,
                )
                id_storage = _resolve_ref(
                    row,
                    id_keys=["id_storage_set", "storage_set_id"],
                    name_keys=["storage_set", "storage_set_code"],
                    ref_map=storage_map,
                )
                if id_storage is None:
                    storage_code = _clean_str(row.get("storage_set") or row.get("storage_set_code"))
                    id_storage = _get_or_create_code_id(
                        db, model=StorageSet, code=storage_code, ref_map=storage_map
                    )
                id_udc = _resolve_ref(
                    row,
                    id_keys=["id_udc_set", "udc_set_id"],
                    name_keys=["udc_set", "udc_set_code"],
                    ref_map=udc_map,
                )
                existing = db.execute(
                    select(OsemosysParamValue).where(
                        and_(
                            OsemosysParamValue.id_scenario == int(scenario_id),
                            OsemosysParamValue.param_name == param_name,
                            OsemosysParamValue.id_region.is_(None)
                            if id_region is None
                            else OsemosysParamValue.id_region == id_region,
                            OsemosysParamValue.id_technology.is_(None)
                            if id_technology is None
                            else OsemosysParamValue.id_technology == id_technology,
                            OsemosysParamValue.id_fuel.is_(None)
                            if id_fuel is None
                            else OsemosysParamValue.id_fuel == id_fuel,
                            OsemosysParamValue.id_emission.is_(None)
                            if id_emission is None
                            else OsemosysParamValue.id_emission == id_emission,
                            OsemosysParamValue.id_timeslice.is_(None)
                            if id_timeslice is None
                            else OsemosysParamValue.id_timeslice == id_timeslice,
                            OsemosysParamValue.id_mode_of_operation.is_(None)
                            if id_mode is None
                            else OsemosysParamValue.id_mode_of_operation == id_mode,
                            OsemosysParamValue.id_season.is_(None)
                            if id_season is None
                            else OsemosysParamValue.id_season == id_season,
                            OsemosysParamValue.id_daytype.is_(None)
                            if id_daytype is None
                            else OsemosysParamValue.id_daytype == id_daytype,
                            OsemosysParamValue.id_dailytimebracket.is_(None)
                            if id_dtb is None
                            else OsemosysParamValue.id_dailytimebracket == id_dtb,
                            OsemosysParamValue.id_storage_set.is_(None)
                            if id_storage is None
                            else OsemosysParamValue.id_storage_set == id_storage,
                            OsemosysParamValue.id_udc_set.is_(None)
                            if id_udc is None
                            else OsemosysParamValue.id_udc_set == id_udc,
                            OsemosysParamValue.year.is_(None)
                            if year is None
                            else OsemosysParamValue.year == year,
                        )
                    )
                ).scalar_one_or_none()
                if existing:
                    existing.value = float(value)
                    stats.updated += 1
                else:
                    db.add(
                        OsemosysParamValue(
                            id_scenario=int(scenario_id),
                            param_name=param_name,
                            id_region=id_region,
                            id_technology=id_technology,
                            id_fuel=id_fuel,
                            id_emission=id_emission,
                            id_timeslice=id_timeslice,
                            id_mode_of_operation=id_mode,
                            id_season=id_season,
                            id_daytype=id_daytype,
                            id_dailytimebracket=id_dtb,
                            id_storage_set=id_storage,
                            id_udc_set=id_udc,
                            year=year,
                            value=float(value),
                        )
                    )
                    stats.inserted += 1

        # Formato SAND (hoja "Parameters"): matriz por año.
        processed_sand_sheets: set[str] = set()
        sand_parameters_sheet = _sheet_by_alias(workbook, {"parameters"})
        if sand_parameters_sheet and _is_sheet_selected(sand_parameters_sheet.title, selected_sheet_name):
            if use_default_scenario:
                _import_sand_matrix_sheet_to_parameter_value(
                    db,
                    sheet=sand_parameters_sheet,
                    stats=stats,
                    param_map=param_map,
                    region_map=region_map,
                    tech_map=tech_map,
                    fuel_map=fuel_map,
                    emission_map=emission_map,
                    timeslice_map=timeslice_map,
                    season_map=season_map,
                    daytype_map=daytype_map,
                    dtb_map=dtb_map,
                    storage_map=storage_map,
                    default_solver_id=default_solver_id,
                )
            else:
                should_run_notebook_preprocess = True
                fallback_scenario = _import_sand_matrix_sheet(
                    db,
                    sheet=sand_parameters_sheet,
                    stats=stats,
                    imported_by=imported_by,
                    fallback_scenario=fallback_scenario,
                    region_map=region_map,
                    tech_map=tech_map,
                    fuel_map=fuel_map,
                    emission_map=emission_map,
                    timeslice_map=timeslice_map,
                    mode_map=mode_map,
                    storage_map=storage_map,
                    season_map=season_map,
                    daytype_map=daytype_map,
                    dtb_map=dtb_map,
                    udc_map=udc_map,
                )
            processed_sand_sheets.add(_normalize_key(sand_parameters_sheet.title))

        if selected_sheet_name:
            selected_sheet = None
            for sheet_name in workbook.sheetnames:
                if _normalize_key(sheet_name) == _normalize_key(selected_sheet_name):
                    selected_sheet = workbook[sheet_name]
                    break
            if selected_sheet and _normalize_key(selected_sheet.title) not in processed_sand_sheets:
                if use_default_scenario:
                    _import_sand_matrix_sheet_to_parameter_value(
                        db,
                        sheet=selected_sheet,
                        stats=stats,
                        param_map=param_map,
                        region_map=region_map,
                        tech_map=tech_map,
                        fuel_map=fuel_map,
                        emission_map=emission_map,
                        timeslice_map=timeslice_map,
                        season_map=season_map,
                        daytype_map=daytype_map,
                        dtb_map=dtb_map,
                        storage_map=storage_map,
                        default_solver_id=default_solver_id,
                    )
                else:
                    should_run_notebook_preprocess = True
                    fallback_scenario = _import_sand_matrix_sheet(
                        db,
                        sheet=selected_sheet,
                        stats=stats,
                        imported_by=imported_by,
                        fallback_scenario=fallback_scenario,
                        region_map=region_map,
                        tech_map=tech_map,
                        fuel_map=fuel_map,
                        emission_map=emission_map,
                        timeslice_map=timeslice_map,
                        mode_map=mode_map,
                        storage_map=storage_map,
                        season_map=season_map,
                        daytype_map=daytype_map,
                        dtb_map=dtb_map,
                        udc_map=udc_map,
                    )

        # Preprocesamiento tipo notebook (sets, matrices, emisiones a la entrada) al subir datos.
        notebook_preprocess: dict[str, int] | None = None
        notebook_preprocess_error: str | None = None
        if should_run_notebook_preprocess and fallback_scenario is not None:
            try:
                notebook_preprocess = run_notebook_preprocess(
                    db,
                    int(fallback_scenario.id),
                    filter_by_sets=True,
                    complete_matrices=False,
                    emission_ratios_at_input=False,
                    generate_udc_matrices=False,
                )
            except Exception as e:
                # No fallar la importación; la simulación usará datos crudos.
                notebook_preprocess_error = str(e)

        db.commit()
        return {
            "filename": filename,
            "imported_at": datetime.now(timezone.utc).isoformat(),
            "imported_by": imported_by,
            "total_rows_read": stats.total_rows_read,
            "inserted": stats.inserted,
            "updated": stats.updated,
            "skipped": stats.skipped,
            "warnings": stats.warnings,
            "notebook_preprocess": notebook_preprocess,
            "notebook_preprocess_error": notebook_preprocess_error,
        }

    @staticmethod
    def update_scenario_from_excel(
        db: Session,
        *,
        scenario_id: int,
        filename: str,
        content: bytes,
        selected_sheet_name: str,
    ) -> dict:
        """Actualiza valores existentes de un escenario desde un Excel SAND.

        Solo actualiza registros que ya existen (por clave compuesta).
        Si un registro del Excel no se encuentra en el escenario, se registra
        como advertencia sin insertar.

        La operación es transaccional: si algo falla, se hace rollback completo.
        """
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        stats = ImportStats()

        selected_names = {_normalize_key(n) for n in workbook.sheetnames}
        if _normalize_key(selected_sheet_name) not in selected_names:
            raise ValueError(f"La hoja seleccionada no existe en el archivo: {selected_sheet_name}")

        region_map = _load_name_map(db, Region)
        tech_map = _load_name_map(db, Technology)
        fuel_map = _load_name_map(db, Fuel)
        emission_map = _load_name_map(db, Emission)
        timeslice_map = _load_code_map(db, Timeslice)
        mode_map = _load_code_map(db, ModeOfOperation)
        season_map = _load_code_map(db, Season)
        daytype_map = _load_code_map(db, Daytype)
        dtb_map = _load_code_map(db, Dailytimebracket)
        storage_map = _load_code_map(db, StorageSet)
        udc_map = _load_code_map(db, UdcSet)

        target_sheet = None
        for sn in workbook.sheetnames:
            if _normalize_key(sn) == _normalize_key(selected_sheet_name):
                target_sheet = workbook[sn]
                break

        if target_sheet is None:
            raise ValueError(f"No se encontró la hoja: {selected_sheet_name}")

        result = _update_sand_matrix_sheet(
            db,
            sheet=target_sheet,
            stats=stats,
            scenario_id=scenario_id,
            region_map=region_map,
            tech_map=tech_map,
            fuel_map=fuel_map,
            emission_map=emission_map,
            timeslice_map=timeslice_map,
            mode_map=mode_map,
            storage_map=storage_map,
            season_map=season_map,
            daytype_map=daytype_map,
            dtb_map=dtb_map,
            udc_map=udc_map,
        )

        db.commit()
        return {
            "updated": result["updated"],
            "not_found": result["not_found"],
            "total_rows_read": stats.total_rows_read,
            "warnings": stats.warnings or [],
        }

    @staticmethod
    def preview_scenario_from_excel(
        db: Session,
        *,
        scenario_id: int,
        filename: str,
        content: bytes,
        selected_sheet_name: str,
    ) -> dict:
        """Genera preview de cambios desde un Excel SAND sin modificar datos.

        Retorna la lista de diferencias (valor actual vs nuevo) para que
        el usuario confirme antes de aplicar.
        """
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        stats = ImportStats()

        selected_names = {_normalize_key(n) for n in workbook.sheetnames}
        if _normalize_key(selected_sheet_name) not in selected_names:
            raise ValueError(f"La hoja seleccionada no existe en el archivo: {selected_sheet_name}")

        region_map = _load_name_map(db, Region)
        tech_map = _load_name_map(db, Technology)
        fuel_map = _load_name_map(db, Fuel)
        emission_map = _load_name_map(db, Emission)
        timeslice_map = _load_code_map(db, Timeslice)
        mode_map = _load_code_map(db, ModeOfOperation)
        season_map = _load_code_map(db, Season)
        daytype_map = _load_code_map(db, Daytype)
        dtb_map = _load_code_map(db, Dailytimebracket)
        storage_map = _load_code_map(db, StorageSet)
        udc_map = _load_code_map(db, UdcSet)

        target_sheet = None
        for sn in workbook.sheetnames:
            if _normalize_key(sn) == _normalize_key(selected_sheet_name):
                target_sheet = workbook[sn]
                break

        if target_sheet is None:
            raise ValueError(f"No se encontró la hoja: {selected_sheet_name}")

        result = _preview_sand_matrix_sheet(
            db,
            sheet=target_sheet,
            stats=stats,
            scenario_id=scenario_id,
            region_map=region_map,
            tech_map=tech_map,
            fuel_map=fuel_map,
            emission_map=emission_map,
            timeslice_map=timeslice_map,
            mode_map=mode_map,
            storage_map=storage_map,
            season_map=season_map,
            daytype_map=daytype_map,
            dtb_map=dtb_map,
            udc_map=udc_map,
        )

        return {
            "changes": result["changes"],
            "not_found": result["not_found"],
            "total_rows_read": stats.total_rows_read,
            "warnings": stats.warnings or [],
        }

    @staticmethod
    def apply_excel_changes(
        db: Session,
        *,
        scenario_id: int,
        changes: list[dict],
    ) -> dict:
        """Aplica cambios confirmados por el usuario tras preview.

        Cada entry en ``changes`` tiene ``row_id`` y ``new_value``.
        Solo actualiza filas que pertenecen al escenario indicado.
        """
        if not changes:
            return {"updated": 0, "skipped": 0}

        valid_ids = set(
            row[0]
            for row in db.execute(
                select(OsemosysParamValue.id).where(
                    OsemosysParamValue.id_scenario == scenario_id
                )
            ).all()
        )

        to_update: list[dict] = []
        skipped = 0
        for ch in changes:
            rid = ch["row_id"]
            if rid in valid_ids:
                to_update.append({"_id": rid, "value": float(ch["new_value"])})
            else:
                skipped += 1

        if to_update:
            tbl = OsemosysParamValue.__table__
            stmt = (
                sa_update(tbl)
                .where(tbl.c.id == bindparam("_id"))
                .values(value=bindparam("value"))
            )
            for i in range(0, len(to_update), BATCH_SIZE):
                db.execute(stmt, to_update[i : i + BATCH_SIZE])

        db.commit()
        return {"updated": len(to_update), "skipped": skipped}
